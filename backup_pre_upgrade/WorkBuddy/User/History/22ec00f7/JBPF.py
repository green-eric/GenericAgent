#!/usr/bin/env python3
"""
验证Excel和JSON输出字段的逻辑一致性
"""

import json
import openpyxl
import os

def verify_json_fields():
    """验证JSON输出字段"""
    print("=== JSON 字段验证 ===")
    with open('季报年报交叉验证_20260426_092933.json', 'r', encoding='utf-8') as f:
        data = json.load(f)

    required_fields = [
        'data_timestamp', 'total_stocks', 'annual_a_count',
        'quarterly_a_count', 'preferred_count', 'preferred_stocks',
        'annual_only_a', 'quarterly_only_a'
    ]

    print("JSON主字段:", [field for field in required_fields if field in data])

    # 验证preferred_stocks结构
    if data['preferred_stocks']:
        sample = data['preferred_stocks'][0]
        expected_keys = ['ts_code', 'name', 'industry_l1', 'annual_score', 'annual_grade', 'quarterly_score', 'quarterly_grade']
        actual_keys = list(sample.keys())
        print("preferred_stocks字段:", actual_keys)
        missing = set(expected_keys) - set(actual_keys)
        if missing:
            print(f"缺少字段: {missing}")
        else:
            print("preferred_stocks字段完整")

    # 验证annual_only_a结构
    if data['annual_only_a']:
        sample = data['annual_only_a'][0]
        expected_keys = ['ts_code', 'name', 'score', 'grade']
        actual_keys = list(sample.keys())
        print("annual_only_a字段:", actual_keys)
        missing = set(expected_keys) - set(actual_keys)
        if missing:
            print(f"缺少字段: {missing}")
        else:
            print("annual_only_a字段完整")

    print(f"股票总数: {data['total_stocks']}")
    print(f"年报A级: {data['annual_a_count']} 只")
    print(f"季报A级: {data['quarterly_a_count']} 只")
    print(f"交集优选: {data['preferred_count']} 只\n")

def verify_excel_fields():
    """验证Excel输出字段"""
    print("=== Excel 字段验证 ===")

    try:
        wb = openpyxl.load_workbook('季度评分_20260426_092933.xlsx')
        sheets = wb.sheetnames
        print(f"✓ 工作表: {sheets}")

        # 验证全部评分表
        ws_all = wb['全部评分']
        headers = [cell.value for cell in ws_all[1]]
        print(f"✓ 全部评分表头: {headers}")

        expected_headers = [
            '代码', '名称', '行业', '年报总分', '年报等级',
            '年报盈利', '年报成长', '年报现金流', '年报负债',
            '季报总分', '季报等级', '季报盈利', '季报成长',
            '季报现金流', '季报负债', '季报毛利率', '季报净利率',
            '季报营收同比', '季报利润同比', '数据来源', '置信度',
            '完整度', '季报日期', '交集优选'
        ]

        missing_headers = set(expected_headers) - set(headers)
        extra_headers = set(headers) - set(expected_headers)

        if missing_headers:
            print(f"缺少列: {missing_headers}")
        if extra_headers:
            print(f"额外列: {extra_headers}")
        if not missing_headers and not extra_headers:
            print("全部评分表字段完整")

        # 验证年报A∩季报A表
        ws_pref = wb['年报A∩季报A']
        pref_headers = [cell.value for cell in ws_pref[1]]
        print(f"✓ 优选股表头: {pref_headers}")

        expected_pref_headers = [
            '代码', '名称', '行业', '年报总分', '年报等级',
            '季报总分', '季报等级', '季报毛利率', '季报净利率',
            '季报营收同比', '季报利润同比', '年报ROE',
            '年报负债率', '年报现金流', '季报日期', '置信度'
        ]

        missing_pref = set(expected_pref_headers) - set(pref_headers)
        extra_pref = set(pref_headers) - set(expected_pref_headers)

        if missing_pref:
            print(f"优选股表缺少列: {missing_pref}")
        if extra_pref:
            print(f"优选股表额外列: {extra_pref}")

        # 验证统计摘要表
        ws_stats = wb['统计摘要']
        print("统计摘要表存在")

        wb.close()
        print("✓ Excel文件验证完成\n")

    except Exception as e:
        print(f"✗ Excel验证失败: {e}\n")

def verify_field_consistency():
    """验证JSON和Excel字段一致性"""
    print("=== 字段一致性验证 ===")

    # 从JSON读取数据
    with open('季报年报交叉验证_20260426_092933.json', 'r', encoding='utf-8') as f:
        json_data = json.load(f)

    # 从Excel读取数据
    wb = openpyxl.load_workbook('季度评分_20260426_092933.xlsx')
    ws_all = wb['全部评分']

    # 获取Excel中的股票代码列表
    excel_codes = []
    for row in ws_all.iter_rows(min_row=2, max_col=1):
        if row[0].value:
            excel_codes.append(row[0].value)

    # 获取JSON中的股票代码列表
    json_preferred_codes = [stock['ts_code'] for stock in json_data['preferred_stocks']]
    json_annual_a_codes = [stock['ts_code'] for stock in json_data['annual_only_a']]

    print(f"Excel总股票数: {len(excel_codes)}")
    print(f"JSON年报A级数: {len(json_annual_a_codes)}")
    print(f"JSON交集优选数: {len(json_preferred_codes)}")

    # 验证交集优选标记
    preferred_marked = 0
    for row in ws_all.iter_rows(min_row=2, max_col=25):
        if len(row) > 24 and row[24].value == '★':  # 第25列是交集优选标记
            preferred_marked += 1

    print(f"Excel中标记为优选的股票数: {preferred_marked}")
    print(f"JSON中优选股票数: {len(json_preferred_codes)}")

    if preferred_marked == len(json_preferred_codes):
        print("优选标记一致")
    else:
        print(f"优选标记不一致: Excel={preferred_marked}, JSON={len(json_preferred_codes)}")

    wb.close()

if __name__ == "__main__":
    verify_json_fields()
    verify_excel_fields()
    verify_field_consistency()
    print("验证完成")