#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用东方财富 API 验证 Excel 数据
"""
import requests
import json
import openpyxl
import re

# 读取Excel
wb = openpyxl.load_workbook(r'D:\Project\AnnualScorer\股票业绩评价_20260426_204545.xlsx')
ws = wb.active
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

targets = {"603350.SH": "安乃达", "000852.SZ": "石化机械"}
excel_data = {}
for r in range(2, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if code in targets:
        row = {}
        for i, h in enumerate(headers, 1):
            row[h] = ws.cell(r, i).value
        excel_data[code] = row

def eastmoney_finance(code):
    """东方财富财务数据"""
    # 转换代码格式
    if code.endswith(".SH"):
        secid = f"1.{code[:-3]}"
    elif code.endswith(".SZ"):
        secid = f"0.{code[:-3]}"
    else:
        return None
    
    # 主要财务指标
    url = f"https://push2his.eastmoney.com/api/qt/stock/kline/get"
    params = {
        "secid": secid,
        "fields1": "f1,f2,f3,f4,f5,f6",
        "fields2": "f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61",
        "klt": "101",  # 年线
        "end": "20251231",
        "lmt": "5",
    }
    
    # 尝试利润表API
    url2 = f"https://push2.eastmoney.com/api/qt/stock/fflow/kline/get"
    params2 = {
        "secid": secid,
        "fields1": "f1,f2,f3,f4,f5,f6",
        "fields2": "f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61",
        "klt": "101",
        "end": "20251231",
        "lmt": "5",
    }
    
    # 财务分析API
    url3 = f"https://push2.eastmoney.com/api/qt/stock/get"
    params3 = {
        "secid": secid,
        "fields": "f170,f162,f167,f168,f169,f111,f112,f113,f114,f115,f116,f117,f118,f119,f120,f121,f122,f123,f124,f125,f126,f127,f128,f129,f130,f131,f132,f133,f134,f135,f136,f137,f138,f139,f140,f141,f142,f143,f144,f145,f146,f147,f148,f149,f150,f151,f152,f153,f154,f155,f156,f157,f158,f159,f160,f161,f162,f163,f164,f165,f166,f167,f168,f169,f170,f171,f172,f173,f174,f175,f176,f177,f178,f179,f180,f181,f182,f183,f184,f185,f186,f187,f188,f189,f190,f191,f192,f193,f194,f195,f196,f197,f198,f199,f200",
    }
    
    try:
        resp = requests.get(url3, params=params3, timeout=10)
        data = resp.json()
        return data.get("data", {})
    except Exception as e:
        print(f"  API错误: {e}")
        return None

for ts_code, name in targets.items():
    row = excel_data.get(ts_code, {})
    
    print(f"\n{'='*60}")
    print(f"股票: {ts_code} {name}")
    print(f"{'='*60}")
    
    print(f"\n  [Excel数据]")
    print(f"    ROE: {row.get('ROE(%)', 'N/A')}%")
    print(f"    毛利率: {row.get('毛利率(%)', 'N/A')}%")
    print(f"    净利率: {row.get('净利率(%)', 'N/A')}%")
    print(f"    营收同比: {row.get('营收同比(%)', 'N/A')}%")
    print(f"    净利润同比: {row.get('净利润同比(%)', 'N/A')}%")
    print(f"    资产负债率: {row.get('资产负债率(%)', 'N/A')}%")
    print(f"    归母净利润: {row.get('净利润(元)', 'N/A')}")
    print(f"    经营现金流: {row.get('经营现金流(元)', 'N/A')}")
    print(f"    OCF/净利润: {row.get('OCF/净利润(%)', 'N/A')}%")
    print(f"    总分: {row.get('总分', 'N/A')}  评级: {row.get('评级', 'N/A')}")
    
    print(f"\n  [东方财富 API]")
    data = eastmoney_finance(ts_code)
    if data:
        # 解析关键字段
        # f170=PE, f162=PB, f167=PS, f168=PCF
        # f111=ROE, f112=毛利率, f113=净利率
        field_map = {
            "f111": "ROE(%)",
            "f112": "毛利率(%)",
            "f113": "净利率(%)",
            "f162": "PB",
            "f170": "PE",
        }
        for field, label in field_map.items():
            val = data.get(field)
            if val:
                print(f"    {label}: {val}")
        
        # 打印所有非空字段
        print(f"\n  [全部字段]")
        for k, v in data.items():
            if v and str(v).strip() and str(v) != "None":
                print(f"    {k}: {v}")
    else:
        print("  查询失败")

print(f"\n{'='*60}")
print("验证完成")
