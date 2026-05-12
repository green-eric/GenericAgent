#!/usr/bin/env python3
"""Read 300308.SZ data from Excel"""
import os
os.environ['PYTHONIOENCODING'] = 'utf-8'

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run(['pip', 'install', 'openpyxl'], capture_output=True)
    import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\green\WorkBuddy\20260424203734\workplace\股票业绩评价_20260425_174247.xlsx', data_only=True)
ws = wb.active

print(f"Sheet: {wb.sheetnames}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Print header row
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    headers.append(str(val))
print("Headers:", " | ".join(headers))
print()

# Find 300308.SZ
for row in range(2, ws.max_row + 1):
    code = ws.cell(row=row, column=1).value
    if code and '300308' in str(code):
        print(f"Found at row {row}:")
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            print(f"  {headers[col-1]}: {val}")
        break
