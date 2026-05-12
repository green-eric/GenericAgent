"""手动验算鼎泰高科(301377)和金海通(603061)的评分"""

# ========== 配置参数 ==========
PROFIT_W = 0.35
GROWTH_W = 0.30
CFSAFE_W = 0.35

ROE_SUB_W = 0.40
GROSS_SUB_W = 0.30
NET_SUB_W = 0.30

REV_YOY_SUB_W = 0.40
PROF_YOY_SUB_W = 0.60

OCF_SUB_W = 0.40
DEBT_SUB_W = 0.60

MARKET_FALLBACK_DISC = 0.95  # 市场备选折扣

# ========== 两只股票的数据 ==========
# 鼎泰高科 301377.SZ
# ROE=9.38, 毛利率=53.25, 净利率=32.01, 营收同比=None, 利润同比=259
# OCF/利润=120.19, 负债率=29.57

# 金海通 603061.SH
# ROE=6.01, 毛利率=52.96, 净利率=29.06, 营收同比=None, 利润同比=221.54
# OCF/利润=75.16, 负债率=17.66

# ========== 从 Excel 读取所有股票数据 ==========
import openpyxl

wb = openpyxl.load_workbook(r'd:\Project\QAScorer\综合评分_20260426_202924.xlsx')
ws = wb.active

all_stocks = []
headers = [cell.value for cell in ws[1]]
print("Headers:", headers)

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        break
    stock = {
        'ts_code': row[0],
        'name': row[1],
        'industry': row[2],
        'total_score': row[3],
        'grade': row[4],
        'profit_score': row[5],
        'growth_score': row[6],
        'cfsafe_score': row[7],
        'roe': row[8],
        'gross_margin': row[9],
        'net_margin': row[10],
        'revenue_yoy': row[11],
        'profit_yoy': row[12],
        'ocf_to_profit': row[13],
        'debt_ratio': row[14],
    }
    all_stocks.append(stock)

print(f"\n共 {len(all_stocks)} 只股票")

# ========== 找出目标股票 ==========
target_codes = ['301377.SZ', '603061.SH']
targets = {s['ts_code']: s for s in all_stocks if s['ts_code'] in target_codes}

for code, s in targets.items():
    print(f"\n{'='*60}")
    print(f"{s['name']} ({s['ts_code']})")
    print(f"  ROE={s['roe']}, 毛利率={s['gross_margin']}, 净利率={s['net_margin']}")
    print(f"  营收同比={s['revenue_yoy']}, 利润同比={s['profit_yoy']}")
    print(f"  OCF/利润={s['ocf_to_profit']}, 负债率={s['debt_ratio']}")
    print(f"  总分={s['total_score']}, 等级={s['grade']}")
    print(f"  盈利={s['profit_score']}, 成长={s['growth_score']}, 现金流={s['cfsafe_score']}")

# ========== 手动计算百分位排名 ==========
def percentile_rank(value, values, reverse=False):
    if not values:
        return 50.0
    n = len(values)
    if n == 1:
        return 50.0
    if max(values) == min(values):
        return 50.0
    sorted_vals = sorted(values, reverse=reverse)
    lo, hi = 0, n - 1
    while lo <= hi:
        mid = (lo + hi) // 2
        if (not reverse and sorted_vals[mid] <= value) or (reverse and sorted_vals[mid] >= value):
            lo = mid + 1
        else:
            hi = mid - 1
    pct = (lo / (n - 1)) * 100.0
    return min(pct, 100.0)

# 构建评分池（全市场，因为行业都是空/None，会触发市场备选）
pool = all_stocks  # 全市场

def pv(key):
    return [s[key] for s in pool if s.get(key) is not None]

print(f"\n{'='*60}")
print("评分池统计（全市场）:")
print(f"  ROE: {len(pv('roe'))} 个值, 范围 [{min(pv('roe')):.2f}, {max(pv('roe')):.2f}]")
print(f"  毛利率: {len(pv('gross_margin'))} 个值, 范围 [{min(pv('gross_margin')):.2f}, {max(pv('gross_margin')):.2f}]")
print(f"  净利率: {len(pv('net_margin'))} 个值, 范围 [{min(pv('net_margin')):.2f}, {max(pv('net_margin')):.2f}]")
print(f"  营收同比: {len(pv('revenue_yoy'))} 个值, 范围 [{min(pv('revenue_yoy')):.2f}, {max(pv('revenue_yoy')):.2f}]" if pv('revenue_yoy') else "  营收同比: 0 个值")
print(f"  利润同比: {len(pv('profit_yoy'))} 个值, 范围 [{min(pv('profit_yoy')):.2f}, {max(pv('profit_yoy')):.2f}]")
print(f"  OCF/利润: {len(pv('ocf_to_profit'))} 个值, 范围 [{min(pv('ocf_to_profit')):.2f}, {max(pv('ocf_to_profit')):.2f}]")
print(f"  负债率: {len(pv('debt_ratio'))} 个值, 范围 [{min(pv('debt_ratio')):.2f}, {max(pv('debt_ratio')):.2f}]")

# 手动计算每只股票
for code in target_codes:
    s = targets[code]
    print(f"\n{'='*60}")
    print(f"手动验算: {s['name']} ({code})")
    
    roe = s['roe']
    gross = s['gross_margin']
    net = s['net_margin']
    rev_yoy = s['revenue_yoy']
    prof_yoy = s['profit_yoy']
    ocf = s['ocf_to_profit']
    debt = s['debt_ratio']
    
    # 盈利能力
    roe_s = 0.0 if roe is None else (0.0 if roe < 0 else percentile_rank(roe, pv('roe')))
    gross_s = percentile_rank(gross, pv('gross_margin')) if gross is not None else 0.0
    net_s = percentile_rank(net, pv('net_margin')) if net is not None else 0.0
    profit_score = (roe_s * ROE_SUB_W + gross_s * GROSS_SUB_W + net_s * NET_SUB_W) * MARKET_FALLBACK_DISC
    
    print(f"\n  盈利能力:")
    print(f"    ROE={roe} → percentile={roe_s:.2f} (权重{ROE_SUB_W})")
    print(f"    毛利率={gross} → percentile={gross_s:.2f} (权重{GROSS_SUB_W})")
    print(f"    净利率={net} → percentile={net_s:.2f} (权重{NET_SUB_W})")
    print(f"    profit_score = ({roe_s:.2f}×{ROE_SUB_W} + {gross_s:.2f}×{GROSS_SUB_W} + {net_s:.2f}×{NET_SUB_W}) × {MARKET_FALLBACK_DISC} = {profit_score:.2f}")
    
    # 成长能力
    rev_s = percentile_rank(rev_yoy, pv('revenue_yoy')) if rev_yoy is not None else 0.0
    prof_s = percentile_rank(prof_yoy, pv('profit_yoy')) if prof_yoy is not None else 0.0
    growth_score = (rev_s * REV_YOY_SUB_W + prof_s * PROF_YOY_SUB_W) * MARKET_FALLBACK_DISC
    
    print(f"\n  成长能力:")
    print(f"    营收同比={rev_yoy} → percentile={rev_s:.2f} (权重{REV_YOY_SUB_W})")
    print(f"    利润同比={prof_yoy} → percentile={prof_s:.2f} (权重{PROF_YOY_SUB_W})")
    print(f"    growth_score = ({rev_s:.2f}×{REV_YOY_SUB_W} + {prof_s:.2f}×{PROF_YOY_SUB_W}) × {MARKET_FALLBACK_DISC} = {growth_score:.2f}")
    
    # 现金流安全
    ocf_s = percentile_rank(ocf, pv('ocf_to_profit')) if ocf is not None else 0.0
    debt_s = percentile_rank(debt, pv('debt_ratio'), reverse=True) if debt is not None else 0.0
    cfsafe_score = (ocf_s * OCF_SUB_W + debt_s * DEBT_SUB_W) * MARKET_FALLBACK_DISC
    
    print(f"\n  现金流安全:")
    print(f"    OCF/利润={ocf} → percentile={ocf_s:.2f} (权重{OCF_SUB_W})")
    print(f"    负债率={debt} → percentile={debt_s:.2f} (reverse, 权重{DEBT_SUB_W})")
    print(f"    cfsafe_score = ({ocf_s:.2f}×{OCF_SUB_W} + {debt_s:.2f}×{DEBT_SUB_W}) × {MARKET_FALLBACK_DISC} = {cfsafe_score:.2f}")
    
    # 总分
    total = profit_score * PROFIT_W + growth_score * GROWTH_W + cfsafe_score * CFSAFE_W
    
    print(f"\n  总分 = {profit_score:.2f}×{PROFIT_W} + {growth_score:.2f}×{GROWTH_W} + {cfsafe_score:.2f}×{CFSAFE_W} = {total:.2f}")
    print(f"  Excel总分 = {s['total_score']}")
    print(f"  差异 = {total - s['total_score']:.2f}")
    
    # 检查等级
    if total >= 75: grade = 'A'
    elif total >= 55: grade = 'B'
    elif total >= 40: grade = 'C'
    elif total >= 25: grade = 'D'
    else: grade = 'E'
    print(f"  计算等级 = {grade}, Excel等级 = {s['grade']}")

# ========== 特别检查：营收同比为 None 的影响 ==========
print(f"\n{'='*60}")
print("特别检查: 营收同比缺失的影响")
print(f"  鼎泰高科: 营收同比=None → rev_s=0 → 成长得分损失")
print(f"  金海通: 营收同比=None → rev_s=0 → 成长得分损失")
print(f"\n  如果鼎泰高科有营收同比数据（假设=19.65%来自年报）:")
rev_test = 19.65
rev_s_test = percentile_rank(rev_test, pv('revenue_yoy')) if pv('revenue_yoy') else 50.0
print(f"    营收同比={rev_test} → percentile={rev_s_test:.2f}")
prof_s_ding = percentile_rank(259, pv('profit_yoy'))
growth_test = (rev_s_test * REV_YOY_SUB_W + prof_s_ding * PROF_YOY_SUB_W) * MARKET_FALLBACK_DISC
profit_score_ding = targets['301377.SZ']['profit_score']
cfsafe_score_ding = targets['301377.SZ']['cfsafe_score']
total_test = profit_score_ding * PROFIT_W + growth_test * GROWTH_W + cfsafe_score_ding * CFSAFE_W
print(f"    成长得分变为: {growth_test:.2f} (当前 {targets['301377.SZ']['growth_score']})")
print(f"    总分变为: {total_test:.2f} (当前 {targets['301377.SZ']['total_score']})")
