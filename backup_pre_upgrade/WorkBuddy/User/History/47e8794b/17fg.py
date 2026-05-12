#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import random

wb = openpyxl.load_workbook(r'D:\Project\AnnualScorer\股票业绩评价_20260426_204545.xlsx')
print('Sheet:', wb.sheetnames)

ws = wb.active
print('总行数:', ws.max_row)
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
print('列头:', headers)
print()

# 随机选2只股票（跳过表头）
if ws.max_row > 3:
    rows = random.sample(range(2, ws.max_row + 1), 2)
    for r in rows:
        print(f'--- 行{r} ---')
        for i, h in enumerate(headers, 1):
            val = ws.cell(r, i).value
            print(f'  {h}: {val}')
        print()
