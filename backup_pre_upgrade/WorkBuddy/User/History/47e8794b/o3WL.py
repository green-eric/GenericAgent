#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import random
import json

wb = openpyxl.load_workbook(r'D:\Project\AnnualScorer\股票业绩评价_20260426_204545.xlsx')
ws = wb.active
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

print('Sheet:', wb.sheetnames)
print('总行数:', ws.max_row)
print('列头:', headers)
print()

# 随机选2只（跳过表头，只选有数据的行）
valid_rows = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 2).value  # 股票名称列
    score = ws.cell(r, 3).value  # 评分列
    if name and score is not None:
        valid_rows.append(r)

print(f'有效数据行: {len(valid_rows)}')
print()

if len(valid_rows) >= 2:
    chosen = random.sample(valid_rows, 2)
    for r in chosen:
        print(f'=== 行{r} ===')
        row_data = {}
        for i, h in enumerate(headers, 1):
            val = ws.cell(r, i).value
            row_data[h] = val
            print(f'  {h}: {val}')
        print()
