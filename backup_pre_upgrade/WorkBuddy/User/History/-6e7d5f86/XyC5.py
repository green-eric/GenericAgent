#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整理数据质量报告并输出Excel到桌面"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = {
  "success": True,
  "data": {
    "report_id": "dqr_20260423_225703",
    "generated_at": "2026-04-23 22:57:03",
    "last_check_time": "2026-04-23 22:56:25",
    "overall": {"score": 94, "level": "优秀"},
    "sources": {
      "news": {
        "source": "news",
        "indicators": [
          {"name": "新闻标题", "description": "新闻标题文本，非空且长度>=2", "unit": "字符", "business_impact": "情绪评分(6分)+业绩评分(20分)核心输入，标题缺失导致情绪/业绩评分失真", "scoring_dimensions": ["sentiment", "performance"], "required_fields": ["title"]},
          {"name": "新闻时间", "description": "新闻发布时间戳，需在合理范围内", "unit": "时间戳", "business_impact": "情绪评分时效性过滤，过期新闻导致情绪评分失真", "scoring_dimensions": ["sentiment"], "required_fields": ["ctime"]},
          {"name": "新闻来源", "description": "新闻来源标识(ths/eastmoney/cls_flash/sina)", "unit": "枚举", "business_impact": "情绪评分来源权重(ths:1.0, sina:0.8)，影响情绪评分质量", "scoring_dimensions": ["sentiment"], "required_fields": ["source"]},
          {"name": "关联股票代码", "description": "新闻关联的A股代码列表", "unit": "代码", "business_impact": "情绪/业绩评分股票-新闻匹配，缺失导致两维度评分无法关联到具体股票", "scoring_dimensions": ["sentiment", "performance"], "required_fields": ["_codes"]},
          {"name": "新闻影响度", "description": "新闻对股票的影响程度(0-16)", "unit": "分数", "business_impact": "情绪评分6级递增核心依据(impact>=1/3/5/8/12/16)", "scoring_dimensions": ["sentiment"], "required_fields": ["impact"]}
        ],
        "quality_checks": {
          "news_completeness": {"dimension": "completeness", "passed": True, "score": 100, "level": "优秀", "total_count": 452, "pass_count": 452, "fail_count": 0},
          "news_validity": {"dimension": "validity", "passed": True, "score": 100, "level": "优秀", "total_count": 452, "pass_count": 452, "fail_count": 0},
          "news_timeliness": {"dimension": "timeliness", "passed": True, "score": 100, "level": "优秀", "total_count": 452, "pass_count": 452, "fail_count": 0},
          "news_consistency": {"dimension": "consistency", "passed": True, "score": 100, "level": "优秀", "total_count": 452, "pass_count": 452, "fail_count": 0},
          "news_accuracy": {"dimension": "accuracy", "passed": True, "score": 100, "level": "优秀", "total_count": 452, "pass_count": 452, "fail_count": 0},
          "news_uniqueness": {"dimension": "uniqueness", "passed": True, "score": 100, "level": "优秀", "total_count": 452, "pass_count": 452, "fail_count": 0}
        },
        "alerts": [], "active_alerts": [], "resolved_alerts": [], "active_alert_count": 0
      },
      "quote": {
        "source": "quote",
        "indicators": [
          {"name": "当前价格", "description": "股票当前/最新价格，必须为正数", "unit": "元", "business_impact": "量价/强势/资金/龙头评分基础，价格异常导致多维度评分失真", "scoring_dimensions": ["vol_price", "seal", "fund", "leader"], "required_fields": ["price"]},
          {"name": "涨跌幅", "description": "当日涨跌幅百分比", "unit": "%", "business_impact": "量价/强势/资金(放量下跌惩罚)/龙头(涨停质量+相对强度)核心输入", "scoring_dimensions": ["vol_price", "seal", "fund", "leader"], "required_fields": ["pct"]},
          {"name": "成交量/额", "description": "当日成交量或成交额", "unit": "手/元", "business_impact": "资金评分量比计算基础，停牌判断依据", "scoring_dimensions": ["fund"], "required_fields": ["vol", "amount"]},
          {"name": "换手率", "description": "当日换手率百分比", "unit": "%", "business_impact": "资金评分子项(权重33.3%)", "scoring_dimensions": ["fund"], "required_fields": ["turnover"]},
          {"name": "量比", "description": "当日量比(当前成交量/前5日同时段均量)，pipeline计算字段", "unit": "倍", "business_impact": "资金评分核心子项(权重33.3%)，量价评分量能确认输入", "scoring_dimensions": ["fund", "vol_price"], "required_fields": ["vol_ratio"]},
          {"name": "股票名称", "description": "股票名称，用于ST/退市风险识别", "unit": "文本", "business_impact": "强势评分ST检测，影响可买入性抑制系数(0.15-0.85)", "scoring_dimensions": ["seal"], "required_fields": ["name"]}
        ],
        "quality_checks": {
          "quote_completeness": {"dimension": "completeness", "passed": True, "score": 100, "level": "优秀", "total_count": 10, "pass_count": 10, "fail_count": 0},
          "quote_validity": {"dimension": "validity", "passed": True, "score": 100, "level": "优秀", "total_count": 10, "pass_count": 10, "fail_count": 0},
          "quote_timeliness": {"dimension": "timeliness", "passed": True, "score": 100, "level": "优秀", "total_count": 10, "pass_count": 10, "fail_count": 0},
          "quote_consistency": {"dimension": "consistency", "passed": False, "score": 0, "level": "极差", "total_count": 10, "pass_count": 0, "fail_count": 10},
          "quote_accuracy": {"dimension": "accuracy", "passed": True, "score": 100, "level": "优秀", "total_count": 10, "pass_count": 10, "fail_count": 0},
          "quote_uniqueness": {"dimension": "uniqueness", "passed": True, "score": 100, "level": "优秀", "total_count": 10, "pass_count": 10, "fail_count": 0}
        },
        "alerts": [], "active_alerts": [], "resolved_alerts": [], "active_alert_count": 0
      },
      "kline": {
        "source": "kline",
        "indicators": [
          {"name": "K线日期", "description": "K线数据日期，格式YYYY-MM-DD", "unit": "日期", "business_impact": "K线时序正确性，日期缺失/错乱导致稳定性/强势/情绪技术指标计算错误", "scoring_dimensions": ["stability", "fund", "seal", "sentiment"], "required_fields": ["date"]},
          {"name": "K线OHLC", "description": "开盘/最高/最低/收盘价，需满足high>=low, open/close在[low,high]内", "unit": "元", "business_impact": "稳定性/资金/强势评分核心输入，情绪技术面兜底代理依赖收盘价，OHLC不一致导致评分严重偏差", "scoring_dimensions": ["stability", "fund", "seal", "sentiment"], "required_fields": ["open", "high", "low", "close"]},
          {"name": "K线成交量", "description": "K线成交量，非负", "unit": "手", "business_impact": "资金评分量比计算、强势评分异常放量阴线检测", "scoring_dimensions": ["fund", "seal"], "required_fields": ["volume"]},
          {"name": "K线连续性", "description": "K线日期连续无断档，至少覆盖20个交易日", "unit": "天", "business_impact": "稳定性评分需20日回看，K线不足导致评分降级", "scoring_dimensions": ["stability"], "required_fields": []}
        ],
        "quality_checks": {
          "kline_accuracy": {"dimension": "accuracy", "passed": True, "score": 100, "level": "优秀", "total_count": 0, "pass_count": 0, "fail_count": 0},
          "kline_completeness": {"dimension": "completeness", "passed": True, "score": 100, "level": "优秀", "total_count": 0, "pass_count": 0, "fail_count": 0},
          "kline_consistency": {"dimension": "consistency", "passed": True, "score": 100, "level": "优秀", "total_count": 0, "pass_count": 0, "fail_count": 0},
          "kline_timeliness": {"dimension": "timeliness", "passed": True, "score": 100, "level": "优秀", "total_count": 0, "pass_count": 0, "fail_count": 0},
          "kline_validity": {"dimension": "validity", "passed": True, "score": 100, "level": "优秀", "total_count": 0, "pass_count": 0, "fail_count": 0},
          "kline_uniqueness": {"dimension": "uniqueness", "passed": True, "score": 100, "level": "优秀", "total_count": 0, "pass_count": 0, "fail_count": 0}
        },
        "alerts": [], "active_alerts": [], "resolved_alerts": [], "active_alert_count": 0
      },
      "earnings": {
        "source": "earnings",
        "indicators": [
          {"name": "净利润增速", "description": "净利润同比增长率", "unit": "%", "business_impact": "业绩评分成长能力子项(权重40%*60%=24%)", "scoring_dimensions": ["performance"], "required_fields": ["net_profit_yoy"]},
          {"name": "营收增速", "description": "营收同比增长率", "unit": "%", "business_impact": "业绩评分成长能力子项(权重40%*40%=16%)", "scoring_dimensions": ["performance"], "required_fields": ["revenue_yoy"]},
          {"name": "ROE", "description": "净资产收益率", "unit": "%", "business_impact": "业绩评分盈利能力子项(权重30%*60%=18%)", "scoring_dimensions": ["performance"], "required_fields": ["roe"]},
          {"name": "预告类型", "description": "业绩预告类型(预增/预减/扭亏/首亏等)", "unit": "枚举", "business_impact": "业绩趋势评分子项(权重30%*60%=18%)", "scoring_dimensions": ["performance"], "required_fields": ["forecast_type"]}
        ],
        "quality_checks": {
          "earnings_completeness": {"dimension": "completeness", "passed": True, "score": 100, "level": "优秀", "total_count": 50, "pass_count": 50, "fail_count": 0},
          "earnings_validity": {"dimension": "validity", "passed": True, "score": 100, "level": "优秀", "total_count": 50, "pass_count": 50, "fail_count": 0},
          "earnings_consistency": {"dimension": "consistency", "passed": True, "score": 100, "level": "优秀", "total_count": 50, "pass_count": 50, "fail_count": 0},
          "earnings_timeliness": {"dimension": "timeliness", "passed": True, "score": 100, "level": "优秀", "total_count": 50, "pass_count": 50, "fail_count": 0},
          "earnings_accuracy": {"dimension": "accuracy", "passed": True, "score": 100, "level": "优秀", "total_count": 50, "pass_count": 50, "fail_count": 0},
          "earnings_uniqueness": {"dimension": "uniqueness", "passed": True, "score": 100, "level": "优秀", "total_count": 50, "pass_count": 50, "fail_count": 0}
        },
        "alerts": [], "active_alerts": [], "resolved_alerts": [], "active_alert_count": 0
      },
      "sector": {
        "source": "sector",
        "indicators": [
          {"name": "板块热度", "description": "板块热度值(0-10+)", "unit": "分数", "business_impact": "板块评分(12分)核心输入，幂函数归一化", "scoring_dimensions": ["sector"], "required_fields": ["heat"]},
          {"name": "板块-股票映射", "description": "板块到成分股的映射关系", "unit": "映射", "business_impact": "龙头评分板块地位/相对强度计算", "scoring_dimensions": ["leader", "sector"], "required_fields": ["sector", "stocks"]}
        ],
        "quality_checks": {
          "sector_completeness": {"dimension": "completeness", "passed": True, "score": 100, "level": "优秀", "total_count": 20, "pass_count": 20, "fail_count": 0},
          "sector_validity": {"dimension": "validity", "passed": True, "score": 100, "level": "优秀", "total_count": 20, "pass_count": 20, "fail_count": 0},
          "sector_consistency": {"dimension": "consistency", "passed": True, "score": 100, "level": "优秀", "total_count": 20, "pass_count": 20, "fail_count": 0},
          "sector_timeliness": {"dimension": "timeliness", "passed": True, "score": 100, "level": "优秀", "total_count": 20, "pass_count": 20, "fail_count": 0},
          "sector_accuracy": {"dimension": "accuracy", "passed": True, "score": 100, "level": "优秀", "total_count": 20, "pass_count": 20, "fail_count": 0},
          "sector_uniqueness": {"dimension": "uniqueness", "passed": True, "score": 100, "level": "优秀", "total_count": 20, "pass_count": 20, "fail_count": 0}
        },
        "alerts": [], "active_alerts": [], "resolved_alerts": [], "active_alert_count": 0
      },
      "hot_stock": {
        "source": "hot_stock",
        "indicators": [
          {"name": "热门股票代码", "description": "热门股票代码，需为有效A股代码", "unit": "代码", "business_impact": "Pipeline核心输入，热门股优先获取K线和评分", "scoring_dimensions": ["leader", "sector"], "required_fields": ["symbol"]},
          {"name": "热度值", "description": "股票热度分数", "unit": "分数", "business_impact": "多源热度融合计算(东财40%+新闻30%+交易20%+动量10%)", "scoring_dimensions": ["sector", "leader"], "required_fields": ["heat"]}
        ],
        "quality_checks": {
          "hot_stock_completeness": {"dimension": "completeness", "passed": True, "score": 100, "level": "优秀", "total_count": 8, "pass_count": 8, "fail_count": 0},
          "hot_stock_validity": {"dimension": "validity", "passed": True, "score": 100, "level": "优秀", "total_count": 8, "pass_count": 8, "fail_count": 0},
          "hot_stock_consistency": {"dimension": "consistency", "passed": True, "score": 100, "level": "优秀", "total_count": 8, "pass_count": 8, "fail_count": 0},
          "hot_stock_timeliness": {"dimension": "timeliness", "passed": True, "score": 100, "level": "优秀", "total_count": 8, "pass_count": 8, "fail_count": 0},
          "hot_stock_accuracy": {"dimension": "accuracy", "passed": True, "score": 100, "level": "优秀", "total_count": 8, "pass_count": 8, "fail_count": 0},
          "hot_stock_uniqueness": {"dimension": "uniqueness", "passed": True, "score": 100, "level": "优秀", "total_count": 8, "pass_count": 8, "fail_count": 0}
        },
        "alerts": [], "active_alerts": [], "resolved_alerts": [], "active_alert_count": 0
      }
    },
    "alerts_summary": {"total": 0, "by_severity": {"critical": 0, "warning": 0, "info": 0}, "unresolved": 0, "recent": []},
    "recommendations": [
      {"source": "quote", "dimension": "consistency", "priority": "high", "current_score": "0.0", "suggestion": "检查OHLC数据一致性(high>=low, price在[low,high]内)；启用多源交叉验证"}
    ]
  }
}

# 样式定义
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
SUBTITLE_FONT = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
NORMAL_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT = Font(name='微软雅黑', bold=True, size=10)
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
HIGHLIGHT_FILL = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

def style_data(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_ALIGN if c > 2 else CENTER_ALIGN
            cell.border = THIN_BORDER

def auto_width(ws, max_col, min_w=10, max_w=50):
    for c in range(1, max_col + 1):
        max_len = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    l = len(str(cell.value))
                    if l > max_len:
                        max_len = min(l * 1.5, max_w)
        ws.column_dimensions[get_column_letter(c)].width = max_len

report = DATA["data"]
wb = Workbook()

# ===================== Sheet1: 报告概览 =====================
ws1 = wb.active
ws1.title = "报告概览"
ws1.merge_cells('A1:F1')
ws1.cell(row=1, column=1, value="数据质量报告概览").font = TITLE_FONT
ws1.cell(row=1, column=1).alignment = CENTER_ALIGN

overview = [
    ["报告ID", report["report_id"]],
    ["生成时间", report["generated_at"]],
    ["上次检查", report["last_check_time"]],
    ["综合评分", f'{report["overall"]["score"]}分'],
    ["综合等级", report["overall"]["level"]],
    ["告警总数", report["alerts_summary"]["total"]],
    ["严重告警", report["alerts_summary"]["by_severity"]["critical"]],
    ["警告", report["alerts_summary"]["by_severity"]["warning"]],
    ["提示", report["alerts_summary"]["by_severity"]["info"]],
]
for i, (k, v) in enumerate(overview, 3):
    ws1.cell(row=i, column=1, value=k).font = BOLD_FONT
    ws1.cell(row=i, column=2, value=v).font = NORMAL_FONT
    ws1.cell(row=i, column=1).border = THIN_BORDER
    ws1.cell(row=i, column=2).border = THIN_BORDER

# 各数据源评分汇总
r = 13
ws1.merge_cells(f'A{r}:F{r}')
ws1.cell(row=r, column=1, value="各数据源质量评分汇总").font = SUBTITLE_FONT
r += 1
headers = ["数据源", "完整性", "有效性", "时效性", "一致性", "准确性", "唯一性", "数据量"]
for c, h in enumerate(headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, len(headers))

source_names = {"news": "新闻(news)", "quote": "行情(quote)", "kline": "K线(kline)", "earnings": "财报(earnings)", "sector": "板块(sector)", "hot_stock": "热门股(hot_stock)"}
dims = ["completeness", "validity", "timeliness", "consistency", "accuracy", "uniqueness"]

for src_key, src_name in source_names.items():
    r += 1
    src = report["sources"][src_key]
    ws1.cell(row=r, column=1, value=src_name).font = BOLD_FONT
    total_data = 0
    for di, dim in enumerate(dims):
        chk_key = f"{src_key}_{dim}"
        chk = src["quality_checks"].get(chk_key, {})
        score = chk.get("score", "N/A")
        cell = ws1.cell(row=r, column=di + 2, value=score)
        if isinstance(score, (int, float)):
            if score >= 80:
                cell.fill = GREEN_FILL
            elif score >= 50:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = RED_FILL
            total_data = max(total_data, chk.get("total_count", 0))
    ws1.cell(row=r, column=8, value=total_data)
style_data(ws1, 14, r, len(headers))
auto_width(ws1, len(headers))

# ===================== Sheet2: 新闻时间(重点) =====================
ws2 = wb.create_sheet("新闻时间(重点分析)")
ws2.merge_cells('A1:F1')
ws2.cell(row=1, column=1, value="新闻时间 - 时效性分析").font = TITLE_FONT
ws2.cell(row=1, column=1).alignment = CENTER_ALIGN

# 新闻时间指标详情
r = 3
ws2.cell(row=r, column=1, value="1. 指标定义").font = SUBTITLE_FONT
r += 1
news_time_info = [
    ["指标名称", "新闻时间"],
    ["描述", "新闻发布时间戳，需在合理范围内"],
    ["单位", "时间戳"],
    ["业务影响", "情绪评分时效性过滤，过期新闻导致情绪评分失真"],
    ["评分维度", "sentiment(情绪)"],
    ["必需字段", "ctime"],
]
for k, v in news_time_info:
    ws2.cell(row=r, column=1, value=k).font = BOLD_FONT
    ws2.cell(row=r, column=2, value=v).font = NORMAL_FONT
    ws2.cell(row=r, column=1).border = THIN_BORDER
    ws2.cell(row=r, column=2).border = THIN_BORDER
    ws2.cell(row=r, column=2).fill = HIGHLIGHT_FILL
    r += 1

# 时效性检查结果
r += 1
ws2.cell(row=r, column=1, value="2. 时效性检查结果").font = SUBTITLE_FONT
r += 1
chk = report["sources"]["news"]["quality_checks"]["news_timeliness"]
chk_info = [
    ["检查维度", "timeliness(时效性)"],
    ["是否通过", "通过" if chk["passed"] else "未通过"],
    ["评分", f'{chk["score"]}分'],
    ["等级", chk["level"]],
    ["总记录数", chk["total_count"]],
    ["通过数", chk["pass_count"]],
    ["失败数", chk["fail_count"]],
]
for k, v in chk_info:
    ws2.cell(row=r, column=1, value=k).font = BOLD_FONT
    ws2.cell(row=r, column=2, value=v).font = NORMAL_FONT
    ws2.cell(row=r, column=1).border = THIN_BORDER
    ws2.cell(row=r, column=2).border = THIN_BORDER
    if k == "是否通过":
        ws2.cell(row=r, column=2).fill = GREEN_FILL if chk["passed"] else RED_FILL
    r += 1

# 新闻源全部指标
r += 1
ws2.cell(row=r, column=1, value="3. 新闻源全部指标一览").font = SUBTITLE_FONT
r += 1
headers = ["指标名称", "描述", "单位", "业务影响", "评分维度", "必需字段"]
for c, h in enumerate(headers, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, len(headers))
for ind in report["sources"]["news"]["indicators"]:
    r += 1
    ws2.cell(row=r, column=1, value=ind["name"])
    ws2.cell(row=r, column=2, value=ind["description"])
    ws2.cell(row=r, column=3, value=ind["unit"])
    ws2.cell(row=r, column=4, value=ind["business_impact"])
    ws2.cell(row=r, column=5, value=", ".join(ind["scoring_dimensions"]))
    ws2.cell(row=r, column=6, value=", ".join(ind["required_fields"]))
    if ind["name"] == "新闻时间":
        for c in range(1, 7):
            ws2.cell(row=r, column=c).fill = YELLOW_FILL
style_data(ws2, r - len(report["sources"]["news"]["indicators"]) + 1, r, len(headers))

# 新闻源6维质量
r += 2
ws2.cell(row=r, column=1, value="4. 新闻源六维质量检查").font = SUBTITLE_FONT
r += 1
headers2 = ["检查项", "维度", "通过", "评分", "等级", "总数", "通过数", "失败数"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, len(headers2))
for chk_key, chk in report["sources"]["news"]["quality_checks"].items():
    r += 1
    ws2.cell(row=r, column=1, value=chk_key)
    ws2.cell(row=r, column=2, value=chk["dimension"])
    ws2.cell(row=r, column=3, value="通过" if chk["passed"] else "未通过")
    ws2.cell(row=r, column=4, value=chk["score"])
    ws2.cell(row=r, column=5, value=chk["level"])
    ws2.cell(row=r, column=6, value=chk["total_count"])
    ws2.cell(row=r, column=7, value=chk["pass_count"])
    ws2.cell(row=r, column=8, value=chk["fail_count"])
    if not chk["passed"]:
        for c in range(1, 9):
            ws2.cell(row=r, column=c).fill = RED_FILL
    elif chk["dimension"] == "timeliness":
        for c in range(1, 9):
            ws2.cell(row=r, column=c).fill = GREEN_FILL
style_data(ws2, r - 5, r, len(headers2))
auto_width(ws2, 8)

# ===================== Sheet3: 全部指标明细 =====================
ws3 = wb.create_sheet("全部指标明细")
ws3.merge_cells('A1:G1')
ws3.cell(row=1, column=1, value="全数据源指标明细").font = TITLE_FONT
ws3.cell(row=1, column=1).alignment = CENTER_ALIGN

r = 3
headers3 = ["数据源", "指标名称", "描述", "单位", "业务影响", "评分维度", "必需字段"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, len(headers3))

src_labels = {"news": "新闻", "quote": "行情", "kline": "K线", "earnings": "财报", "sector": "板块", "hot_stock": "热门股"}
for src_key, src_data in report["sources"].items():
    for ind in src_data["indicators"]:
        r += 1
        ws3.cell(row=r, column=1, value=src_labels.get(src_key, src_key))
        ws3.cell(row=r, column=2, value=ind["name"])
        ws3.cell(row=r, column=3, value=ind["description"])
        ws3.cell(row=r, column=4, value=ind["unit"])
        ws3.cell(row=r, column=5, value=ind["business_impact"])
        ws3.cell(row=r, column=6, value=", ".join(ind["scoring_dimensions"]))
        ws3.cell(row=r, column=7, value=", ".join(ind["required_fields"]))
        if "时间" in ind["name"] or "时效" in ind["description"]:
            for c in range(1, 8):
                ws3.cell(row=r, column=c).fill = YELLOW_FILL
style_data(ws3, 4, r, len(headers3))
auto_width(ws3, len(headers3))

# ===================== Sheet4: 质量检查明细 =====================
ws4 = wb.create_sheet("质量检查明细")
ws4.merge_cells('A1:I1')
ws4.cell(row=1, column=1, value="全数据源质量检查明细").font = TITLE_FONT
ws4.cell(row=1, column=1).alignment = CENTER_ALIGN

r = 3
headers4 = ["数据源", "检查项", "维度", "是否通过", "评分", "等级", "总数", "通过数", "失败数"]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, len(headers4))

for src_key, src_data in report["sources"].items():
    for chk_key, chk in src_data["quality_checks"].items():
        r += 1
        ws4.cell(row=r, column=1, value=src_labels.get(src_key, src_key))
        ws4.cell(row=r, column=2, value=chk_key)
        ws4.cell(row=r, column=3, value=chk["dimension"])
        ws4.cell(row=r, column=4, value="通过" if chk["passed"] else "未通过")
        ws4.cell(row=r, column=5, value=chk["score"])
        ws4.cell(row=r, column=6, value=chk["level"])
        ws4.cell(row=r, column=7, value=chk["total_count"])
        ws4.cell(row=r, column=8, value=chk["pass_count"])
        ws4.cell(row=r, column=9, value=chk["fail_count"])
        if not chk["passed"]:
            for c in range(1, 10):
                ws4.cell(row=r, column=c).fill = RED_FILL
        elif chk["dimension"] == "timeliness":
            for c in range(1, 10):
                ws4.cell(row=r, column=c).fill = GREEN_FILL
style_data(ws4, 4, r, len(headers4))
auto_width(ws4, len(headers4))

# ===================== Sheet5: 改进建议 =====================
ws5 = wb.create_sheet("改进建议")
ws5.merge_cells('A1:E1')
ws5.cell(row=1, column=1, value="数据质量改进建议").font = TITLE_FONT
ws5.cell(row=1, column=1).alignment = CENTER_ALIGN

r = 3
headers5 = ["数据源", "维度", "优先级", "当前评分", "改进建议"]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, len(headers5))

for rec in report["recommendations"]:
    r += 1
    ws5.cell(row=r, column=1, value=rec["source"])
    ws5.cell(row=r, column=2, value=rec["dimension"])
    ws5.cell(row=r, column=3, value=rec["priority"])
    ws5.cell(row=r, column=4, value=rec["current_score"])
    ws5.cell(row=r, column=5, value=rec["suggestion"])
    if rec["priority"] == "high":
        for c in range(1, 6):
            ws5.cell(row=r, column=c).fill = RED_FILL
style_data(ws5, 4, r, len(headers5))
auto_width(ws5, len(headers5))

# 保存
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
outpath = os.path.join(desktop, "数据质量报告_新闻时效性分析.xlsx")
wb.save(outpath)
print(f"Excel saved: {outpath}")
