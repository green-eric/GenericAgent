#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 Excel 导出模块 V6.0.0
 负责将评分结果导出为带格式的 Excel 文件
================================================================================
"""
import os
import logging
from datetime import datetime
from typing import List, Dict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger('ScoreSys')


# ============================================================================
# Excel 输出字段映射
# ============================================================================
EXCEL_COLUMNS = {
    # 基础信息
    'symbol': ('股票代码', 'Code'),
    'name': ('股票名称', 'Name'),
    'industry': ('申万一级行业', 'Industry'),
    'annual_report_date': ('年报日期', 'AnnualReportDate'),
    'latest_quarter': ('最新季报期', 'LatestQuarter'),
    'data_completeness': ('数据完整度', 'DataCompleteness'),
    'quarter_coverage': ('季度覆盖', 'QuarterCov'),
    'field_gaps': ('缺失字段', 'FieldGaps'),
    # 成长性
    'q_revenue_yoy': ('营收同比(%)(单季)', 'RevenueYoY'),
    'q_net_profit_yoy': ('净利润同比(%)(单季)', 'ProfitYoY'),
    'growth': ('成长性', 'Growth'),
    # 盈利能力
    'roe_ttm': ('ROE(%)(TTM)', 'ROE_TTM'),
    'gross_margin_ttm': ('毛利率(%)(TTM)', 'GrossMargin_TTM'),
    'profitability': ('盈利能力', 'Profitability'),
    # 现金流质量
    'net_profit_ratio': ('净现比', 'OCFtoProfit'),
    'fcf_yield': ('FCF收益率', 'FCFYield'),
    'cash_recovery_rate': ('收现比', 'CashRecoveryRate'),
    'cash_flow_quality': ('现金流质量', 'CashFlow'),
    # 偿债风险
    'de_ratio': ('D/E(倍)', 'DebtEquityRatio'),
    'current_ratio': ('流动比率(倍)', 'CurrentRatio'),
    'asset_liability_ratio': ('资产负债率(%)(单季)', 'AssetLiabilityRatio'),
    'leverage_risk': ('偿债风险', 'Leverage'),
    # 估值
    'pe_ttm': ('市盈率TTM', 'PE_TTM'),
    'total_mv': ('总市值(亿)', 'MarketValue'),
    # 综合
    'total_score': ('总分', 'TotalScore'),
    'rating': ('评级', 'Rating'),
    'confidence': ('置信度', 'Confidence'),
    # 否决
    'veto': ('触发否决', 'Veto'),
    'veto_reason': ('否决原因', 'VetoReason'),
}

# 维度分组背景色
DIMENSION_FILLS = {
    'info': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
    'growth': PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),
    'profitability': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'cashflow': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'leverage': PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),
    'valuation': PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),
    'total': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'veto': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
}

# 列到维度的映射
COL_DIMENSIONS = {
    'symbol': 'info', 'name': 'info', 'industry': 'info',
    'annual_report_date': 'info', 'latest_quarter': 'info', 'data_completeness': 'info',
    'quarter_coverage': 'info', 'field_gaps': 'info',
    'total_score': 'total', 'rating': 'total', 'confidence': 'total',
    'q_revenue_yoy': 'growth', 'q_net_profit_yoy': 'growth', 'growth': 'growth',
    'roe_ttm': 'profitability', 'gross_margin_ttm': 'profitability', 'profitability': 'profitability',
    'net_profit_ratio': 'cashflow', 'fcf_yield': 'cashflow',
    'cash_recovery_rate': 'cashflow', 'cash_flow_quality': 'cashflow',
    'net_profit_ttm': 'cashflow', 'ocf_ttm': 'cashflow',
    'de_ratio': 'leverage', 'current_ratio': 'leverage',
    'asset_liability_ratio': 'leverage', 'leverage_risk': 'leverage',
    'pe_ttm': 'valuation', 'total_mv': 'valuation',
    'veto': 'veto', 'veto_reason': 'veto',
}

# 评级颜色
RATING_COLORS = {
    'A+': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
    'A': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
    'B+': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
    'B': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
    'C': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
    'D': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
}

# 默认列顺序
DEFAULT_COLUMN_ORDER = [
    'symbol', 'name',
    'total_score', 'rating', 'confidence',
    'industry',
    'q_revenue_yoy', 'q_net_profit_yoy', 'growth',
    'roe_ttm', 'gross_margin_ttm', 'profitability',
    'net_profit_ratio', 'fcf_yield', 'cash_recovery_rate', 'cash_flow_quality',
    'de_ratio', 'current_ratio', 'asset_liability_ratio', 'leverage_risk',
    'pe_ttm', 'total_mv',
    'annual_report_date', 'latest_quarter', 'data_completeness', 'quarter_coverage', 'field_gaps',
    'veto', 'veto_reason',
]


def save_to_excel(results: List[Dict], filename: str = None, max_keep: int = 10) -> str:
    """
    将评分结果保存为带格式的 Excel 文件。

    Args:
        results: 评分结果列表
        filename: 输出文件名（默认自动生成带时间戳）
        max_keep: 保留最近N份评分结果，超出自动删除

    Returns:
        保存的文件路径
    """
    if not results:
        logger.warning("无结果可保存")
        return ""

    df = pd.DataFrame(results)
    df = df.sort_values('total_score', ascending=False)

    # 默认文件名带时间戳
    if filename is None:
        filename = f"评分结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # 自动清理旧文件：保留最近N份
    output_dir = os.path.dirname(os.path.abspath(filename))
    prefix = '评分结果_'
    if os.path.basename(filename).startswith(prefix):
        old_files = sorted(
            [f for f in os.listdir(output_dir) if f.startswith(prefix) and f.endswith('.xlsx')],
            reverse=True
        )
        for old_f in old_files[max_keep:]:
            try:
                os.remove(os.path.join(output_dir, old_f))
            except OSError:
                pass

    # 排序列顺序
    cols = [c for c in DEFAULT_COLUMN_ORDER if c in df.columns]
    df = df[cols]

    # 构建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '评分结果'

    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 表头
    headers = []
    header_fills = []
    for col in df.columns:
        cn, en = EXCEL_COLUMNS.get(col, (col, col))
        headers.append(f"{cn}\n({en})")
        dim = COL_DIMENSIONS.get(col, 'info')
        header_fills.append(DIMENSION_FILLS[dim])

    for col_idx, (header, hfill) in enumerate(zip(headers, header_fills), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hfill
        cell.font = Font(bold=True, size=10)
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行 — 先批量写入所有值
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 按维度设置整列背景色
    for col_idx in range(1, len(df.columns) + 1):
        col_name = df.columns[col_idx - 1]
        dim = COL_DIMENSIONS.get(col_name, 'info')
        col_fill = DIMENSION_FILLS[dim]
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = col_fill
            cell.alignment = cell_align
            cell.border = thin_border

    # 特殊列的逐单元格样式
    special_col_indices = {}
    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name in ('rating', 'data_completeness', 'quarter_coverage', 'field_gaps', 'total_score'):
            special_col_indices[col_name] = col_idx

    for row_idx, row in enumerate(df.values, 2):
        # 评级列特殊颜色
        if 'rating' in special_col_indices:
            col_idx = special_col_indices['rating']
            value = row[col_idx - 1]
            if value in RATING_COLORS:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = RATING_COLORS[value]
                if value in ['A+', 'A', 'D']:
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.font = Font(bold=True, color='000000')

        # 数据完整度条件着色
        if 'data_completeness' in special_col_indices:
            col_idx = special_col_indices['data_completeness']
            value = row[col_idx - 1]
            if isinstance(value, (int, float)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if value < 60:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
                elif value < 80:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=True)

        # 季度覆盖高亮
        if 'quarter_coverage' in special_col_indices:
            col_idx = special_col_indices['quarter_coverage']
            value = row[col_idx - 1]
            if isinstance(value, str) and '缺' in value:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                cell.font = Font(bold=True)

        # 缺失字段高亮
        if 'field_gaps' in special_col_indices:
            col_idx = special_col_indices['field_gaps']
            value = row[col_idx - 1]
            if isinstance(value, str) and value != '无':
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                cell.font = Font(bold=True)

        # 总分加粗
        if 'total_score' in special_col_indices:
            col_idx = special_col_indices['total_score']
            ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    # 列宽（采样估算）
    for col_idx in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        header_len = len(headers[col_idx - 1])
        sample_size = min(100, len(df))
        max_len = header_len
        for r in range(2, sample_size + 2):
            v = ws.cell(row=r, column=col_idx).value
            if v:
                max_len = max(max_len, len(str(v)))
        for r in range(max(2, len(df) - 48), len(df) + 2):
            v = ws.cell(row=r, column=col_idx).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 28)

    ws.freeze_panes = 'A2'
    wb.save(filename)
    logger.info(f"结果已保存: {filename}")
    print(f"\n结果已保存: {filename}")
    return filename
