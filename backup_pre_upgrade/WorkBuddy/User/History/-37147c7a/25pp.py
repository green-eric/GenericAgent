#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 A股智能选股系统 v3.1.2
 参考格式输出：股票代码 | 股票名称 | 行业 | 财务指标 | 评分 | 评级
 数据流程：AKShare → SQLite → 计算 → 评分 → Excel
================================================================================
"""
import argparse
from datetime import datetime
from typing import List, Optional, Dict
import os
import sys
import json
import sqlite3
import logging
import threading
import time
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import VETO_RULES
from scorer import Scorer
from data_provider import DataProvider
from calculator import IndicatorCalculator


# 全局变量
_eval_date = None
_calendar = None


# ============================================================================
# 模拟数据
# ============================================================================
MOCK_DATA = {
    '600519': {
        'name': '贵州茅台', 'industry': '食品饮料',
        'roe_ttm': 30.5, 'gross_margin_ttm': 91.5, 'net_margin_ttm': 52.3,
        'q_revenue_yoy': 18.5, 'q_net_profit_yoy': 20.2,
        'asset_liability_ratio': 15.2, 'de_ratio': 0.18, 'current_ratio': 4.5,
        'net_profit_ratio': 136.5, 'net_profit_ttm': 880.5, 'ocf_ttm': 1200.5,
        'pe_ttm': 28.5, 'total_mv': 25000.0,
        'annual_report_date': '2024-04-27', 'latest_quarter': '2024-12-31', 'data_completeness': 100,
        'quarter_coverage': '4/4', 'field_gaps': '无',
        'veto': '否', 'veto_reason': '',
    },
    '000858': {
        'name': '五粮液', 'industry': '食品饮料',
        'roe_ttm': 18.2, 'gross_margin_ttm': 75.8, 'net_margin_ttm': 35.2,
        'q_revenue_yoy': 25.3, 'q_net_profit_yoy': 45.7,
        'asset_liability_ratio': 32.5, 'de_ratio': 0.48, 'current_ratio': 2.8,
        'net_profit_ratio': 115.2, 'net_profit_ttm': 215.5, 'ocf_ttm': 295.5,
        'pe_ttm': 22.5, 'total_mv': 8500.0,
        'annual_report_date': '2024-04-27', 'latest_quarter': '2024-12-31', 'data_completeness': 100,
        'quarter_coverage': '4/4', 'field_gaps': '无',
        'veto': '否', 'veto_reason': '',
    },
    '002415': {
        'name': '海康威视', 'industry': '电子',
        'roe_ttm': 22.5, 'gross_margin_ttm': 45.2, 'net_margin_ttm': 18.5,
        'q_revenue_yoy': 12.8, 'q_net_profit_yoy': 15.3,
        'asset_liability_ratio': 38.5, 'de_ratio': 0.62, 'current_ratio': 2.2,
        'net_profit_ratio': 95.5, 'net_profit_ttm': 145.2, 'ocf_ttm': 138.5,
        'pe_ttm': 25.8, 'total_mv': 3200.0,
        'annual_report_date': '2024-04-27', 'latest_quarter': '2024-12-31', 'data_completeness': 100,
        'quarter_coverage': '4/4', 'field_gaps': '无',
        'veto': '否', 'veto_reason': '',
    },
    '999999': {
        'name': '高杠杆股', 'industry': '钢铁',
        'roe_ttm': 5.2, 'gross_margin_ttm': 15.5, 'net_margin_ttm': 3.2,
        'q_revenue_yoy': -10.5, 'q_net_profit_yoy': -25.0,
        'asset_liability_ratio': 92.5, 'de_ratio': 12.5, 'current_ratio': 0.45,
        'net_profit_ratio': -45.0, 'net_profit_ttm': -20.5, 'ocf_ttm': -50.0,
        'pe_ttm': 0, 'total_mv': 500.0,
        'annual_report_date': '2024-04-27', 'latest_quarter': '2024-12-31', 'data_completeness': 100,
        'quarter_coverage': '4/4', 'field_gaps': '无',
        'veto': '是', 'veto_reason': '经营现金流为负',
    },
}


# ============================================================================
# Excel 输出字段映射（参考格式）
# ============================================================================
EXCEL_COLUMNS = {
    # 基础信息
    'symbol': ('股票代码', 'Code'),
    'name': ('股票名称', 'Name'),
    'industry': ('申万一级行业', 'Industry'),
    'annual_report_date': ('年报日期', 'AnnualReportDate'),
    'latest_quarter': ('最新季报期', 'LatestQuarter'),
    'data_completeness': ('数据完整度', 'DataCompleteness'),
    'quarter_coverage': ('季度覆盖', 'QuarterCov'),
    'field_gaps': ('缺失字段', 'FieldGaps'),
    
    # 成长性
    'q_revenue_yoy': ('营收同比(%)(单季)', 'RevenueYoY'),
    'q_net_profit_yoy': ('净利润同比(%)(单季)', 'ProfitYoY'),
    'growth': ('成长性', 'Growth'),
    
    # 盈利能力
    'roe_ttm': ('ROE(%)(TTM)', 'ROE_TTM'),
    'gross_margin_ttm': ('毛利率(%)(TTM)', 'GrossMargin_TTM'),
    'profitability': ('盈利能力', 'Profitability'),
    
    # 现金流质量
    'net_profit_ratio': ('净现比', 'OCFtoProfit'),
    'fcf_yield': ('FCF收益率', 'FCFYield'),
    'cash_recovery_rate': ('收现比', 'CashRecoveryRate'),
    'cash_flow_quality': ('现金流质量', 'CashFlow'),
    
    # 偿债风险
    'de_ratio': ('D/E(倍)', 'DebtEquityRatio'),
    'current_ratio': ('流动比率(倍)', 'CurrentRatio'),
    'asset_liability_ratio': ('资产负债率(%)(单季)', 'AssetLiabilityRatio'),
    'leverage_risk': ('偿债风险', 'Leverage'),
    
    # 估值
    'pe_ttm': ('市盈率TTM', 'PE_TTM'),
    'total_mv': ('总市值(亿)', 'MarketValue'),
    
    # 综合
    'total_score': ('总分', 'TotalScore'),
    'rating': ('评级', 'Rating'),
    'confidence': ('置信度', 'Confidence'),
    
    # 否决
    'veto': ('触发否决', 'Veto'),
    'veto_reason': ('否决原因', 'VetoReason'),
}


def get_rating(score: float) -> str:
    """根据总分返回评级"""
    if score >= 80:
        return 'A+'
    elif score >= 70:
        return 'A'
    elif score >= 60:
        return 'B+'
    elif score >= 50:
        return 'B'
    elif score >= 40:
        return 'C'
    else:
        return 'D'


def get_confidence(score: float, completeness: float) -> str:
    """置信度评估"""
    if completeness < 80:
        return '低（数据不足）'
    elif score >= 70:
        return '高'
    elif score >= 50:
        return '中'
    else:
        return '中（需验证）'


def _mock_evaluate(symbol: str) -> Optional[Dict]:
    """模拟评估（通用，随机生成真实感数据）"""
    import random
    from scorer import Scorer
    
    # 基础行业属性
    industry_defaults = {
        '食品饮料': {'roe': (20, 35), 'gm': (40, 95), 'nm': (10, 55)},
        '电子': {'roe': (10, 25), 'gm': (25, 50), 'nm': (5, 20)},
        '医药': {'roe': (12, 25), 'gm': (50, 80), 'nm': (10, 30)},
        '电力': {'roe': (8, 18), 'gm': (20, 40), 'nm': (5, 15)},
        '钢铁': {'roe': (2, 15), 'gm': (5, 20), 'nm': (-5, 10)},
        '房地产': {'roe': (5, 20), 'gm': (15, 35), 'nm': (3, 15)},
    }
    
    # 根据代码规律推断行业和名称
    prefix = symbol[:3]
    if prefix in ['600', '601', '603', '605']:
        industry = random.choice(list(industry_defaults.keys()))
    elif prefix in ['000', '001']:
        industry = random.choice(['电子', '医药', '电力', '房地产'])
    elif prefix in ['002']:
        industry = random.choice(['电子', '医药', '化工', '机械'])
    elif prefix in ['300']:
        industry = random.choice(['电子', '医药', '软件', '化工'])
    elif prefix in ['301']:
        industry = random.choice(['软件', '医药', '电子', '化工'])
    else:
        industry = random.choice(list(industry_defaults.keys()))
    
    # 生成模拟股票名称
    name = f"{industry[:2]}{symbol[-2:]}"
    
    ranges = industry_defaults.get(industry, industry_defaults['电子'])
    
    # 生成数据
    roe = round(random.uniform(*ranges['roe']), 2)
    gm = round(random.uniform(*ranges['gm']), 2)
    rev_yoy = round(random.uniform(-10, 50), 2)
    profit_yoy = round(random.uniform(-20, 80), 2)
    alr = round(random.uniform(20, 80), 2)
    de = round(random.uniform(0.1, 2.0), 2)
    cr = round(random.uniform(0.8, 4.0), 2)
    ocf_ratio = round(random.uniform(50, 150), 2)
    profit = round(random.uniform(1, 500), 2)
    ocf = round(profit * random.uniform(0.5, 1.5), 2)
    pe = round(random.uniform(10, 60), 2)
    mv = round(random.uniform(50, 5000), 2)
    
    # 特殊处理高风险
    if symbol == '999999':
        roe, gm = 5.2, 15.5
        rev_yoy, profit_yoy = -10.5, -25.0
        alr, de, cr = 92.5, 12.5, 0.45
        ocf_ratio, profit, ocf = -45.0, -20.5, -50.0
        industry = '钢铁'
    
    # 一票否决
    veto = '否'
    veto_reason = ''
    if ocf < 0:
        veto = '是'
        veto_reason = '经营现金流为负'
    elif de > 3.0:
        veto = '是'
        veto_reason = f'D/E={de:.1f}超过阈值'
    elif alr > 90:
        veto = '是'
        veto_reason = f'资产负债率={alr:.1f}%超过90%'
    
    # MockCalc
    class MockCalc:
        roe_ttm = roe
        gross_margin_ttm = gm
        de_ratio = de
        asset_liability_ratio = alr
        ocf_ttm = ocf * 100000000
        _ttm_capex = ocf * 20000000
        fcf_ttm = (ocf * 100000000) - (ocf * 20000000)
        net_profit_ratio = ocf_ratio
        cash_recovery_rate = 1.1
        q_net_profit_yoy = profit_yoy
        q_revenue_yoy = rev_yoy
        current_ratio = cr
        net_profit_ttm = profit * 100000000
    
    class MockQuote(dict):
        def __init__(self):
            super().__init__()
            self['total_mv'] = mv * 100000000
            self['pe_ttm'] = pe
    
    calc = MockCalc()
    quote = MockQuote()
    scorer = Scorer(calc, quote)
    scores = scorer.total_score()
    
    if veto == '是':
        scores['total_score'] = 0.0
        scores['veto'] = True
        scores['veto_reason'] = veto_reason
    
    return {
        'symbol': symbol,
        'name': symbol,
        'industry': industry,
        # 成长性
        'q_revenue_yoy': rev_yoy,
        'q_net_profit_yoy': profit_yoy,
        # 盈利能力
        'roe_ttm': roe,
        'gross_margin_ttm': gm,
        # 现金流质量
        'net_profit_ttm': profit,
        'ocf_ttm': ocf,
        'capex_ttm': ocf * 0.2,
        'fcf_ttm': ocf * 0.8,
        'net_profit_ratio': ocf_ratio,
        'fcf_yield': round((ocf * 0.8) / mv if mv > 0 else 0, 4),
        'cash_recovery_rate': 1.1,
        # 偿债风险
        'de_ratio': de,
        'current_ratio': cr,
        'asset_liability_ratio': alr,
        # 估值
        'pe_ttm': pe,
        'total_mv': mv,
        # 基础信息
        'annual_report_date': '2024-04-27',
        'latest_quarter': '2024-12-31',
        'data_completeness': 100,
        'quarter_coverage': '4/4',
        'field_gaps': '无',
        # 综合评分
        'total_score': scores['total_score'],
        'rating': get_rating(scores['total_score']),
        'confidence': get_confidence(scores['total_score'], 100),
        'profitability': round(scores['profitability'], 2),
        'growth': round(scores['growth'], 2),
        'cash_flow_quality': round(scores['cash_flow'], 2),
        'leverage_risk': round(scores['leverage'], 2),
        # 否决
        'veto': veto,
        'veto_reason': veto_reason,
    }


def batch_evaluate(symbols: List[str], eval_date: datetime, mock: bool = True, db=None, 
                     workers: int = 5, rate_limit: float = 0.3) -> List[Dict]:
    """批量评估（并发）"""
    global _eval_date
    _eval_date = eval_date
    
    print(f"\n评估日期: {eval_date.date()}")
    print(f"股票数量: {len(symbols)}")
    print(f"并发数: {workers}")
    print(f"模式: {'模拟测试' if mock else '真实数据'}")
    print("-" * 50)

    results = []
    lock = threading.Lock()
    semaphore = threading.Semaphore(workers)

    def evaluate_one(symbol):
        semaphore.acquire()
        try:
            if mock:
                res = _mock_evaluate(symbol)
            else:
                try:
                    provider = DataProvider()

                    name = provider.get_stock_name(symbol)
                    industry = provider.get_industry(symbol)
                    df = provider.get_combined_financials(symbol)

                    if df.empty:
                        return 'no_data'

                    df = df[df['ann_date'] <= pd.Timestamp(eval_date)].copy()
                    if len(df) < 4:
                        return 'no_data'

                    calc = IndicatorCalculator(df, eval_date=pd.Timestamp(eval_date))
                    quote = provider.get_stock_quote(symbol)
                    scorer = Scorer(calc, quote, industry=industry)
                    scores = scorer.total_score()
                    
                    # 数据完整度详情
                    comp = calc.get_completeness_info()

                    res = {
                        'symbol': symbol, 'name': name, 'industry': industry,
                        'q_revenue_yoy': round(calc.q_revenue_yoy, 2),
                        'q_net_profit_yoy': round(calc.q_net_profit_yoy, 2),
                        'roe_ttm': round(calc.roe_ttm, 2),
                        'gross_margin_ttm': round(calc.gross_margin_ttm, 2),
                        'net_profit_ratio': round(calc.net_profit_ratio, 2),
                        'fcf_yield': round(calc.fcf_ttm / quote.get('total_mv', 1) if quote.get('total_mv', 0) > 0 else 0, 4),
                        'cash_recovery_rate': round(calc.cash_recovery_rate, 2),
                        'de_ratio': round(calc.de_ratio, 2),
                        'current_ratio': round(calc.current_ratio, 2),
                        'asset_liability_ratio': round(calc.asset_liability_ratio, 2),
                        'pe_ttm': round(quote.get('pe_ttm', 0), 2),
                        'total_mv': round(quote.get('total_mv', 0) / 100000000, 2),
                        'annual_report_date': str(df[df['report_date'].dt.month == 12]['report_date'].max().date()) if len(df[df['report_date'].dt.month == 12]) > 0 else str(df['report_date'].max().date()),
                        'latest_quarter': str(df['report_date'].max().date()),
                        'data_completeness': comp['score'],
                        'quarter_coverage': comp['quarter_coverage'],
                        'field_gaps': comp['field_gaps'],
                        'total_score': scores['total_score'],
                        'rating': get_rating(scores['total_score']),
                        'confidence': get_confidence(scores['total_score'], comp['score']),
                        'profitability': round(scores['profitability'], 2),
                        'growth': round(scores['growth'], 2),
                        'cash_flow_quality': round(scores['cash_flow'], 2),
                        'leverage_risk': round(scores['leverage'], 2),
                        'veto': '是' if scores['veto'] else '否',
                        'veto_reason': scores.get('veto_reason', ''),
                    }
                except Exception as e:
                    print(f"\n  [ERROR] {symbol}: {e}")
                    return 'error'

            with lock:
                results.append(res)
                print(f"\r[{len(results)}/{len(symbols)}] {symbol} 总分={res.get('total_score', 0):.1f} 评级={res.get('rating', '-')}", end='', flush=True)

            return 'ok'
        finally:
            semaphore.release()
            if not mock:
                time.sleep(rate_limit)

    # 并发执行
    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(evaluate_one, s): s for s in symbols}
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"\n  [EXCEPTION] {futures[future]}: {e}")

    print()  # 换行
    return results


def save_to_excel(results: List[Dict], filename: str = None):
    """保存为 Excel（默认带时间戳，自动保留最近10份）"""
    if not results:
        print("无结果可保存")
        return

    df = pd.DataFrame(results)
    df = df.sort_values('total_score', ascending=False)
    
    # 默认文件名带时间戳
    if filename is None:
        filename = f"评分结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 自动清理旧文件：保留最近10份
    output_dir = os.path.dirname(os.path.abspath(filename))
    output_name = os.path.basename(filename)
    prefix = '评分结果_'
    if output_name.startswith(prefix):
        old_files = sorted(
            [f for f in os.listdir(output_dir) if f.startswith(prefix) and f.endswith('.xlsx')],
            reverse=True  # 最新的排前面
        )
        # 保留最近10份，删除更早的
        for old_f in old_files[10:]:
            try:
                os.remove(os.path.join(output_dir, old_f))
            except OSError:
                pass
    
    # 排序列顺序 - 按维度分组
    order = [
        # 基础信息（代码/名称）
        'symbol', 'name',
        # 综合评分（前置，紧跟代码）
        'total_score', 'rating', 'confidence',
        # 行业
        'industry',
        # 成长性（蓝）
        'q_revenue_yoy', 'q_net_profit_yoy', 'growth',
        # 盈利能力（绿）
        'roe_ttm', 'gross_margin_ttm', 'profitability',
        # 现金流质量（橙）
        'net_profit_ratio', 'fcf_yield', 'cash_recovery_rate', 'cash_flow_quality',
        # 偿债风险（红）
        'de_ratio', 'current_ratio', 'asset_liability_ratio', 'leverage_risk',
        # 估值（紫）
        'pe_ttm', 'total_mv',
        # 基础信息（日期/完整度，后置）
        'annual_report_date', 'latest_quarter', 'data_completeness', 'quarter_coverage', 'field_gaps',
        # 否决
        'veto', 'veto_reason',
    ]
    
    cols = [c for c in order if c in df.columns]
    df = df[cols]
    
    # Excel 样式
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '评分结果'
    
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 维度分组背景色
    DIMENSION_FILLS = {
        'info': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),        # 白色
        'growth': PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),      # 蓝色
        'profitability': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'), # 绿色
        'cashflow': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),    # 橙色
        'leverage': PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),   # 深橙
        'valuation': PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),    # 紫色
        'total': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),       # 金色
        'veto': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),         # 灰色
    }
    
    # 列到维度的映射（统一完整定义，避免重复和遗漏）
    COL_DIMENSIONS = {
        # 基础信息
        'symbol': 'info', 'name': 'info', 'industry': 'info',
        'annual_report_date': 'info', 'latest_quarter': 'info', 'data_completeness': 'info',
        'quarter_coverage': 'info', 'field_gaps': 'info',
        # 综合评分
        'total_score': 'total', 'rating': 'total', 'confidence': 'total',
        # 成长性
        'q_revenue_yoy': 'growth', 'q_net_profit_yoy': 'growth', 'growth': 'growth',
        # 盈利能力
        'roe_ttm': 'profitability', 'gross_margin_ttm': 'profitability', 'profitability': 'profitability',
        # 现金流质量
        'net_profit_ratio': 'cashflow', 'fcf_yield': 'cashflow',
        'cash_recovery_rate': 'cashflow', 'cash_flow_quality': 'cashflow',
        'net_profit_ttm': 'cashflow', 'ocf_ttm': 'cashflow',
        # 偿债风险
        'de_ratio': 'leverage', 'current_ratio': 'leverage',
        'asset_liability_ratio': 'leverage', 'leverage_risk': 'leverage',
        # 估值
        'pe_ttm': 'valuation', 'total_mv': 'valuation',
        # 否决
        'veto': 'veto', 'veto_reason': 'veto',
    }
    
    # 表头（中英文）- 按维度着色
    headers = []
    header_fills = []
    for col in df.columns:
        cn, en = EXCEL_COLUMNS.get(col, (col, col))
        headers.append(f"{cn}\n({en})")
        dim = COL_DIMENSIONS.get(col, 'info')
        header_fills.append(DIMENSION_FILLS[dim])
    
    for col_idx, (header, hfill) in enumerate(zip(headers, header_fills), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hfill
        cell.font = Font(bold=True, size=10)
        cell.alignment = header_align
        cell.border = thin_border
    
    # 评级颜色方案
    RATING_COLORS = {
        'A+': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
        'A': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
        'B+': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
        'B': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
        'C': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        'D': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
    }
    
    # 数据行
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            col_name = df.columns[col_idx - 1]
            dim = COL_DIMENSIONS.get(col_name, 'info')
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            
            # 同维度同背景色
            cell.fill = DIMENSION_FILLS[dim]
            
            # 评级特殊处理
            if col_name == 'rating' and value in RATING_COLORS:
                cell.fill = RATING_COLORS[value]
                # 深色背景用白色字体，浅色背景用黑色字体
                if value in ['A+', 'A', 'D']:
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.font = Font(bold=True, color='000000')
            # 数据完整度条件着色
            elif col_name == 'data_completeness':
                if isinstance(value, (int, float)):
                    if value < 60:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                    elif value < 80:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        cell.font = Font(bold=True)
                    else:
                        cell.font = Font(bold=True)
            elif col_name == 'quarter_coverage':
                # 非完整覆盖时高亮
                if isinstance(value, str) and '缺' in value:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.font = Font(bold=True)
            elif col_name == 'field_gaps':
                # 有缺失字段时高亮
                if isinstance(value, str) and value != '无':
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.font = Font(bold=True)
            elif col_name == 'total_score':
                cell.font = Font(bold=True)
    
    # 列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 28)
    
    ws.freeze_panes = 'A2'
    wb.save(filename)
    print(f"\n结果已保存: {filename}")


def main():
    # ---- 全局日志配置（UTF-8 文件 + 控制台） ----
    import sys as _sys
    _sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    _sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    
    # 创建 logs 目录
    _log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    os.makedirs(_log_dir, exist_ok=True)
    _log_ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    _log_path = os.path.join(_log_dir, f'run_{_log_ts}.log')
    
    # 配置 root logger：同时写文件(UTF-8)和控制台
    _root_logger = logging.getLogger()
    _root_logger.setLevel(logging.INFO)
    _root_logger.handlers.clear()
    _file_handler = logging.FileHandler(_log_path, encoding='utf-8')
    _file_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S'))
    _console_handler = logging.StreamHandler(sys.stdout)
    _console_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S'))
    _root_logger.addHandler(_file_handler)
    _root_logger.addHandler(_console_handler)
    
    logging.info(f"日志文件: {_log_path}")
    
    parser = argparse.ArgumentParser(description='A股智能选股系统 v3.1.2')
    parser.add_argument('--stocks', nargs='+', help='股票代码列表')
    parser.add_argument('--pool', type=str, default='stock_pool.txt', help='股票池文件')
    parser.add_argument('--workers', type=int, default=5, help='并发线程数')
    parser.add_argument('--rate-limit', type=float, default=0.3, help='请求间隔(秒)')
    parser.add_argument('--date', type=str, help='评估日期 (YYYY-MM-DD)')
    parser.add_argument('--output', type=str, help='输出Excel文件名')
    parser.add_argument('--disable-veto', action='store_true', help='禁用否决机制')
    parser.add_argument('--mock', action='store_true', help='模拟模式（无需网络）')
    parser.add_argument('--real', action='store_true', help='真实数据模式')
    parser.add_argument('--db', type=str, default='stock_data.db', help='数据库路径')
    parser.add_argument('--save-db', action='store_true', help='保存结果到数据库')
    parser.add_argument('--from-db', action='store_true', help='从数据库读取数据评分')
    parser.add_argument('--fetch-only', action='store_true', help='仅获取数据存DB，不评分')
    args = parser.parse_args()

    if args.disable_veto:
        VETO_RULES['enable'] = False
        print("[提示] 一票否决机制已禁用")

    # 股票池
    if args.stocks:
        symbols = args.stocks
    else:
        from utils import load_stock_pool
        symbols = load_stock_pool(args.pool)
    
    if not symbols:
        print("股票池为空")
        return

    # 评估日期
    if args.date:
        eval_date = datetime.strptime(args.date, '%Y-%m-%d')
    else:
        eval_date = datetime.today()

    # 数据库
    db = None
    if args.save_db or args.from_db or args.fetch_only:
        from database import StockDatabase
        db = StockDatabase(args.db)
        print(f"[数据库] {args.db}")

    # ========== 模式1：从DB读取数据评分 ==========
    if args.from_db:
        print(f"\n从数据库读取数据评分...")
        results = []
        for i, symbol in enumerate(symbols, 1):
            print(f"\r[{i}/{len(symbols)}] {symbol}...", end='', flush=True)
            df = db.get_financials(symbol)
            if df is None or df.empty or len(df) < 4:
                continue
            
            from calculator import IndicatorCalculator
            
            # DB读取后日期列可能是字符串，确保转datetime
            if not pd.api.types.is_datetime64_any_dtype(df['report_date']):
                df['report_date'] = pd.to_datetime(df['report_date'], errors='coerce')
            if 'ann_date' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['ann_date']):
                df['ann_date'] = pd.to_datetime(df['ann_date'], errors='coerce')
            
            df = df[df['ann_date'] <= pd.Timestamp(eval_date)].copy()
            if len(df) < 4:
                continue
            
            # 从DB读取股票信息（不再调用API）
            stock_info = db.get_stock_info(symbol)
            name = stock_info.get('name', symbol) if stock_info else symbol
            industry = stock_info.get('industry', None) if stock_info else None
            quote = db.get_quote(symbol)
            
            calc = IndicatorCalculator(df, eval_date=pd.Timestamp(eval_date))
            scorer = Scorer(calc, quote, industry=industry)
            scores = scorer.total_score()
            
            # 数据完整度详情
            comp = calc.get_completeness_info()
            
            res = {
                'symbol': symbol, 'name': name, 'industry': industry or '',
                'q_revenue_yoy': round(calc.q_revenue_yoy, 2),
                'q_net_profit_yoy': round(calc.q_net_profit_yoy, 2),
                'roe_ttm': round(calc.roe_ttm, 2),
                'gross_margin_ttm': round(calc.gross_margin_ttm, 2),
                'net_profit_ratio': round(calc.net_profit_ratio, 2),
                'fcf_yield': round(calc.fcf_ttm / quote.get('total_mv', 1) if quote.get('total_mv', 0) > 0 else 0, 4),
                'cash_recovery_rate': round(calc.cash_recovery_rate, 2),
                'de_ratio': round(calc.de_ratio, 2),
                'current_ratio': round(calc.current_ratio, 2),
                'asset_liability_ratio': round(calc.asset_liability_ratio, 2),
                'pe_ttm': round(quote.get('pe_ttm', 0), 2),
                'total_mv': round(quote.get('total_mv', 0) / 100000000, 2),
                'annual_report_date': str(df[df['report_date'].dt.month == 12]['report_date'].max().date()) if len(df[df['report_date'].dt.month == 12]) > 0 else str(df['report_date'].max().date()) if hasattr(df['report_date'].max(), 'date') else str(df['report_date'].max()),
                'latest_quarter': str(df['report_date'].max().date()) if hasattr(df['report_date'].max(), 'date') else str(df['report_date'].max()),
                'data_completeness': comp['score'],
                'quarter_coverage': comp['quarter_coverage'],
                'field_gaps': comp['field_gaps'],
                'total_score': scores['total_score'],
                'rating': get_rating(scores['total_score']),
                'confidence': get_confidence(scores['total_score'], comp['score']),
                'profitability': round(scores['profitability'], 2),
                'growth': round(scores['growth'], 2),
                'cash_flow_quality': round(scores['cash_flow'], 2),
                'leverage_risk': round(scores['leverage'], 2),
                'veto': '是' if scores['veto'] else '否',
                'veto_reason': scores.get('veto_reason', ''),
            }
            results.append(res)
        
        print(f"\n完成，共 {len(results)} 条结果")
        if results:
            save_to_excel(results, args.output or None)
        return

    # ========== 模式2：获取数据并存DB（并发+断点续传+完整日志） ==========
    if args.real and db:
        import time as _time
        t_batch_start = _time.time()
        
        # ---- 日志文件（纯TXT，UTF-8，绝不乱码） ----
        log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
        os.makedirs(log_dir, exist_ok=True)
        log_ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = os.path.join(log_dir, f'fetch_{log_ts}.log')
        fail_file = os.path.join(log_dir, f'failed_stocks_{log_ts}.txt')
        report_file = os.path.join(log_dir, f'fetch_report_{log_ts}.txt')
        
        def log_write(msg, to_console=True):
            """写日志文件（UTF-8）+ 可选控制台输出"""
            ts = datetime.now().strftime('%H:%M:%S')
            line = f"[{ts}] {msg}"
            if to_console:
                try:
                    print(line)
                except UnicodeEncodeError:
                    print(line.encode('ascii', errors='replace').decode('ascii'))
            with open(log_file, 'a', encoding='utf-8') as lf:
                lf.write(line + '\n')
        
        log_write('='*60)
        log_write(f"数据获取启动 | workers={args.workers} rate_limit={args.rate_limit}s")
        log_write(f"日志文件: {log_file}")
        log_write('='*60)
        
        # 预加载全市场行情缓存
        from data_provider import preload_market_data
        preload_ok = preload_market_data()
        log_write(f"[预加载] {'成功' if preload_ok else '失败（将逐只降级获取）'}")
        
        from concurrent.futures import ThreadPoolExecutor, as_completed
        
        # ---- 增量更新逻辑 ----
        # 分类：full_fetch（DB无数据/不足4季） vs incremental（DB有数据，追加新报告期）
        full_fetch, incremental_fetch = db.get_stale_symbols(symbols, min_quarters=4)
        log_write(f"[增量分析] 总池 {len(symbols)} 只 | 全量获取 {len(full_fetch)} 只 | 增量更新 {len(incremental_fetch)} 只")
        
        # 合并获取列表（全量+增量都需要拉取API）
        fetch_symbols = full_fetch + incremental_fetch
        # 记录每只股票的获取模式
        fetch_mode = {s: 'full' for s in full_fetch}
        fetch_mode.update({s: 'incr' for s in incremental_fetch})
        
        if not fetch_symbols:
            log_write("所有股票数据已存在且最新，无需获取")
        else:
            # 统计计数器
            success_count = 0
            fail_count = 0
            empty_count = 0
            new_count = 0       # 增量：新增的报告期行数
            update_count = 0    # 增量：更新的报告期行数
            timeout_count = 0
            network_err_count = 0
            parse_err_count = 0
            other_err_count = 0
            lock = threading.Lock()
            info_batch = []
            failed_symbols = []
            failed_details = []
            
            def classify_error(e):
                ename = type(e).__name__
                emsg = str(e).lower()
                if 'timeout' in ename.lower() or 'timeout' in emsg:
                    return 'timeout'
                elif any(k in emsg for k in ['connection', 'refused', 'reset', 'disconnected', 'network', 'remote']):
                    return 'network'
                elif any(k in emsg for k in ['json', 'parse', 'decode', 'keyerror', 'indexerror']):
                    return 'parse'
                else:
                    return 'other'
            
            def fetch_one(symbol):
                nonlocal success_count, fail_count, empty_count
                nonlocal timeout_count, network_err_count, parse_err_count, other_err_count
                t0 = _time.time()
                mode = fetch_mode.get(symbol, 'full')
                mode_tag = '[全量]' if mode == 'full' else '[增量]'
                try:
                    from data_provider import DataProvider
                    provider = DataProvider()
                    
                    # 获取财务数据（API始终返回全量历史）
                    df = provider.get_combined_financials(symbol)
                    if df is not None and not df.empty:
                        # 增量模式：只保存DB中不存在的新报告期
                        if mode == 'incr':
                            existing_dates = set()
                            conn = sqlite3.connect(db.db_path)
                            cur = conn.cursor()
                            cur.execute('SELECT report_date FROM financials WHERE symbol = ?', (symbol,))
                            existing_dates = {r[0] for r in cur.fetchall()}
                            conn.close()
                            new_df = df[~df['report_date'].astype(str).isin(existing_dates)]
                            if new_df.empty:
                                # 没有新数据，也算成功（已是最新）
                                elapsed = _time.time() - t0
                                with lock:
                                    success_count += 1
                                    done = success_count + fail_count + empty_count
                                    pct = done / len(fetch_symbols) * 100
                                    print(f"\r  [{done}/{len(fetch_symbols)}] {pct:.0f}% | OK:{success_count} FAIL:{fail_count} EMPTY:{empty_count} | {mode_tag} {symbol} 已是最新 {elapsed:.1f}s    ", end='', flush=True)
                                _time.sleep(args.rate_limit)
                                return
                            # 有新数据，保存新增行
                            rows_saved = db.save_financials(symbol, new_df)
                        else:
                            # 全量模式：直接保存全部
                            rows_saved = db.save_financials(symbol, df)
                        
                        name = provider.get_stock_name(symbol)
                        industry = provider.get_industry(symbol)
                        quote = provider.get_stock_quote(symbol)
                        db.save_quote(symbol, quote)
                        
                        elapsed = _time.time() - t0
                        with lock:
                            success_count += 1
                            info_batch.append((symbol, name, industry))
                            done = success_count + fail_count + empty_count
                            pct = done / len(fetch_symbols) * 100
                            rows_info = f" {rows_saved}行" if mode == 'incr' else ""
                            print(f"\r  [{done}/{len(fetch_symbols)}] {pct:.0f}% | OK:{success_count} FAIL:{fail_count} EMPTY:{empty_count} | {mode_tag} {symbol} {name}{rows_info} {elapsed:.1f}s    ", end='', flush=True)
                    else:
                        elapsed = _time.time() - t0
                        with lock:
                            empty_count += 1
                            fail_count += 1
                            failed_symbols.append(symbol)
                            failed_details.append(f"{symbol}\t{mode_tag}\t数据为空\t{elapsed:.1f}s")
                            done = success_count + fail_count + empty_count
                            pct = done / len(fetch_symbols) * 100
                            print(f"\r  [{done}/{len(fetch_symbols)}] {pct:.0f}% | OK:{success_count} FAIL:{fail_count} EMPTY:{empty_count} | {mode_tag} {symbol} 数据为空 {elapsed:.1f}s    ", end='', flush=True)
                            
                except Exception as e:
                    elapsed = _time.time() - t0
                    err_type = classify_error(e)
                    with lock:
                        fail_count += 1
                        if err_type == 'timeout':
                            timeout_count += 1
                        elif err_type == 'network':
                            network_err_count += 1
                        elif err_type == 'parse':
                            parse_err_count += 1
                        else:
                            other_err_count += 1
                        failed_symbols.append(symbol)
                        failed_details.append(f"{symbol}\t{mode_tag}\t{type(e).__name__}: {str(e)[:100]}\t{elapsed:.1f}s")
                        done = success_count + fail_count + empty_count
                        pct = done / len(fetch_symbols) * 100
                        print(f"\r  [{done}/{len(fetch_symbols)}] {pct:.0f}% | OK:{success_count} FAIL:{fail_count} EMPTY:{empty_count} | {mode_tag} {symbol} [{type(e).__name__}] {elapsed:.1f}s    ", end='', flush=True)
                
                _time.sleep(args.rate_limit)
            
            # ---- 并发执行 ----
            log_write(f"开始并发获取 {len(fetch_symbols)} 只 (workers={args.workers})...")
            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                futures = {executor.submit(fetch_one, s): s for s in fetch_symbols}
                for f in as_completed(futures):
                    try:
                        f.result()
                    except Exception as e:
                        sym = futures[f]
                        log_write(f"  [FATAL] {sym}: {type(e).__name__}: {e}")
                        with lock:
                            fail_count += 1
                            failed_symbols.append(sym)
                            failed_details.append(f"{sym}\tFATAL\t{type(e).__name__}: {str(e)[:100]}")
            
            print()  # 换行
            
            # 批量保存股票信息
            if info_batch:
                db.save_stock_info_batch(info_batch)
                log_write(f"股票信息批量保存完成: {len(info_batch)} 条")
            
            t_batch_elapsed = _time.time() - t_batch_start
            avg_time = t_batch_elapsed / len(fetch_symbols) if fetch_symbols else 0
            
            # ---- 最终汇总报告（纯TXT） ----
            rate_str = f"{success_count/len(fetch_symbols)*100:.1f}%" if fetch_symbols else "N/A"
            
            report_lines = [
                '',
                '='*60,
                '数据获取汇总报告',
                '='*60,
                f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"总池: {len(symbols)} 只",
                f"  - 全量获取: {len(full_fetch)} 只",
                f"  - 增量更新: {len(incremental_fetch)} 只",
                f"实际获取: {len(fetch_symbols)} 只",
                f"",
                f"结果统计:",
                f"  成功:   {success_count} 只 ({rate_str})",
                f"  失败:   {fail_count} 只",
                f"    - 数据为空: {empty_count}",
                f"    - 网络错误: {network_err_count}",
                f"    - 超时:     {timeout_count}",
                f"    - 解析错误: {parse_err_count}",
                f"    - 其他错误: {other_err_count}",
                f"",
                f"耗时: {t_batch_elapsed:.0f}s ({t_batch_elapsed/60:.1f}min) | 平均: {avg_time:.1f}s/只",
                f"日志: {log_file}",
            ]
            
            if failed_symbols:
                report_lines.append(f"失败列表: {fail_file} ({len(failed_symbols)} 只)")
                report_lines.append('')
                report_lines.append('失败明细:')
                report_lines.append(f"{'代码':<10} {'模式':<8} {'原因':<30} {'耗时':<8}")
                report_lines.append('-'*60)
                for d in failed_details:
                    report_lines.append(d)
            
            report_lines.append('='*60)
            
            # 写汇总到日志
            for l in report_lines:
                log_write(l)
            
            # 写独立报告文件
            with open(report_file, 'w', encoding='utf-8') as rf:
                rf.write('\n'.join(report_lines))
            
            # 写失败股票列表（纯代码，方便重跑）
            if failed_symbols:
                with open(fail_file, 'w', encoding='utf-8') as ff:
                    ff.write('# 失败股票列表，可复制到 --pool 参数重跑\n')
                    ff.write(f'# 生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
                    ff.write(f'# 共 {len(failed_symbols)} 只\n')
                    for s in failed_symbols:
                        ff.write(s + '\n')
        
        if args.fetch_only:
            stats = db.get_db_stats()
            log_write(f"\n[DB统计] 财务记录: {stats['financials']} | 股票数: {stats['symbols_with_fin']} | DB大小: {stats['db_size_mb']}MB")
            
            # 提示重跑命令
            if failed_symbols:
                log_write(f"\n[提示] 重跑失败股票:")
                log_write(f"  python main.py --real --pool {fail_file} --workers 8 --rate-limit 0.1 --save-db --fetch-only --db {args.db}")
            return

    # ========== 模式3：评分（从DB或实时） ==========
    if args.real and not db:
        # 实时获取+评分（原流程）
        results = batch_evaluate(symbols, eval_date, mock=False, db=db,
                                 workers=args.workers, rate_limit=args.rate_limit)
    else:
        # 从DB读取财务数据评分
        print(f"\n从数据库读取数据评分...")
        results = []
        for i, symbol in enumerate(symbols, 1):
            print(f"\r[{i}/{len(symbols)}] {symbol}...", end='', flush=True)
            
            if db:
                df = db.get_financials(symbol)
            else:
                df = None
            
            if df is None or df.empty:
                # 回退到实时获取
                try:
                    from data_provider import DataProvider
                    provider = DataProvider()
                    df = provider.get_combined_financials(symbol)
                    if db and df is not None and not df.empty:
                        db.save_financials(symbol, df)
                except:
                    pass
            
            if df is None or df.empty or len(df) < 4:
                continue
            
            from calculator import IndicatorCalculator
            
            df = df[df['ann_date'] <= pd.Timestamp(eval_date)].copy()
            if len(df) < 4:
                continue
            
            # 优先从DB读取，DB没有则回退到API
            if db:
                stock_info = db.get_stock_info(symbol)
                name = stock_info.get('name', symbol) if stock_info else symbol
                industry = stock_info.get('industry', None) if stock_info else None
                quote = db.get_quote(symbol)
                # 如果DB中没有quote数据，回退到API
                if quote.get('total_mv', 0) == 0:
                    try:
                        from data_provider import DataProvider
                        provider = DataProvider()
                        quote = provider.get_stock_quote(symbol)
                    except:
                        pass
            else:
                from data_provider import DataProvider
                provider = DataProvider()
                name = provider.get_stock_name(symbol)
                industry = provider.get_industry(symbol)
                quote = provider.get_stock_quote(symbol)
            
            calc = IndicatorCalculator(df, eval_date=pd.Timestamp(eval_date))
            scorer = Scorer(calc, quote, industry=industry)
            scores = scorer.total_score()
            
            # 数据完整度详情
            comp = calc.get_completeness_info()
            
            res = {
                'symbol': symbol, 'name': name, 'industry': industry,
                'q_revenue_yoy': round(calc.q_revenue_yoy, 2),
                'q_net_profit_yoy': round(calc.q_net_profit_yoy, 2),
                'roe_ttm': round(calc.roe_ttm, 2),
                'gross_margin_ttm': round(calc.gross_margin_ttm, 2),
                'net_profit_ratio': round(calc.net_profit_ratio, 2),
                'fcf_yield': round(calc.fcf_ttm / quote.get('total_mv', 1) if quote.get('total_mv', 0) > 0 else 0, 4),
                'cash_recovery_rate': round(calc.cash_recovery_rate, 2),
                'de_ratio': round(calc.de_ratio, 2),
                'current_ratio': round(calc.current_ratio, 2),
                'asset_liability_ratio': round(calc.asset_liability_ratio, 2),
                'pe_ttm': round(quote.get('pe_ttm', 0), 2),
                'total_mv': round(quote.get('total_mv', 0) / 100000000, 2),
                'annual_report_date': str(df[df['report_date'].dt.month == 12]['report_date'].max().date()) if len(df[df['report_date'].dt.month == 12]) > 0 else str(df['report_date'].max().date()) if hasattr(df['report_date'].max(), 'date') else str(df['report_date'].max()),
                'latest_quarter': str(df['report_date'].max().date()) if hasattr(df['report_date'].max(), 'date') else str(df['report_date'].max()),
                'data_completeness': comp['score'],
                'quarter_coverage': comp['quarter_coverage'],
                'field_gaps': comp['field_gaps'],
                'total_score': scores['total_score'],
                'rating': get_rating(scores['total_score']),
                'confidence': get_confidence(scores['total_score'], comp['score']),
                'profitability': round(scores['profitability'], 2),
                'growth': round(scores['growth'], 2),
                'cash_flow_quality': round(scores['cash_flow'], 2),
                'leverage_risk': round(scores['leverage'], 2),
                'veto': '是' if scores['veto'] else '否',
                'veto_reason': scores.get('veto_reason', ''),
            }
            results.append(res)
        
        print()

    # 保存结果
    if results:
        save_to_excel(results, args.output or None)
    else:
        print("无结果可保存")


if __name__ == '__main__':
    main()