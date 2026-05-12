#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""快速验证：打印Excel数据 + 手动对比东方财富"""
import openpyxl

wb = openpyxl.load_workbook(r'D:\Project\AnnualScorer\股票业绩评价_20260426_204545.xlsx')

# 打印各Sheet统计
print("=== Sheet 统计 ===")
for sn in wb.sheetnames:
    ws2 = wb[sn]
    print(f"  {sn}: {ws2.max_row}行")

# 读取综合评价结果
ws = wb["综合评价结果"]
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

# 找目标股票
targets = {"603350.SH": "安乃达", "000852.SZ": "石化机械"}

print(f"\n=== 列头 ===")
print(headers)

for r in range(2, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if code in targets:
        name = ws.cell(r, 2).value
        print(f"\n{'='*60}")
        print(f"股票: {code} {name}")
        print(f"{'='*60}")
        for i, h in enumerate(headers, 1):
            val = ws.cell(r, i).value
            print(f"  {h}: {val}")

# 打印统计概览
print(f"\n=== 统计概览 ===")
ws_stats = wb["统计概览"]
for r in range(1, ws_stats.max_row + 1):
    for c in range(1, ws_stats.max_column + 1):
        val = ws_stats.cell(r, c).value
        if val:
            print(f"  {val}", end="\t")
    print()
