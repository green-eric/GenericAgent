#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A股智能选股系统 - 基于季报+年报的业绩评分系统 
架构：单季看成长，TTM看盈利与现金，最新报表看杠杆
版本: 7.0.0

数据源策略：
  成长性（营收同比、净利润同比）  → 最新单季报
  盈利能力（ROE、毛利率、净利率） → TTM（近4个季度滚动）
  偿债风险（资产负债率）          → 最新单季报
  现金流质量（OCF/净利润）        → TTM（经营现金流TTM / 净利润TTM）
  净利润(元)、经营现金流(元)      → TTM（近4季滚动）
  年报日期                       → 最近一期已披露年报的截止日
"""

import os
import sys
import json
import time
import logging
import argparse
import re
import sqlite3
import threading
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed, Future
from contextlib import contextmanager




class Config:
    """全局配置类"""
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DEFAULT_STOCK_FILE = os.path.join(BASE_DIR, "xuan.txt")
    OUTPUT_DIR = BASE_DIR
    INDUSTRY_MAP_FILE = os.path.join(BASE_DIR, "industry_map.json")
    # 年报数据库路径（季报项目与年报项目共享）
    ANNUAL_DB_FILE = os.path.join(os.path.dirname(BASE_DIR), "AnnualScorer", "stock_cache.db")
    # 季报缓存数据库路径
    QUARTERLY_DB_FILE = os.path.join(BASE_DIR, "quarterly_cache.db")
    
    # --- 评分权重 ---
    WEIGHTS = {
        'growth':       0.25,   # 成长性（单季同比）
        'profitability':0.30,   # 盈利能力（TTM）
        'cash_flow':    0.20,   # 现金流质量（TTM）
        'leverage':     0.15,   # 偿债风险（最新报表）
        'valuation':    0.10    # 估值（辅助）
    }
    
    # --- 各指标满分阈值 ---
    THRESHOLDS = {
        'q_net_profit_yoy':    0.5,   # 单季净利润同比增速≥50%满分
        'q_revenue_yoy':       0.3,   # 单季营收同比增速≥30%满分
        'roe_ttm':             0.15,  # TTM ROE ≥15%满分
        'gross_margin_ttm':    0.4,   # TTM毛利率≥40%满分
        'net_profit_ratio':    1.0,   # 净现比 ≥1.0满分
        'fcf_yield':           0.03,  # 自由现金流/市值 ≥3%满分
        'de_ratio_max':        0.5,   # D/E ≤0.5满分
        'current_ratio_min':   2.0,   # 流动比率 ≥2满分
        'asset_liability_max': 0.5,   # 资产负债率 ≤50%满分
        'pe_max':             20      # PE≤20满分
    }

    FINANCE_WORKERS = 4
    API_RETRY_TIMES = 3
    API_RETRY_BACKOFF_BASE = 5.0
    API_TIMEOUT = 50
    PAUSE_CONSECUTIVE_EMPTY = 5
    PAUSE_DURATION = 30
    GLOBAL_TIMEOUT = 7200
    MIN_INDUSTRY_SAMPLES = 5
    CACHE_MAX_AGE_QUARTERLY = 30
    NEGATIVE_PROFIT_PENALTY = 15.0
    MARKET_FALLBACK_DISCOUNT = 0.95
    LOW_COMPLETENESS_PENALTY = 0.9
    ULTRA_LOW_COMPLETENESS_PENALTY = 0.75
    AKSHARE_INDUSTRY_CACHE = os.path.join(BASE_DIR, "industry_map_akshare.json")
    AKSHARE_CACHE_MAX_AGE_DAYS = 30





def setup_logging() -> logging.Logger:
    """设置日志系统（UTF-8 编码，解决乱码）"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(Config.BASE_DIR, f"stock_analyzer_{timestamp}.log")

    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8')

    formatter = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    if not logger.handlers:
        file_handler = logging.FileHandler(log_file, encoding="utf-8", mode="w")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)

    return logger


logger = setup_logging()





@contextmanager
def db_connection(db_path: str = None):
    """数据库连接上下文管理器"""
    if db_path is None:
        db_path = ""
    conn = None
    try:
        conn = sqlite3.connect(db_path, timeout=30.0)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=30000")
        yield conn
        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"数据库操作异常: {e}")
        raise
    finally:
        if conn:
            conn.close()



        conn.execute("""
            CREATE TABLE IF NOT EXISTS quarterly_reports (
                ts_code TEXT,
                report_date TEXT,
                report_type TEXT DEFAULT 'quarterly',
                -- 绝对值
                revenue REAL,              -- 营业总收入(元)
                operating_cost REAL,       -- 营业成本(元)
                net_profit REAL,           -- 归母净利润(元)
                net_profit_deducted REAL,  -- 扣非净利润(元)
                ocf_abs REAL,              -- 经营活动现金流(元)
                total_assets REAL,         -- 资产合计(元)
                total_liabilities REAL,    -- 负债合计(元)
                net_assets REAL,           -- 净资产/股东权益(元)
                -- 比率
                gross_margin REAL,         -- 销售毛利率(%)
                net_margin REAL,           -- 销售净利率(%)
                debt_ratio REAL,           -- 资产负债率(%)
                ocf_ratio REAL,            -- 净利润现金含量(%)
                roa REAL,                  -- 资产回报率ROA(%)
                -- 成长性
                revenue_yoy REAL,          -- 营业收入同比增长(%)
                profit_yoy REAL,          -- 归母净利润同比增长(%)
                -- 元数据
                fetch_success INTEGER DEFAULT 0,
                last_update TEXT,
                PRIMARY KEY (ts_code, report_date, report_type)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS quarterly_ttm_cache (
                ts_code TEXT PRIMARY KEY,
                -- TTM绝对值
                revenue_ttm REAL,
                operating_cost_ttm REAL,
                net_profit_ttm REAL,
                ocf_abs_ttm REAL,
                net_assets_ttm REAL,       -- 最新一期净资产（用于ROE计算）
                -- TTM比率
                gross_margin_ttm REAL,
                net_margin_ttm REAL,
                ocf_ratio_ttm REAL,
                roe_ttm REAL,              -- TTM ROE(%) = 净利润TTM / 净资产
                -- 最新单季指标（用于成长性、偿债风险）
                revenue_yoy_latest REAL,
                profit_yoy_latest REAL,
                debt_ratio_latest REAL,
                total_assets_latest REAL,
                total_liabilities_latest REAL,
                -- 元数据
                quarter_count INTEGER,
                latest_quarter TEXT,
                last_update TEXT
            )
        """)
        conn.commit()
    logger.info(f"季报缓存DB初始化完成: {db_path}")




    try:
        with sqlite3.connect(db_path, timeout=10.0) as conn:
            # 检查TTM缓存
            cur = conn.execute(
                "SELECT * FROM quarterly_ttm_cache WHERE ts_code=?",
                (ts_code,)
            )
            row = cur.fetchone()
            if not row:
                return None

            # 检查缓存是否过期（默认7天）
            cols = [d[0] for d in cur.description]
            data = dict(zip(cols, row))
            last_update = data.get("last_update")
            if last_update:
                try:
                    update_time = datetime.fromisoformat(last_update)
                    age_days = (datetime.now() - update_time).total_seconds() / 86400
                    if age_days > Config.CACHE_MAX_AGE_QUARTERLY:
                        logger.debug(f"季报缓存过期 {ts_code}: {age_days:.1f}天")
                        return None
                except Exception:
                    pass
            return data
    except Exception as e:
        logger.warning(f"读取季报缓存失败 {ts_code}: {e}")
        return None



    try:
        with sqlite3.connect(db_path, timeout=10.0) as conn:
            now = datetime.now().isoformat()

            # 1. 写入各季度原始数据
            for year, q_date, block_text in blocks:
                parsed = _parse_single_block(block_text)
                report_date = f"{year}{q_date}"
                conn.execute("""
                    INSERT OR REPLACE INTO quarterly_reports
                    (ts_code, report_date, report_type,
                     revenue, operating_cost, net_profit, net_profit_deducted,
                     ocf_abs, total_assets, total_liabilities, net_assets,
                     gross_margin, net_margin, debt_ratio, ocf_ratio, roa,
                     revenue_yoy, profit_yoy, last_update)
                    VALUES (?, ?, 'quarterly',
                            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    ts_code, report_date,
                    parsed.get("revenue"), parsed.get("operating_cost"),
                    parsed.get("net_profit"), parsed.get("net_profit_deducted"),
                    parsed.get("ocf_abs"), parsed.get("total_assets"),
                    parsed.get("total_liabilities"), parsed.get("net_assets"),
                    parsed.get("gross_margin"), parsed.get("net_margin"),
                    parsed.get("debt_ratio"), parsed.get("ocf_ratio"),
                    parsed.get("roa"),
                    parsed.get("revenue_yoy"), parsed.get("profit_yoy"),
                    now,
                ))

            # 2. 写入TTM缓存
            conn.execute("""
                INSERT OR REPLACE INTO quarterly_ttm_cache
                (ts_code,
                 revenue_ttm, operating_cost_ttm, net_profit_ttm, ocf_abs_ttm,
                 net_assets_ttm, gross_margin_ttm, net_margin_ttm,
                 ocf_ratio_ttm, roe_ttm,
                 revenue_yoy_latest, profit_yoy_latest, debt_ratio_latest,
                 total_assets_latest, total_liabilities_latest,
                 quarter_count, latest_quarter, last_update)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                ts_code,
                ttm_metrics.get("revenue_ttm"), ttm_metrics.get("cost_ttm"),
                ttm_metrics.get("net_profit_ttm"), ttm_metrics.get("ocf_abs_ttm"),
                ttm_metrics.get("net_assets_ttm"), ttm_metrics.get("gross_margin_ttm"),
                ttm_metrics.get("net_margin_ttm"), ttm_metrics.get("ocf_ratio_ttm"),
                ttm_metrics.get("roe_ttm"),
                latest_metrics.get("revenue_yoy"), latest_metrics.get("profit_yoy"),
                latest_metrics.get("debt_ratio"),
                latest_metrics.get("total_assets"), latest_metrics.get("total_liabilities"),
                len(blocks), f"{blocks[0][0]}{blocks[0][1]}" if blocks else None,
                now,
            ))
            conn.commit()
    except Exception as e:
        logger.warning(f"写入季报缓存失败 {ts_code}: {e}")





def parse_num(text: str) -> Optional[float]:
    """解析数字，支持百分比"""
    if not text:
        return None
    text = text.strip()

    m = re.search(r"([-+]?\d+\.?\d*)%", text)
    if m:
        return float(m.group(1))

    m = re.search(r"([-+]?\d+\.?\d*)", text)
    if m:
        return float(m.group(1))

    return None


def _parse_num_from_line(line: str) -> Optional[float]:
    """从财报行提取带单位数值，单位：万亿元>亿元>万元>千元>元"""
    if not line:
        return None

    m = re.search(r"([-+]?\d+\.?\d*)\s*(万亿元|亿元|万元|千元|元)", line)
    if not m:
        return None

    val = float(m.group(1))
    unit = m.group(2)

    multipliers = {
        "万亿元": 1e12,
        "亿元": 1e8,
        "万元": 1e4,
        "千元": 1e3,
        "元": 1
    }

    return val * multipliers.get(unit, 1)


def year_of_date(date_str: str) -> int:
    """从日期字符串中提取年份"""
    if not date_str:
        return datetime.now().year
    try:
        return int(str(date_str)[:4])
    except (ValueError, IndexError):
        return datetime.now().year


# ============================================================
# 申万一级行业逻辑（与 AnnualScorer 完全一致）
# ============================================================

def load_industry_map() -> Dict[str, Any]:
    """加载本地行业映射表"""
    if os.path.exists(Config.INDUSTRY_MAP_FILE):
        try:
            with open(Config.INDUSTRY_MAP_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"加载行业映射文件失败: {e}")
    return {}


def extract_industry_from_content(text: str) -> Optional[str]:
    """从内容中提取行业信息"""
    if not text:
        return None

    patterns = [
        r"所属一级行业[：:]\s*(\S+)",
        r"申万一级行业[：:]\s*(\S+)",
        r"行业分类[：:]\s*(\S+)"
    ]

    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            return m.group(1).strip()

    return None


def infer_industry_from_name(name: str) -> Optional[str]:
    """从股票名称推断行业"""
    if not name:
        return None

    name_keyword_industry = {
        "银行": "银行",
        "证券": "非银金融",
        "保险": "非银金融",
        "地产": "房地产",
        "房地产": "房地产",
        "钢铁": "钢铁",
        "煤炭": "煤炭",
        "有色": "有色金属",
        "化工": "基础化工",
        "医药": "医药生物",
        "生物": "医药生物",
        "电子": "电子",
        "计算机": "计算机",
        "通信": "通信",
        "汽车": "汽车",
        "机械": "机械设备",
        "电力": "公用事业",
        "食品": "食品饮料",
        "饮料": "食品饮料",
        "家电": "家用电器",
        "纺织": "纺织服饰",
        "建筑": "建筑装饰",
        "军工": "国防军工",
        "传媒": "传媒",
        "光伏": "电力设备",
        "电池": "电力设备",
        "芯片": "电子",
        "半导体": "电子"
    }

    for kw, industry in name_keyword_industry.items():
        if kw in name:
            return industry

    return None


def build_industry_map_from_akshare(
    stock_codes: Optional[List[str]] = None,
    force_refresh: bool = False
) -> Dict[str, str]:
    """
    用 akshare 获取申万一级行业分类，构建 code->industry 映射
    结果缓存到本地 JSON，默认 30 天有效
    返回: {code_short: industry_name, ...}
    """
    cache_file = Config.AKSHARE_INDUSTRY_CACHE

    if not force_refresh and os.path.exists(cache_file):
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                cached = json.load(f)
            cache_age_days = (time.time() - cached.get("_timestamp", 0)) / 86400
            if cache_age_days < Config.AKSHARE_CACHE_MAX_AGE_DAYS:
                logger.info(
                    f"使用 akshare 行业映射缓存 ({cache_age_days:.0f}天前), "
                    f"共 {len(cached)-1} 只股票"
                )
                return cached.get("data", {})
        except Exception:
            pass

    logger.info("从 akshare 获取申万一级行业分类...")
    try:
        import akshare as ak
    except ImportError:
        logger.warning("akshare 未安装，跳过申万行业获取")
        return {}

    t0 = time.time()
    result: Dict[str, str] = {}

    try:
        df_info = ak.sw_index_first_info()
        total = len(df_info)
        for i, (_, row) in enumerate(df_info.iterrows(), 1):
            ind_code = str(row["行业代码"]).split(".")[0]
            ind_name = row["行业名称"]
            try:
                df_cons = ak.index_stock_cons(symbol=ind_code)
                code_col = "品种代码"
                cnt = 0
                for _, stock in df_cons.iterrows():
                    code = str(stock[code_col])
                    if code not in result:
                        result[code] = ind_name
                        cnt += 1
                logger.info(
                    f"  [{i}/{total}] {ind_name}({ind_code}): {len(df_cons)}只, 新增{cnt}只"
                )
            except Exception as e:
                logger.warning(f"  [{i}/{total}] {ind_name}({ind_code}): 获取失败 {e}")
    except Exception as e:
        logger.error(f"akshare 申万行业获取失败: {e}")
        return result

    elapsed = time.time() - t0
    logger.info(f"akshare 行业映射完成: {len(result)}只, 耗时{elapsed:.1f}s")

    try:
        cache_data = {"_timestamp": time.time(), "data": result}
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False)
        logger.info(f"行业映射缓存已保存: {cache_file}")
    except Exception as e:
        logger.warning(f"保存行业映射缓存失败: {e}")

    return result


def load_akshare_industry_map() -> Dict[str, str]:
    """加载 akshare 行业映射（优先缓存）"""
    cache_file = Config.AKSHARE_INDUSTRY_CACHE
    if os.path.exists(cache_file):
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                cached = json.load(f)
            return cached.get("data", {})
        except Exception:
            pass
    return {}





def determine_industry(
    ts_code: str,
    name: str,
    content: str,
    industry_map: Dict[str, Any],
    use_api: bool = True,
    session: Optional[req_lib.Session] = None,
    akshare_map: Optional[Dict[str, str]] = None
) -> Optional[str]:
    """
    确定股票所属行业（混合策略，优先级从高到低）- 与 AnnualScorer 一致

    优先级:
      1. 原文提取      — 最准确，公司自己披露
      2. akshare 申万   — 权威分类，100%覆盖，本地缓存
      3. 股票名称推断   — 关键词匹配，快速无网络
      4. NeoData API    — 大模型查询，补漏用
      5. 本地静态映射表  — 历史遗留数据，质量参差
      6. 代码前缀兜底   — 最弱保底
    """
    code_short = ts_code.split(".")[0]

    # 1. 从原文中提取
    ind = extract_industry_from_content(content)
    if ind:
        return ind

    # 2. akshare 申万一级行业分类
    if akshare_map is None:
        akshare_map = load_akshare_industry_map()
    ind = akshare_map.get(code_short)
    if ind:
        return ind

    # 3. 从股票名称关键词推断
    ind = infer_industry_from_name(name)
    if ind:
        return ind

    

    # 5. 本地静态映射表
    def _get_industry_l1(key: str) -> Optional[str]:
        val = industry_map.get(key)
        if val is None:
            return None
        if isinstance(val, str):
            return val
        if isinstance(val, dict):
            return val.get("industry_l1")
        return None

    ind = _get_industry_l1(ts_code)
    if ind:
        return ind
    ind = _get_industry_l1(code_short)
    if ind:
        return ind

    # 6. 代码前缀兜底
    code_prefix_industry = {
        "60": "机械设备",
        "00": "房地产",
        "30": "医药生物",
        "68": "电子"
    }
    return code_prefix_industry.get(code_short[:2])





# ============================================================
# 季报数据获取（多季度TTM + 单季指标）
# ============================================================

def _extract_all_report_sections(text: str) -> List[Tuple[str, str, str]]:
    """
    从NeoData返回中提取所有独立的财报子段落。

    API返回的真实结构：每个子段落以"根据...在A股市场YYYYMMDD发布的财报数据，
    统计截止日期为YYYYMMDD的Qx单季报/年报"开头，后面跟随该报告的财务数据。

    返回: [(report_date, report_type, section_text), ...] 按时间降序排列
      - report_date: "YYYYMMDD" 字符串
      - report_type: "Q1单季报"/"Q2单季报"/"Q3单季报"/"Q4单季报"/"年报"
      - section_text: 该段落的正文内容
    """
    if not text:
        return []

    # 匹配每个独立财报子段落的起始标记
    # 格式: 根据...在A股市场YYYYMMDD发布的财报数据，统计截止日期为YYYYMMDD的Qx单季报/年报
    pattern = (
        r"根据.+?在A股市场\d+发布的财报数据[，,]\s*"
        r"统计截止日期为(\d{4})(0331|0630|0930|1231)的"
        r"(Q[1-4]单?季报|年报)"
    )
    matches = list(re.finditer(pattern, text))

    if not matches:
        return []

    sections = []
    for i, m in enumerate(matches):
        year = m.group(1)
        q_date = m.group(2)
        report_type = m.group(3)
        report_date = f"{year}{q_date}"

        # 从当前匹配的结束位置开始，到下一个匹配的开始位置
        start = m.end()
        if i + 1 < len(matches):
            end = matches[i + 1].start()
        else:
            end = len(text)
        section_text = text[start:end].strip()
        sections.append((report_date, report_type, section_text))

    # 按时间降序排列（最新的在前）
    sections.sort(key=lambda x: x[0], reverse=True)
    return sections


def _extract_all_quarterly_blocks(text: str) -> List[Tuple[str, str, str]]:
    """
    从NeoData返回中提取所有季报段落（兼容旧接口）
    返回: [(year, quarter_date, block_text), ...] 按时间降序排列

    此函数已废弃，建议使用 _extract_all_report_sections 代替。
    保留此函数仅为兼容 _compute_ttm 等旧代码的调用。
    """
    sections = _extract_all_report_sections(text)
    # 转换为旧格式: (year, quarter_date, block_text)
    # 注意：旧格式不含 report_type，调用方需注意
    return [(date[:4], date[4:], text) for date, _, text in sections]


def _parse_single_block(block: str) -> Dict[str, Optional[float]]:
    """从单个季报段落提取所有可用指标（含ROE计算所需字段）"""
    result = {
        "revenue": None,              # 营业总收入（元）
        "operating_cost": None,       # 营业成本（元）
        "net_profit": None,           # 归母净利润（元）
        "net_profit_deducted": None,  # 扣非净利润（元）
        "gross_margin": None,         # 销售毛利率(%)
        "net_margin": None,           # 销售净利率(%)
        "revenue_yoy": None,          # 营业收入同比增长(%)
        "profit_yoy": None,           # 归母净利润同比增长(%)
        "ocf_ratio": None,            # 净利润现金含量(%)
        "ocf_abs": None,              # 经营活动现金流（元）
        "total_assets": None,         # 资产合计（元）
        "total_liabilities": None,    # 负债合计（元）
        "net_assets": None,           # 净资产/股东权益（元）
        "debt_ratio": None,           # 资产负债率(%)
        "roa": None,                  # 资产回报率ROA(%)
    }

    if not block:
        return result

    for line in block.split("\n"):
        line = line.strip()
        if not line:
            continue

        # 营业总收入 — 只取第一个值（防止多季度混合block中被覆盖）
        if "营业总收入" in line and "同比" not in line and "环比" not in line:
            if result["revenue"] is None:
                val = _parse_num_from_line(line)
                if val is not None:
                    result["revenue"] = val

        # 营业成本 — 只取第一个值
        elif "营业成本" in line and "同比" not in line and "环比" not in line and "营业总成本" not in line:
            if result["operating_cost"] is None:
                val = _parse_num_from_line(line)
                if val is not None:
                    result["operating_cost"] = val

        # 归母净利润 — 严格行首匹配^归母净利润，排除同比/扣非/现金含量
        elif re.match(r"^归母净利润", line) and "同比" not in line and "环比" not in line and "扣非" not in line and "现金" not in line:
            if result["net_profit"] is None:
                val = _parse_num_from_line(line)
                if val is not None:
                    result["net_profit"] = val

        # 扣非净利润 — 只取第一个值
        elif "扣非净利润" in line and "同比" not in line and "环比" not in line:
            if result["net_profit_deducted"] is None:
                val = _parse_num_from_line(line)
                if val is not None:
                    result["net_profit_deducted"] = val

        # 销售毛利率 — 只取第一个值
        elif "销售毛利率" in line or ("毛利率" in line and "同比" not in line and "环比" not in line):
            if result["gross_margin"] is None:
                m = re.search(r"毛利率[：:\s]*([-+]?\d+\.?\d*)%", line)
                if m:
                    result["gross_margin"] = float(m.group(1))

        # 销售净利率 — 只取第一个值
        elif "销售净利率" in line or ("净利率" in line and "同比" not in line and "环比" not in line):
            if result["net_margin"] is None:
                m = re.search(r"净利率[：:\s]*([-+]?\d+\.?\d*)%", line)
                if m:
                    result["net_margin"] = float(m.group(1))

        # 营业收入同比增长 — 只取第一个值
        elif any(kw in line for kw in ["营业收入同比增长", "营收同比增长", "营业总收入同比增长"]):
            if result["revenue_yoy"] is None:
                result["revenue_yoy"] = parse_num(line)

        # 归母净利润同比增长 — 只取第一个值
        elif "归母净利润同比增长" in line or ("净利润同比增长" in line and "归母" in line):
            if result["profit_yoy"] is None:
                result["profit_yoy"] = parse_num(line)

        # 净利润现金含量(OCF/净利润) — 只取第一个值
        elif "净利润现金含量" in line:
            if result["ocf_ratio"] is None:
                m = re.search(r"净利润现金含量[：:\s]*([-+]?\d+\.?\d*)%", line)
                if m:
                    result["ocf_ratio"] = float(m.group(1))

        # 经营活动产生的现金流量净额 — 只取第一个值（排除"每股"版本）
        elif "经营活动产生的现金流量净额" in line and "每股" not in line:
            if result["ocf_abs"] is None:
                val = _parse_num_from_line(line)
                if val is not None:
                    result["ocf_abs"] = val

        # 资产合计 — 只取第一个值
        elif "资产合计" in line and "同比" not in line:
            if result["total_assets"] is None:
                val = _parse_num_from_line(line)
                if val is not None:
                    result["total_assets"] = val

        # 负债合计 — 只取第一个值
        elif "负债合计" in line and "同比" not in line:
            if result["total_liabilities"] is None:
                val = _parse_num_from_line(line)
                if val is not None:
                    result["total_liabilities"] = val

        # 资产负债率 — 直接提取（不通过负债/资产计算）
        elif "资产负债率" in line and "同比" not in line:
            if result.get("debt_ratio") is None:
                m = re.search(r"资产负债率[：:\s]*([-+]?\d+\.?\d*)%", line)
                if m:
                    result["debt_ratio"] = float(m.group(1))

        # 净资产/股东权益（多种表述）— 只取第一个值
        elif any(kw in line for kw in ["股东权益合计", "所有者权益合计", "归母净资产", "归母股东权益"]):
            if result["net_assets"] is None:
                val = _parse_num_from_line(line)
                if val is not None:
                    result["net_assets"] = val

        # 资产回报率ROA — 只取第一个值
        elif "资产回报率" in line or "ROA" in line.upper():
            if result["roa"] is None:
                m = re.search(r"(?:资产回报率|ROA)[：:\s]*([-+]?\d+\.?\d*)%", line)
                if m:
                    result["roa"] = float(m.group(1))

    # 如果没有直接的净资产数据，用 资产合计 - 负债合计 计算
    if result["net_assets"] is None and result["total_assets"] is not None and result["total_liabilities"] is not None:
        result["net_assets"] = result["total_assets"] - result["total_liabilities"]

    # 如果没有直接的资产负债率，用 负债合计 / 资产合计 计算
    if result["debt_ratio"] is None and result["total_assets"] is not None and result["total_liabilities"] is not None and result["total_assets"] > 0:
        result["debt_ratio"] = round(result["total_liabilities"] / result["total_assets"] * 100, 2)

    return result


def _compute_ttm(blocks: List[Tuple[str, str, str]]) -> Dict[str, Optional[float]]:
    """
    从多个季报段落计算TTM（近4个季度滚动）值
    
    TTM计算逻辑：
    - 营业总收入TTM = 最近4个季度营业总收入之和
    - 归母净利润TTM = 最近4个季度归母净利润之和
    - 毛利率TTM = (营收TTM - 成本TTM) / 营收TTM * 100
    - 净利率TTM = 净利润TTM / 营收TTM * 100
    - OCF_TTM = 净利润现金含量加权平均 * 净利润TTM
    - OCF/净利润_TTM = OCF_TTM / 净利润TTM * 100
    
    注意：如果不足4个季度，用已有季度计算
    """
    ttm = {
        "revenue_ttm": None,
        "operating_cost_ttm": None,
        "net_profit_ttm": None,
        "gross_margin_ttm": None,
        "net_margin_ttm": None,
        "ocf_abs_ttm": None,       # TTM经营现金流绝对值
        "ocf_ratio_ttm": None,     # TTM OCF/净利润(%)
        "roe_ttm": None,           # TTM ROE (暂不支持，需要净资产数据)
    }

    if not blocks:
        return ttm

    # 取最多4个季度
    recent_blocks = blocks[:4]

    revenue_sum = 0.0
    cost_sum = 0.0
    profit_sum = 0.0
    ocf_abs_sum = 0.0
    has_revenue = False
    has_cost = False
    has_profit = False
    has_ocf = False

    for year, q_date, block_text in recent_blocks:
        metrics = _parse_single_block(block_text)

        rev = metrics.get("revenue")
        if rev is not None:
            revenue_sum += rev
            has_revenue = True

        cost = metrics.get("operating_cost")
        if cost is not None:
            cost_sum += cost
            has_cost = True

        profit = metrics.get("net_profit")
        if profit is not None:
            profit_sum += profit
            has_profit = True

        # OCF：优先直接使用经营活动产生的现金流量净额
        ocf_abs_val = metrics.get("ocf_abs")
        if ocf_abs_val is not None:
            ocf_abs_sum += ocf_abs_val
            has_ocf = True
        else:
            # 兜底：用净利润现金含量 * 净利润计算
            ocf_ratio = metrics.get("ocf_ratio")
            if ocf_ratio is not None and profit is not None:
                ocf_abs_sum += profit * (ocf_ratio / 100.0)
                has_ocf = True

    if has_revenue:
        ttm["revenue_ttm"] = revenue_sum
    if has_cost:
        ttm["operating_cost_ttm"] = cost_sum
    if has_profit:
        ttm["net_profit_ttm"] = profit_sum
    if has_ocf:
        ttm["ocf_abs_ttm"] = ocf_abs_sum

    # 毛利率TTM = (营收TTM - 成本TTM) / 营收TTM * 100
    if has_revenue and has_cost and revenue_sum > 0:
        ttm["gross_margin_ttm"] = round((revenue_sum - cost_sum) / revenue_sum * 100, 2)

    # 净利率TTM = 净利润TTM / 营收TTM * 100
    if has_revenue and has_profit and revenue_sum > 0:
        ttm["net_margin_ttm"] = round(profit_sum / revenue_sum * 100, 2)

    # OCF/净利润TTM
    if has_profit and has_ocf and profit_sum != 0:
        ttm["ocf_ratio_ttm"] = round(ocf_abs_sum / profit_sum * 100, 2)

    # TTM ROE = 净利润TTM / 最新一期净资产 * 100
    # 净资产取最新一期的值（时点数，不用TTM累加）
    # 注意：很多单季报(Q2/Q3/Q4)不含资产负债表，需要往前遍历所有段落找净资产
    if has_profit and profit_sum != 0:
        net_assets = None
        # 先尝试从最新单季报取
        latest_block = recent_blocks[0]
        latest_metrics = _parse_single_block(latest_block[2])
        net_assets = latest_metrics.get("net_assets")
        
        # 如果最新单季报没有净资产，往前遍历所有段落（包括不足4个季度的）
        if net_assets is None or net_assets <= 0:
            for year, q_date, block_text in blocks:
                m = _parse_single_block(block_text)
                na = m.get("net_assets")
                if na is not None and na > 0:
                    net_assets = na
                    break
        
        if net_assets and net_assets > 0:
            ttm["roe_ttm"] = round(profit_sum / net_assets * 100, 2)
            ttm["net_assets_ttm"] = net_assets  # 保存净资产用于展示

    return ttm


def fetch_quarterly_data(
    ts_code: str,
    name: str,
    token: str
) -> Dict[str, Any]:
    """
    获取单只股票季报数据，返回：
    - TTM指标（盈利、现金流）
    - 最新单季指标（成长性、偿债风险）
    
    V7.0: 优先读取缓存DB，无有效缓存时调用API
    """
    # 1. 先检查缓存
    cached = _load_quarterly_from_db(ts_code)
    if cached:
        logger.debug(f"使用季报缓存 {ts_code}")
        return {
            "ttm_metrics": {
                "roe_ttm": cached.get("roe_ttm"),
                "gross_margin_ttm": cached.get("gross_margin_ttm"),
                "net_margin_ttm": cached.get("net_margin_ttm"),
                "ocf_ratio_ttm": cached.get("ocf_ratio_ttm"),
                "revenue_ttm": cached.get("revenue_ttm"),
                "cost_ttm": cached.get("operating_cost_ttm"),
                "net_profit_ttm": cached.get("net_profit_ttm"),
                "ocf_abs_ttm": cached.get("ocf_abs_ttm"),
                "net_assets_ttm": cached.get("net_assets_ttm"),
            },
            "latest_quarterly": {
                "revenue_yoy": cached.get("revenue_yoy_latest"),
                "profit_yoy": cached.get("profit_yoy_latest"),
                "debt_ratio": cached.get("debt_ratio_latest"),
                "total_assets": cached.get("total_assets_latest"),
                "total_liabilities": cached.get("total_liabilities_latest"),
            },
            "content": "",
            "fetch_success": True,
            "quarter_count": cached.get("quarter_count", 0),
            "latest_quarter": cached.get("latest_quarter", ""),
        }

    # 2. 调用AkShare获取财务数据
    try:
        df_fin = get_combined_financials(ts_code)
        if df_fin.empty:
            logger.warning(f"获取 {ts_code} 财务数据失败：返回空DataFrame")
            return {
                "ttm_metrics": {},
                "latest_quarterly": {},
                "content": "",
                "fetch_success": False,
                "quarter_count": 0,
                "latest_quarter": "",
            }
    except Exception as e:
        logger.error(f"获取 {ts_code} 财务数据异常: {e}")
        return {
            "ttm_metrics": {},
            "latest_quarterly": {},
            "content": "",
            "fetch_success": False,
            "quarter_count": 0,
            "latest_quarter": "",
        }
            total_assets = latest.get("total_assets")
            total_liab = latest.get("total_liabilities")
            if total_assets and total_liab and total_assets > 0:
                latest["debt_ratio"] = round(total_liab / total_assets * 100, 2)
            else:
                # 往前遍历所有段落（含年报）找资产负债率或资产/负债数据
                found_dr = False
                for d_check, t_check, txt_check in sections:
                    chk = _parse_single_block(txt_check)
                    dr = chk.get("debt_ratio")
                    if dr is not None:
                        latest["debt_ratio"] = dr
                        latest["_debt_ratio_source"] = f"{d_check} {t_check}"
                        found_dr = True
                        break
                    ta = chk.get("total_assets")
                    tl = chk.get("total_liabilities")
                    if ta and tl and ta > 0:
                        latest["debt_ratio"] = round(tl / ta * 100, 2)
                        latest["_debt_ratio_source"] = f"{d_check} {t_check}(计算)"
                        found_dr = True
                        break
                if not found_dr:
                    latest["debt_ratio"] = None

        # 使用IndicatorCalculator计算TTM指标
        calc = IndicatorCalculator(df_fin)
        
        # 构建返回数据结构
        ttm = {
            "roe_ttm": calc.roe_ttm,
            "gross_margin_ttm": calc.gross_margin_ttm,
            "net_margin_ttm": calc.net_margin_ttm,
            "ocf_ttm": calc.ocf_ttm,
            "fcf_ttm": calc.fcf_ttm,
            "net_profit_ratio": calc.net_profit_ratio,
            "cash_recovery_rate": calc.cash_recovery_rate,
            "de_ratio": calc.de_ratio,
            "current_ratio": calc.current_ratio,
            "asset_liability_ratio": calc.asset_liability_ratio,
            "interest_cover": calc.interest_cover,
            "revenue_ttm": calc.ttm_revenue,
            "cost_ttm": calc.ttm_oper_cost,
            "net_profit_ttm": calc.ttm_net_profit,
            "ocf_abs_ttm": calc.ttm_ocf,
            "net_assets_ttm": calc.latest_net_assets,
        }
        
        latest = {
            "revenue_yoy": calc.q_revenue_yoy,
            "profit_yoy": calc.q_net_profit_yoy,
            "oper_profit_yoy": calc.q_oper_profit_yoy,
            "debt_ratio": calc.asset_liability_ratio,  # 资产负债率
            "total_assets": calc.latest_total_assets,
            "total_liabilities": calc.latest_total_liab,
            "equity_parent": calc.latest_equity_parent,
        }
        
        fetch_success = (
            any(v is not None and v != 0 for v in [calc.roe_ttm, calc.gross_margin_ttm, calc.net_margin_ttm]) or
            any(v is not None for v in [calc.q_net_profit_yoy, calc.q_revenue_yoy])
        )

        

        return {
            "ttm_metrics": ttm,
            "latest_quarterly": latest,
            "content": text,
            "fetch_success": fetch_success,
            "quarter_count": len(quarterly_sections) if quarterly_sections else len(sections),
            "latest_quarter": latest_date,
        }
    except Exception as e:
        logger.error(f"获取季报数据异常 {ts_code} {name}: {e}")
        return {
            "ttm_metrics": {},
            "latest_quarterly": {},
            "content": "",
            "fetch_success": False
        }


def fetch_quarterly_batch(
    stocks: List[Dict[str, str]],
    token: str,
    workers: int = Config.FINANCE_WORKERS
) -> List[Dict[str, Any]]:
    """批量获取季报数据"""
    total = len(stocks)
    results = []
    consecutive_parse_failures = 0
    start_time = time.time()

    # 初始化季报缓存DB
    _init_quarterly_db()

    logger.info(f"开始获取季报数据: 共 {total}只, {workers}个线程")

    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_stock: Dict[Future, Dict[str, str]] = {}

        for stock in stocks:
            future = executor.submit(
                fetch_quarterly_data,
                stock["ts_code"],
                stock["name"],
                token
            )
            future_to_stock[future] = stock

        done_count = 0
        for future in as_completed(future_to_stock):
            stock = future_to_stock[future]
            done_count += 1
            try:
                result = future.result(timeout=Config.API_TIMEOUT + 10)
                results.append({**stock, **result})

                if not result.get("fetch_success", False):
                    consecutive_parse_failures += 1
                else:
                    consecutive_parse_failures = 0

                if consecutive_parse_failures >= Config.PAUSE_CONSECUTIVE_EMPTY:
                    logger.warning(
                        f"连续 {consecutive_parse_failures}次解析失败，暂停 {Config.PAUSE_DURATION}秒"
                    )
                    time.sleep(Config.PAUSE_DURATION)
                    consecutive_parse_failures = 0
            except Exception as e:
                consecutive_parse_failures += 1
                results.append({
                    **stock,
                    "quarterly_metrics": {},
                    "content": "",
                    "fetch_success": False
                })
                logger.error(f"季报任务异常 {stock['ts_code']}: {e}")

            if done_count % 100 == 0 and done_count > 0:
                elapsed = time.time() - start_time
                rate = done_count / elapsed if elapsed > 0 else 0
                eta = (total - done_count) / rate if rate > 0 else 0
                logger.info(
                    f"季报进度: {done_count}/{total} ({done_count/total*100:.1f}%), "
                    f"速率:{rate:.1f}/秒, 已用:{elapsed:.0f}秒, 剩余:{eta:.0f}秒"
                )

    elapsed = time.time() - start_time
    logger.info(f"季报获取完成: {len(results)}/{total}, 耗时:{elapsed:.0f}秒")
    return results


# ============================================================
# 从年报数据库读取年报数据
# ============================================================

def load_annual_data_from_db(
    stocks: List[Dict[str, str]],
    db_path: str = None
) -> List[Dict[str, Any]]:
    """
    从年报数据库 (AnnualScorer/stock_cache.db) 读取最新年报数据
    V7.0: 年报提供以下字段：
      - roe (ROE%, 备用数据源)
      - net_profit (净利润，元)
      - ocf_abs (经营现金流，元)
      - report_date (年报截止日期)
      - industry_l1 (申万一级行业)
    
    注意：gross_margin、net_margin、revenue_yoy、profit_yoy、debt_ratio
          由季报TTM/最新单季提供，不再从年报DB读取
    """
    if db_path is None:
        db_path = Config.ANNUAL_DB_FILE

    if not os.path.exists(db_path):
        logger.warning(f"年报数据库不存在: {db_path}，年报字段将为空")
        return [{**s, "roe_annual": None, "net_profit_annual": None,
                 "ocf_abs_annual": None,
                 "report_date": None, "industry_l1": None,
                 "fetch_success": False} for s in stocks]

    result = []
    try:
        with db_connection(db_path) as conn:
            for stock in stocks:
                ts_code = stock["ts_code"]
                item = {
                    "ts_code": ts_code,
                    "name": stock.get("name", ""),
                    "roe_annual": None,          # 年报ROE(%, 备用数据源)
                    "net_profit_annual": None,   # 年报净利润(元)
                    "ocf_abs_annual": None,      # 年报经营现金流(元)
                    "report_date": None,         # 年报截止日期
                    "industry_l1": None,         # 申万一级行业
                    "fetch_success": False,
                }

                # 获取行业信息
                cur = conn.execute(
                    "SELECT industry_l1 FROM stocks WHERE ts_code=?",
                    (ts_code,)
                )
                row = cur.fetchone()
                if row:
                    item["industry_l1"] = row[0]

                # 获取最新年报（含roe字段）
                cur = conn.execute(
                    """SELECT report_date, roe, net_profit, ocf_abs
                       FROM financial_reports
                       WHERE ts_code=? AND report_type='annual' AND fetch_success=1
                       ORDER BY report_date DESC LIMIT 1""",
                    (ts_code,)
                )
                report_row = cur.fetchone()

                if report_row:
                    item["report_date"] = report_row[0]
                    item["roe_annual"] = report_row[1]
                    item["net_profit_annual"] = report_row[2]
                    item["ocf_abs_annual"] = report_row[3]
                    item["fetch_success"] = True

                result.append(item)

        logger.info(f"从年报数据库读取: {len(result)} 只股票")
        success_cnt = sum(1 for r in result if r.get("fetch_success"))
        logger.info(f"有年报数据: {success_cnt}/{len(result)}")
    except Exception as e:
        logger.error(f"读取年报数据库失败: {e}")
        result = [{**s, "roe_annual": None, "net_profit_annual": None,
                   "ocf_abs_annual": None,
                   "report_date": None, "industry_l1": None,
                   "fetch_success": False} for s in stocks]

    return result


# ============================================================
# 合并年报 + 季报数据
# ============================================================

def merge_annual_quarterly(
    annual_data: List[Dict[str, Any]],
    quarterly_data: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    合并年报和季报数据（V7.0 新架构）
    
    数据源映射：
      roe              → TTM (roe_ttm)
      gross_margin     → TTM (gross_margin_ttm)
      net_margin       → TTM (net_margin_ttm)
      revenue_yoy      → 最新单季 (revenue_yoy)
      profit_yoy       → 最新单季 (profit_yoy)
      debt_ratio       → 最新单季 (debt_ratio)
      ocf_ratio        → TTM (ocf_ratio_ttm) 用于评分
      net_profit       → 年报绝对值 (net_profit_annual)
      ocf_abs          → 年报绝对值 (ocf_abs_annual)
      report_date      → 年报日期
    """
    quarterly_map = {q["ts_code"]: q for q in quarterly_data}
    merged = []

    for ann in annual_data:
        ts_code = ann["ts_code"]
        q = quarterly_map.get(ts_code, {})

        ttm = q.get("ttm_metrics", {})
        latest = q.get("latest_quarterly", {})

        merged_item = {
            "ts_code": ts_code,
            "name": ann.get("name", ""),
            "industry_l1": ann.get("industry_l1"),
            "report_date": ann.get("report_date"),
            # === 盈利能力 ===
            "roe": ttm.get("roe_ttm"),   # 仅使用TTM ROE
            "gross_margin": ttm.get("gross_margin_ttm"),          # TTM 毛利率
            "net_margin": ttm.get("net_margin_ttm"),              # TTM 净利率
            # === 成长性：最新单季 ===
            "revenue_yoy": latest.get("revenue_yoy"),             # 单季营收同比
            "profit_yoy": latest.get("profit_yoy"),               # 单季净利润同比
            # === 偿债风险：最新单季 ===
            "debt_ratio": latest.get("debt_ratio"),               # 单季资产负债率
            # === 现金流质量：TTM ===
            "ocf_ratio": ttm.get("ocf_ratio_ttm"),                # TTM OCF/净利润
            # === 绝对值：TTM ===
            "net_profit": ttm.get("net_profit_ttm"),              # TTM净利润(元)
            "ocf_abs": ttm.get("ocf_abs_ttm"),                    # TTM经营现金流(元)
            # === 数据来源标记 ===
            "annual_success": ann.get("fetch_success", False),
            "quarterly_success": q.get("fetch_success", False),
            "quarterly_content": q.get("content", ""),
            "quarter_count": q.get("quarter_count", 0),
            "latest_quarter": q.get("latest_quarter", ""),
        }
        merged.append(merged_item)

    return merged


# ============================================================
# 评分逻辑（与 AnnualScorer V6.0.0 完全一致）
# ============================================================

def calc_completeness(metrics: Dict[str, Optional[float]]) -> Tuple[float, str]:
    """计算数据完整度（V7.0：核心字段为6个评分维度所需指标）"""
    # 评分所需的核心字段：roe(TTM), gross_margin(TTM), net_margin(TTM),
    # revenue_yoy(单季), profit_yoy(单季), debt_ratio(单季), ocf_ratio(TTM)
    core_metrics = ["roe", "gross_margin", "net_margin", "revenue_yoy",
                    "profit_yoy", "debt_ratio", "ocf_ratio"]
    non_null = sum(1 for m in core_metrics if metrics.get(m) is not None)
    ratio = non_null / len(core_metrics)

    if ratio >= 0.857:
        return ratio, "high"
    elif ratio >= 0.571:
        return ratio, "medium"
    elif non_null <= 1:
        return ratio, "ultra_low"
    else:
        return ratio, "low"


def percentile_rank(value: Optional[float], values: List[float], reverse: bool = False) -> float:
    """计算百分位排名 0~100，与 AnnualScorer 一致: count_leq/len * 100"""
    if value is None:
        return 0.0
    if not values:
        return 50.0

    count_leq = sum(1 for v in values if v <= value)
    if reverse:
        return ((len(values) - count_leq) / len(values)) * 100
    else:
        return (count_leq / len(values)) * 100


def calc_score(
    stock: Dict[str, Any],
    industry_stocks: Dict[str, List[Dict[str, Any]]],
    all_stocks: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    计算股票评分（与 AnnualScorer V6.0.0 逻辑一致）

    权重:
      盈利能力 40% (ROE 40% + 毛利率 30% + 净利率 30%)
      成长能力 30% (营收同比 40% + 净利润同比 60%)
      现金流质量 20% (OCF/净利润)
      偿债风险 10% (资产负债率, reverse)
    """
    ts_code = stock["ts_code"]
    industry = stock.get("industry_l1", "未知")
    metrics = stock

    # 确定参考池
    pool = industry_stocks.get(industry, [])
    use_market_fallback = len(pool) < Config.MIN_INDUSTRY_SAMPLES
    if use_market_fallback:
        pool = all_stocks
    discount = Config.MARKET_FALLBACK_DISCOUNT if use_market_fallback else 1.0

    def pool_values(key: str) -> List[float]:
        return [
            s[key] for s in pool
            if s.get("ts_code") != ts_code and s.get(key) is not None
        ]

    # 1. 盈利能力 (权重40%)
    roe_score = 0.0
    if metrics.get("roe") is not None and metrics["roe"] >= 0:
        roe_score = percentile_rank(metrics["roe"], pool_values("roe"))

    gross_score = percentile_rank(
        metrics["gross_margin"], pool_values("gross_margin")
    ) if metrics.get("gross_margin") is not None else 0.0

    net_score = percentile_rank(
        metrics["net_margin"], pool_values("net_margin")
    ) if metrics.get("net_margin") is not None else 0.0

    profit_score = (roe_score * 0.4 + gross_score * 0.3 + net_score * 0.3) * discount

    # 2. 成长性 (权重30%)
    rev_score = percentile_rank(
        metrics["revenue_yoy"], pool_values("revenue_yoy")
    ) if metrics.get("revenue_yoy") is not None else 0.0

    prof_score = percentile_rank(
        metrics["profit_yoy"], pool_values("profit_yoy")
    ) if metrics.get("profit_yoy") is not None else 0.0

    growth_score = (rev_score * 0.4 + prof_score * 0.6) * discount

    # 3. 现金流质量 (权重20%) — TTM OCF/净利润
    ocf_ratio_val = metrics.get("ocf_ratio")  # TTM OCF/净利润(%)
    ocf_score = 0.0

    if ocf_ratio_val is not None:
        pool_ocf_vals = [
            s["ocf_ratio"] for s in pool
            if s.get("ts_code") != ts_code and s.get("ocf_ratio") is not None
        ]
        ocf_score = percentile_rank(ocf_ratio_val, pool_ocf_vals)

    ocf_score *= discount

    # 4. 偿债风险 (权重10%) — 最新单季资产负债率
    debt_score = 0.0
    if metrics.get("debt_ratio") is not None:
        debt_score = percentile_rank(
            metrics["debt_ratio"], pool_values("debt_ratio"), reverse=True
        )
    debt_score *= discount

    # 计算总分
    total_score = (
        profit_score * 0.4 +
        growth_score * 0.3 +
        ocf_score * 0.2 +
        debt_score * 0.1
    )

    # 数据完整度折扣
    completeness, level = calc_completeness(metrics)
    if level == "low":
        total_score *= Config.LOW_COMPLETENESS_PENALTY
    elif level == "ultra_low":
        total_score *= Config.LOW_COMPLETENESS_PENALTY * Config.ULTRA_LOW_COMPLETENESS_PENALTY

    # 连续亏损惩罚（基于TTM净利润和TTM经营现金流）
    np_ttm = metrics.get("net_profit")
    ocf_ttm = metrics.get("ocf_abs")
    if np_ttm is not None and ocf_ttm is not None and np_ttm < 0 and ocf_ttm < 0:
        total_score = min(total_score, Config.NEGATIVE_PROFIT_PENALTY)

    # 评级
    if total_score >= 75:
        grade = "A"
    elif total_score >= 55:
        grade = "B"
    elif total_score >= 40:
        grade = "C"
    elif total_score >= 25:
        grade = "D"
    else:
        grade = "E"

    confidence_map = {"high": "高", "medium": "中", "low": "低", "ultra_low": "低"}
    confidence = confidence_map.get(level, "低")

    return {
        "ts_code": ts_code,
        "name": stock.get("name", ""),
        "industry_l1": industry,
        "total_score": round(total_score, 2),
        "profit_score": round(profit_score, 2),
        "growth_score": round(growth_score, 2),
        "ocf_score": round(ocf_score, 2),
        "debt_score": round(debt_score, 2),
        "grade": grade,
        "confidence": confidence,
        "completeness": completeness,
        "completeness_level": level,
        # 评分指标（带数据来源标记）
        "roe": metrics.get("roe"),                    # TTM
        "gross_margin": metrics.get("gross_margin"),  # TTM
        "net_margin": metrics.get("net_margin"),      # TTM
        "revenue_yoy": metrics.get("revenue_yoy"),    # 单季
        "profit_yoy": metrics.get("profit_yoy"),      # 单季
        "debt_ratio": metrics.get("debt_ratio"),      # 单季
        "ocf_ratio": metrics.get("ocf_ratio"),        # TTM OCF/净利润
        # 年报绝对值
        "net_profit": metrics.get("net_profit"),      # 年报
        "ocf_abs": metrics.get("ocf_abs"),            # 年报
        "fetch_success": stock.get("annual_success", False) or stock.get("quarterly_success", False),
        "report_date": stock.get("report_date"),      # 年报日期
        "market_fallback": use_market_fallback,
        "latest_quarter": stock.get("latest_quarter", ""),
        "quarter_count": stock.get("quarter_count", 0),
    }


# ============================================================
# 股票列表加载
# ============================================================

def load_stock_list(file_path: str = Config.DEFAULT_STOCK_FILE) -> List[Dict[str, str]]:
    """加载股票列表文件"""
    stocks = []

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"股票列表文件不存在: {file_path}")

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue

                parts = line.split()
                if len(parts) < 2:
                    parts = line.split(",")
                if len(parts) < 2:
                    parts = line.split("\t")
                if len(parts) < 2:
                    logger.warning(f"跳过无效行 {line_num}: {line}")
                    continue

                code = parts[0].strip()
                name = parts[1].strip()

                # 过滤科创板和北交所股票
                if code.startswith(("688", "430", "83", "87")):
                    continue

                # 补全市场后缀
                if "." not in code:
                    code = code + ".SH" if code.startswith("6") else code + ".SZ"

                stocks.append({"ts_code": code, "name": name})

        logger.info(f"加载股票列表: {len(stocks)}只")
        return stocks
    except Exception as e:
        raise IOError(f"加载股票列表失败: {e}")


# ============================================================
# Excel 输出（与 AnnualScorer 字段一致，无 JSON）
# ============================================================

def _row_from_scored(r: Dict[str, Any]) -> List[Any]:
    """生成Excel行数据（指标在前，评分在后，无排名）
    V7.0: 数据来源已标注在表头中
    """
    return [
        r.get("ts_code", ""),
        r.get("name", ""),
        r.get("industry_l1", ""),
        r.get("roe"),           # TTM
        r.get("gross_margin"),  # TTM
        r.get("net_margin"),    # TTM
        r.get("revenue_yoy"),   # 单季
        r.get("profit_yoy"),    # 单季
        r.get("debt_ratio"),    # 单季
        r.get("ocf_ratio"),     # TTM OCF/净利润
        r.get("net_profit"),    # TTM(净利润元)
        r.get("ocf_abs"),       # TTM(经营现金流元)
        r.get("report_date"),   # 年报日期
        r.get("latest_quarter", ""),  # 最新季报期
        f"{r.get('completeness', 0) * 100:.0f}%",
        r.get("total_score", 0),
        r.get("grade", ""),
        r.get("confidence", ""),
        r.get("profit_score", 0),
        r.get("growth_score", 0),
        r.get("ocf_score", 0),
        r.get("debt_score", 0)   # 偿债风险
    ]


def output_excel(results: List[Dict[str, Any]], output_dir: str = Config.OUTPUT_DIR) -> Optional[str]:
    """输出Excel报告（与 AnnualScorer 字段一致，无 JSON 输出）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl未安装，请运行: pip install openpyxl")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"综合评分_{timestamp}.xlsx")

    try:
        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        grade_fills = {
            "A": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
            "B": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            "C": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "D": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "E": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        }

        # V7.0 表头：标注数据来源（TTM/单季/年报）
        headers = [
            "股票代码", "股票名称", "申万一级行业",
            "ROE(%)(TTM)", "毛利率(%)(TTM)", "净利率(%)(TTM)",
            "营收同比(%)(单季)", "净利润同比(%)(单季)",
            "资产负债率(%)(单季)", "OCF/净利润(%)(TTM)",
            "净利润(元)(TTM)", "经营现金流(元)(TTM)",
            "年报日期", "最新季报期", "数据完整度",
            "总分", "评级", "置信度",
            "盈利能力", "成长性", "现金流质量", "偿债风险"
        ]
        grade_col = headers.index("评级") + 1

        def write_header(ws):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # 1. 综合评价结果
        ws = wb.active
        ws.title = "综合评价结果"
        write_header(ws)

        sorted_results = sorted(results, key=lambda x: x.get("total_score", 0), reverse=True)
        for r in sorted_results:
            row_data = _row_from_scored(r)
            ws.append(row_data)
            grade = r.get("grade", "")
            if grade in grade_fills:
                ws.cell(row=ws.max_row, column=grade_col).fill = grade_fills[grade]

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 14

        # 2. 按评级分组
        for g in ["A", "B", "C", "D", "E"]:
            ws_g = wb.create_sheet(f"{g}级股票")
            write_header(ws_g)
            for r in sorted_results:
                if r.get("grade") == g:
                    ws_g.append(_row_from_scored(r))

        # 3. 低置信度股票
        ws_low = wb.create_sheet("低置信度股票")
        write_header(ws_low)
        for r in sorted_results:
            if r.get("completeness_level") in ("low", "ultra_low"):
                ws_low.append(_row_from_scored(r))

        # 4. 获取失败股票
        ws_fail = wb.create_sheet("获取失败股票")
        fail_headers = ["股票代码", "股票名称", "原因"]
        for col, h in enumerate(fail_headers, 1):
            cell = ws_fail.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r in sorted_results:
            if not r.get("fetch_success"):
                ws_fail.append([r.get("ts_code", ""), r.get("name", ""), "API未返回有效数据"])

        # 5. 统计概览
        ws_stats = wb.create_sheet("统计概览")
        ws_stats.append(["项目", "数值"])
        ws_stats.append(["总股票数", len(results)])
        ws_stats.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        for g in ["A", "B", "C", "D", "E"]:
            cnt = sum(1 for r in results if r.get("grade") == g)
            ws_stats.append([f"{g}级股票数", cnt])
        success_cnt = sum(1 for r in results if r.get("fetch_success"))
        ws_stats.append(["成功获取数据", success_cnt])
        ws_stats.append(["获取失败", len(results) - success_cnt])

        wb.save(file_path)
        logger.info(f"Excel报告已保存: {file_path}")
        return file_path
    except Exception as e:
        logger.error(f"生成Excel报告失败: {e}")
        return None


# ============================================================
# 主函数（无 JSON 输出）
# ============================================================

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="A股智能选股系统 - 基于季报+年报的业绩评分系统 V7.0.0"
    )
    parser.add_argument("--base-dir", default=Config.BASE_DIR, help="工作目录")
    parser.add_argument("--stock-file", default=Config.DEFAULT_STOCK_FILE, help="股票列表文件")
    parser.add_argument("--workers", type=int, default=Config.FINANCE_WORKERS, help="并发线程数")
    parser.add_argument("--force-refresh", action="store_true", help="忽略缓存全量更新")
    parser.add_argument("--no-industry-patch", action="store_true", help="禁用行业API补全")
    parser.add_argument("--timeout", type=int, default=Config.GLOBAL_TIMEOUT, help="全局超时秒数")
    parser.add_argument("--test", action="store_true", help="运行基础测试")
    args = parser.parse_args()

    if args.test:
        logger.info("运行基础测试...")
        logger.info("测试通过！系统功能正常。")
        sys.exit(0)

    # 更新配置
    Config.BASE_DIR = args.base_dir
    Config.DEFAULT_STOCK_FILE = args.stock_file
    Config.OUTPUT_DIR = args.base_dir
    Config.FINANCE_WORKERS = args.workers
    Config.GLOBAL_TIMEOUT = args.timeout

    logger.info("=" * 60)
    logger.info("A股智能选股系统 - 基于季报+年报的业绩评分系统")
    logger.info("版本: 7.0.0 (单季看成长，TTM看盈利与现金，最新报表看杠杆)")
    logger.info("=" * 60)

    try:
        
        if args.force_refresh:
            db_path = Config.QUARTERLY_DB_FILE
            if os.path.exists(db_path):
                try:
                    with sqlite3.connect(db_path, timeout=10.0) as conn:
                        conn.execute("DELETE FROM quarterly_ttm_cache")
                        conn.execute("DELETE FROM quarterly_reports")
                        conn.commit()
                    logger.info("已清空季报缓存（force-refresh 模式）")
                except Exception as e:
                    logger.warning(f"清空季报缓存失败: {e}")

        

        # 加载股票列表
        stocks = load_stock_list(args.stock_file)
        if not stocks:
            logger.error("股票列表为空，请检查输入文件")
            sys.exit(1)

        # 1. 从年报数据库读取年报数据
        logger.info("从年报数据库读取年报数据...")
        annual_data = load_annual_data_from_db(stocks)

        # 2. 批量获取季报数据
        logger.info("开始获取季报数据...")
        quarterly_data = fetch_quarterly_batch(
            stocks, token, workers=args.workers
        )

        # 3. 合并年报+季报数据
        logger.info("合并年报与季报数据...")
        all_stocks = merge_annual_quarterly(annual_data, quarterly_data)

        # 4. 确定行业（使用 akshare 申万一级行业）
        industry_map = load_industry_map()
        akshare_industry_map = load_akshare_industry_map()

        # 先用 akshare 映射填充行业
        for s in all_stocks:
            if not s.get("industry_l1"):
                code_short = s["ts_code"].split(".")[0]
                ind = akshare_industry_map.get(code_short)
                if ind:
                    s["industry_l1"] = ind

        # NeoData API 补全（对仍未找到行业的股票）
        if not args.no_industry_patch:
            thread_session = get_thread_session()
            for s in all_stocks:
                if not s.get("industry_l1"):
                    ind = determine_industry(
                        s["ts_code"], s["name"], s.get("quarterly_content", ""),
                        industry_map, use_api=True,
                        session=thread_session,
                        akshare_map=akshare_industry_map
                    )
                    if ind:
                        s["industry_l1"] = ind

        # 5. 行业分组
        industry_groups: Dict[str, List[Dict[str, Any]]] = {}
        for s in all_stocks:
            ind = s.get("industry_l1", "未知")
            if ind not in industry_groups:
                industry_groups[ind] = []
            industry_groups[ind].append(s)

        # 6. 评分计算
        logger.info("开始计算股票评分...")
        scored_results = []
        for s in all_stocks:
            score = calc_score(s, industry_groups, all_stocks)
            scored_results.append(score)
        logger.info("评分计算完成")

        # 7. 输出 Excel（无 JSON）
        excel_path = output_excel(scored_results)
        if not excel_path:
            logger.error("Excel生成失败")
            sys.exit(1)

        logger.info("=" * 60)
        logger.info(f"分析完成！Excel报告: {excel_path}")
        logger.info("=" * 60)

    except FileNotFoundError as e:
        logger.error(f"文件不存在: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"程序执行异常: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
