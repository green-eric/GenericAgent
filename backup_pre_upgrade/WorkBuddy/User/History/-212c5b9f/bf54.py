#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""使用东方财富API批量查询 2025年报 & 2026一季报 归母净利润增长率"""

import io, json, os, re, sys, time
import concurrent.futures

if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'buffer'):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_BASE = "https://mkapi2.dfcfs.com"
API_KEY = os.environ.get('MX_APIKEY', os.environ.get('EASTMONEY_APIKEY', ''))
HEADERS = {"Content-Type": "application/json", "apikey": API_KEY}

# 归母净利润字段ID（东方财富内部编码）
ATTR_NET_PROFIT = "100000000000308"  # 归属于母公司所有者的净利润
ATTR_NET_PROFIT2 = "100000000000200" # 归母净利润(利润表)

# =========== 股票列表 ===========
RAW = """002692 远程股份 002705 新宝股份 002718 友邦吊顶 002738 中矿资源 600103 青山纸业 600105 永鼎股份 002787 华源控股 600114 东睦股份 600118 中国卫星 002792 通宇通讯 002793 罗欣药业 600126 杭钢股份 002802 洪汇新材 002810 山东赫达 002824 和胜股份 600166 福田汽车 600176 中国巨石 002843 泰嘉股份 002850 科达利 600183 生益科技 600184 光电股份 600186 莲花控股 002866 传艺科技 600206 有研新材 600208 衢州发展 002885 京泉华 600234 科新发展 002916 深南电路 600259 中稀有色 002935 天奥电子 002937 兴瑞科技 002938 鹏鼎控股 002940 昂利康 002943 宇晶股份 002947 恒铭达 002957 科瑞技术 002962 五方光电 002975 博杰股份 002979 雷赛智能 600330 天通股份 002980 华盛昌 600337 美克家居 600338 西藏珠峰 600343 航天动力 002990 盛视科技 600345 长江通信 003018 金富科技 003022 联泓新科 003023 彩虹集团 003031 中瓷电子 003036 泰坦股份 300006 莱美药业 300012 华测检测 300051 琏升科技 300057 万顺新材 300058 蓝色光标 600482 中国动力 600487 亨通光电 600488 津药药业 600502 安徽建工 300082 奥克股份 600510 黑牡丹 300097 智云股份 600522 中天科技 600539 狮头股份 600552 凯盛科技 300131 英唐智控 300136 信维通信 300165 天瑞仪器 300179 四方达 600590 泰豪科技 600594 益佰制药 300199 翰宇药业 300204 舒泰神 300209 行云科技 300243 瑞丰高材 000026 飞亚达 600641 先导基电 300270 中威电子 000070 特发信息 600666 奥瑞德 300283 温州宏丰 600683 京投发展 000338 潍柴动力 000404 长虹华意 300308 中际旭创 600707 彩虹股份 300322 硕贝德 600724 宁波富达 300331 苏大维格 300342 天银机电 300351 永贵电器 600736 苏州高新 600743 华远控股 300382 斯莱克 600749 西藏旅游 300390 天华新能 600769 祥龙电业 600773 西藏城投 300398 飞凯材料 000570 苏常柴A 600791 京能置业 300408 三环集团 000586 汇源通信 000593 德龙汇能 300436 广生堂 300440 运达科技 000628 高新发展 300450 先导智能 300461 田中精机 300476 胜宏科技 300478 杭州高新 300480 光力科技 600869 远东股份 600879 航天电子 300489 光智科技 300502 新易盛 300503 昊志机电 300518 新迅达 300522 世名科技 600929 雪天盐业 300540 蜀道装备 000762 西藏矿业 000766 通化金马 300547 川环科技 300548 长芯博创 600955 维远股份 000782 恒申新材 600961 株冶集团 300558 贝达药业 000807 云铝股份 300580 贝斯特 600986 浙文互联 300585 奥联电子 300590 移为通信 300593 新雷能 000829 天音控股 300604 长川科技 300607 拓斯达 000880 潍柴重机 300620 光库科技 300626 华瑞股份 000889 中嘉博创 000890 法尔胜 300649 杭州园林 601133 柏诚股份 300657 弘信电子 601138 工业富联 300661 圣邦股份 300668 杰恩股份 300679 电连技术 000960 锡业股份 300681 英搏尔 300684 中石科技 300686 智动力 000967 盈峰环境 601231 环旭电子 000977 浪潮信息 300696 爱乐达 300700 岱勒新材 601339 百隆东方 000990 诚志股份 300721 怡达股份 601512 中新集团 001215 千味央厨 601579 会稽山 300736 百邦科技 300747 锐科激光 001230 劲旅环境 001234 泰慕士 300756 金马游乐 300757 罗博特科 300762 上海瀚讯 601677 明泰铝业 300767 震安科技 001268 联合精密 601698 中国卫通 300776 帝尔激光 300782 卓胜微 300788 中信出版 001309 德明利 300790 宇瞳光学 001313 粤海饲料 300801 泰和科技 601869 长飞光纤 001330 博纳影业 300806 斯迪克 601886 江河集团 300819 聚杰微纤 300821 东岳硅材 001389 广合科技 300834 星辉环材 300835 龙磁科技 300843 胜蓝股份 300845 捷安高科 002008 大族激光 603002 宏昌电子 300853 申昊科技 300857 协创数据 300858 科拓生物 603010 万盛股份 603016 新宏泰 002023 海特高新 300868 杰美特 300870 欧陆通 603026 石大胜华 002033 丽江股份 603032 德新科技 300890 翔丰华 002046 国机精工 603045 福达合金 603052 可川科技 002051 中工国际 300900 广联航空 300903 科翔股份 603061 金海通 300905 宝丽迪 603063 禾望电气 002062 宏润建设 002066 瑞泰科技 300919 中伟新材 603083 剑桥科技 002080 中材科技 002081 金螳螂 002082 万邦德 300936 中英科技 300938 信测标准 603101 汇嘉时代 300953 震裕科技 300959 线上线下 603112 华翔股份 002107 沃华医药 603115 海星股份 002108 沧州明珠 603121 华培动力 300970 华绿生物 603124 江南新材 603127 昭衍新药 002124 天邦食品 603132 金徽股份 603135 中重科技 603139 康惠股份 603150 万朗磁塑 002132 恒星科技 301002 崧盛股份 301003 江苏博云 301005 超捷股份 002149 西部材料 603178 圣龙股份 301018 申菱环境 301021 英诺激光 603193 润本股份 603196 璞源材料 603198 迎驾贡酒 002176 江特电机 603203 快克智能 301053 远信工业 301055 张小泉 603217 元利科技 002192 融捷股份 603220 中贝通信 002201 九鼎新材 002203 海亮股份 301070 开勒股份 301071 力量钻石 301079 邵阳液压 603256 宏和科技 301086 鸿富瀚 002222 福晶科技 301095 广立微 603271 永杰新材 002237 恒邦股份 002240 盛新锂能 603285 键邦股份 301110 青木科技 002245 蔚蓝锂芯 603303 得邦照明 603306 华懋科技 301123 奕东电子 603308 应流股份 301125 腾亚精工 301128 强瑞技术 002263 大东南 603315 福鞍股份 603318 水发燃气 301148 嘉戎技术 002273 水晶光电 002281 光迅科技 301157 华塑科技 002283 天润工业 603336 宏辉果蔬 002290 禾盛新材 301169 零点有数 301172 君逸数码 002297 博云新材 301181 标榜股份 301182 凯旺科技 301183 东田微 301186 超达装备 301188 力诺药包 301189 奥尼电子 301197 工大科雅 301205 联特科技 002328 新朋股份 603399 永杉锂业 301216 万凯新材 301217 铜冠铜箔 301219 腾远钴业 002338 奥普光电 002342 巨力索具 301228 实朴检测 301230 泓博医药 301232 飞沃科技 002348 高乐股份 002353 杰瑞股份 301237 和顺科技 002361 神剑股份 002364 中恒电气 301259 艾布鲁 301265 华新科技 603538 美诺华 301280 珠城科技 301282 金禄电子 002384 东山精密 301285 鸿日达 002392 北京利尔 301292 海科新源 301295 美硕科技 301297 富乐德 603601 再升科技 301306 西测测试 603608 天创时尚 301310 鑫宏业 301312 智立方 301317 鑫磊股份 301319 唯特偶 301321 翰博高新 603618 杭电股份 301322 绿通科技 301323 新莱福 301326 捷邦科技 603629 利通电子 301328 维峰电子 603637 镇海股份 301345 涛涛车业 002418 康盛股份 301360 荣旗科技 301362 民爆光电 603668 天马科技 002428 云南锗业 301366 一博科技 002429 兆驰股份 301369 联动科技 002432 九安医疗 301371 敷尔佳 002436 兴森科技 301372 科净源 301373 凌玮科技 301377 鼎泰高科 603687 大胜达 603688 石英股份 002443 金洲管道 301387 光大同创 603698 航天工程 301392 汇成真空 301393 昊帆生物 002454 松芝股份 301396 宏景科技 301397 溯联股份 002458 益生股份 301408 华人健康 002463 沪电股份 002466 天齐锂业 002468 申通快递 002471 中超控股 301486 致尚科技 002475 立讯精密 603738 泰晶科技 301489 思泉新材 603757 大元泵业 301500 飞南资源 002484 江海股份 603773 沃格光电 301511 德福科技 002491 通鼎互联 301517 陕西华达 603798 康普顿 002497 雅化集团 603800 洪田股份 301526 国际复材 301528 多浦乐 301548 崇德科技 603826 坤彩科技 301566 达利凯普 301603 乔锋智能 301607 富特科技 002536 飞龙股份 603890 春秋电子 301629 矽电股份 603897 长城科技 301631 壹连科技 002552 宝鼎科技 603906 龙蟠科技 002560 通达股份 002565 顺灏股份 603936 博敏电子 603938 三孚股份 603950 长源东谷 002580 圣阳股份 603958 哈森股份 603966 法兰泰克 603985 恒润股份 605006 山东玻纤 605018 长华集团 605055 迎丰股份 002635 安洁科技 002636 金安国纪 605098 行动教育 605100 华丰股份 002645 华宏科技 605133 嵘泰股份 002647 仁东控股 605168 三人行 002655 共达电声 605189 富春染织 605198 安德利 605222 起帆电缆 605298 必得科技 605303 园林股份 605318 法狮龙 605365 立达信 605366 宏柏新材 605376 博迁新材 605566 福莱蒽特 605589 圣泉集团"""

def parse_stocks():
    stocks = []
    parts = RAW.split()
    i = 0
    while i < len(parts) - 1:
        code, name = parts[i], parts[i+1]
        i += 2
        if code.startswith('688'):
            continue
        if code.startswith(('000','001','002','003','300','301')):
            ts_code = f"{code}.SZ"
        elif code.startswith(('600','601','603','605')):
            ts_code = f"{code}.SH"
        else:
            continue
        stocks.append({'code': code, 'name': name, 'ts_code': ts_code})
    return stocks

def query_em(tool_query, timeout=20):
    """调用东方财富数据查询API"""
    url = f"{API_BASE}/finskillshub/api/claw/query"
    payload = {"toolQuery": tool_query}
    try:
        r = requests.post(url, headers=HEADERS, json=payload, timeout=timeout)
        return r.json()
    except:
        return None

def extract_profits(data, target_name):
    """从东方财富API返回中提取归母净利润"""
    if not data or data.get('status') != 0:
        return None, None
    
    try:
        d = data['data']['data']['searchDataResultDTO']['dataTableDTOList'][0]
        table = d.get('table', {})
        raw = d.get('rawTable', {})
        name_map = d.get('nameMap', {})
        heads = table.get('headName', [])
        
        # 找归母净利润字段ID
        profit_ids = []
        for fid, fname in name_map.items():
            if '归母' in fname or '归属' in fname:
                profit_ids.append(fid)
        
        if not profit_ids:
            profit_ids = [ATTR_NET_PROFIT, ATTR_NET_PROFIT2]
        
        # 从rawTable获取精确数值
        val1, val2 = None, None
        for pid in profit_ids:
            vals = raw.get(pid, [])
            if len(vals) >= 2:
                try:
                    v1 = float(vals[0])
                    v2 = float(vals[1])
                    return v1, v2  # 返回原始元值
                except:
                    continue
        
        # fallback: 从table获取（带单位的字符串）
        for pid in profit_ids:
            vals = table.get(pid, [])
            if len(vals) >= 2:
                v1 = parse_chinese_num(vals[0])
                v2 = parse_chinese_num(vals[1])
                if v1 is not None and v2 is not None:
                    return v1, v2
    except Exception as e:
        pass
    return None, None

def parse_chinese_num(s):
    """解析 '32.76亿元' 等中文数字"""
    if not s:
        return None
    m = re.match(r'([\d.]+)亿元', s)
    if m:
        return float(m.group(1)) * 1e8
    m = re.match(r'([\d.]+)万元', s)
    if m:
        return float(m.group(1)) * 1e4
    m = re.match(r'([\d.]+)元', s)
    if m:
        return float(m.group(1))
    return None

def process_stock(stock):
    """查询单只股票的年报和季报增长率"""
    code = stock['code']
    name = stock['name']
    ts_code = stock['ts_code']
    
    result = {
        'code': code, 'name': name, 'ts_code': ts_code,
        'p_2025ann': None, 'p_2024ann': None,
        'p_2026q1': None, 'p_2025q1': None,
        'growth_2025ann': None, 'growth_2026q1': None,
        'ann_head': '', 'q1_head': '',
    }
    
    # 查询年报
    ann_data = query_em(f"{name}{ts_code} 年报归母净利润")
    p1, p2 = extract_profits(ann_data, name)
    if p1 is not None and p2 is not None:
        result['p_2025ann'] = round(p1 / 1e8, 2)
        result['p_2024ann'] = round(p2 / 1e8, 2)
        if p2 != 0:
            result['growth_2025ann'] = round((p1 - p2) / abs(p2) * 100, 1)
        else:
            result['growth_2025ann'] = 9999.0
    # 提取表头
    try:
        if ann_data and ann_data.get('status') == 0:
            heads = ann_data['data']['data']['searchDataResultDTO']['dataTableDTOList'][0]['table']['headName']
            result['ann_head'] = ' / '.join(heads)
    except:
        pass
    
    time.sleep(0.2)
    
    # 查询一季报
    q1_data = query_em(f"{name}{ts_code} 一季报归母净利润")
    p1, p2 = extract_profits(q1_data, name)
    if p1 is not None and p2 is not None:
        result['p_2026q1'] = round(p1 / 1e8, 2)
        result['p_2025q1'] = round(p2 / 1e8, 2)
        if p2 != 0:
            result['growth_2026q1'] = round((p1 - p2) / abs(p2) * 100, 1)
        else:
            result['growth_2026q1'] = 9999.0
    try:
        if q1_data and q1_data.get('status') == 0:
            heads = q1_data['data']['data']['searchDataResultDTO']['dataTableDTOList'][0]['table']['headName']
            result['q1_head'] = ' / '.join(heads)
    except:
        pass
    
    return result

def main():
    stocks = parse_stocks()
    total = len(stocks)
    print(f"Total: {total} stocks (excl 688)")
    
    results = []
    completed = 0
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
        future_map = {executor.submit(process_stock, s): s for s in stocks}
        for future in concurrent.futures.as_completed(future_map):
            completed += 1
            r = future.result()
            results.append(r)
            if completed % 30 == 0 or completed == total:
                g25 = f"{r['growth_2025ann']}%" if r['growth_2025ann'] is not None else "N/A"
                g26 = f"{r['growth_2026q1']}%" if r['growth_2026q1'] is not None else "N/A"
                print(f"[{completed}/{total}] {r['name']}({r['code']}) 25Ann:{g25} 26Q1:{g26}")
    
    results.sort(key=lambda x: x['code'])
    
    # 筛选
    high_2025 = [r for r in results if r['growth_2025ann'] is not None and r['growth_2025ann'] > 50]
    high_2026q1 = [r for r in results if r['growth_2026q1'] is not None and r['growth_2026q1'] > 50]
    high_either = [r for r in results if (r['growth_2025ann'] is not None and r['growth_2025ann'] > 50) or (r['growth_2026q1'] is not None and r['growth_2026q1'] > 50)]
    
    high_2025.sort(key=lambda x: x['growth_2025ann'], reverse=True)
    high_2026q1.sort(key=lambda x: x['growth_2026q1'], reverse=True)
    high_either.sort(key=lambda x: max(x['growth_2025ann'] or 0, x['growth_2026q1'] or 0), reverse=True)
    
    print(f"\n===== SUMMARY =====")
    print(f"Total: {total}")
    print(f"2025 Ann >50%: {len(high_2025)}")
    print(f"2026 Q1 >50%: {len(high_2026q1)}")
    print(f"\n--- 2025年报增长>50% ---")
    for r in high_2025[:20]:
        print(f"  {r['name']}({r['code']}) {r['growth_2025ann']}%")
    print(f"\n--- 2026一季报增长>50% ---")
    for r in high_2026q1[:20]:
        print(f"  {r['name']}({r['code']}) {r['growth_2026q1']}%")
    
    # ===== 生成 Excel =====
    wb = Workbook()
    
    HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    TITLE_FONT = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
    SUB_FONT = Font(name='微软雅黑', bold=True, size=12, color='2F5496')
    NORM = Font(name='微软雅黑', size=10)
    BOLD = Font(name='微软雅黑', bold=True, size=10)
    RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def hdr(ws, row, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    
    def sty(ws, r1, r2, nc):
        for r in range(r1, r2+1):
            for c in range(1, nc+1):
                cell = ws.cell(row=r, column=c)
                cell.font = NORM; cell.alignment = CENTER if c <= 3 else LEFT; cell.border = BORDER
    
    def wid(ws, nc, mx=45):
        for c in range(1, nc+1):
            mx_l = 8
            for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
                for v in row:
                    if v: mx_l = max(mx_l, min(len(str(v))*1.3, mx))
            ws.column_dimensions[get_column_letter(c)].width = mx_l
    
    # ---- Sheet1: 2025年报增长>50% ----
    ws1 = wb.active; ws1.title = "2025年报增长>50%"
    ws1.merge_cells('A1:H1')
    ws1.cell(row=1, column=1, value=f"2025年报归母净利润增长率 > 50%（共{len(high_2025)}只，排除科创板）").font = TITLE_FONT
    ws1.cell(row=1, column=1).alignment = CENTER
    r = 3
    hdr(ws1, r, ["排名","股票名称","代码","2025年报净利润(亿)","2024年报净利润(亿)","增长率(%)","报告期","数据来源"])
    for i, s in enumerate(high_2025, 1):
        r += 1
        ws1.cell(row=r, column=1, value=i)
        ws1.cell(row=r, column=2, value=s['name'])
        ws1.cell(row=r, column=3, value=s['code'])
        ws1.cell(row=r, column=4, value=s['p_2025ann'])
        ws1.cell(row=r, column=5, value=s['p_2024ann'])
        ws1.cell(row=r, column=6, value=s['growth_2025ann'])
        ws1.cell(row=r, column=7, value=s['ann_head'])
        ws1.cell(row=r, column=8, value="东方财富")
        if s['growth_2025ann'] and s['growth_2025ann'] > 50:
            ws1.cell(row=r, column=6).fill = RED_FILL
    sty(ws1, 4, r, 8); wid(ws1, 8)
    
    # ---- Sheet2: 2026一季报增长>50% ----
    ws2 = wb.create_sheet("2026一季报增长>50%")
    ws2.merge_cells('A1:H1')
    ws2.cell(row=1, column=1, value=f"2026一季报归母净利润增长率 > 50%（共{len(high_2026q1)}只，排除科创板）").font = TITLE_FONT
    ws2.cell(row=1, column=1).alignment = CENTER
    r = 3
    hdr(ws2, r, ["排名","股票名称","代码","2026Q1净利润(亿)","2025Q1净利润(亿)","增长率(%)","报告期","数据来源"])
    for i, s in enumerate(high_2026q1, 1):
        r += 1
        ws2.cell(row=r, column=1, value=i)
        ws2.cell(row=r, column=2, value=s['name'])
        ws2.cell(row=r, column=3, value=s['code'])
        ws2.cell(row=r, column=4, value=s['p_2026q1'])
        ws2.cell(row=r, column=5, value=s['p_2025q1'])
        ws2.cell(row=r, column=6, value=s['growth_2026q1'])
        ws2.cell(row=r, column=7, value=s['q1_head'])
        ws2.cell(row=r, column=8, value="东方财富")
        if s['growth_2026q1'] and s['growth_2026q1'] > 50:
            ws2.cell(row=r, column=6).fill = RED_FILL
    sty(ws2, 4, r, 8); wid(ws2, 8)
    
    # ---- Sheet3: 任一增长>50% ----
    ws3 = wb.create_sheet("合并(任一>50%)")
    ws3.merge_cells('A1:I1')
    ws3.cell(row=1, column=1, value=f"2025年报或2026一季报增长率 > 50%（共{len(high_either)}只，排除科创板）").font = TITLE_FONT
    ws3.cell(row=1, column=1).alignment = CENTER
    r = 3
    hdr(ws3, r, ["排名","股票名称","代码","2025年报增长率(%)","2026Q1增长率(%)","2025年报净利(亿)","2024年报净利(亿)","2026Q1净利(亿)","2025Q1净利(亿)"])
    for i, s in enumerate(high_either, 1):
        r += 1
        ws3.cell(row=r, column=1, value=i)
        ws3.cell(row=r, column=2, value=s['name'])
        ws3.cell(row=r, column=3, value=s['code'])
        ws3.cell(row=r, column=4, value=s['growth_2025ann'])
        ws3.cell(row=r, column=5, value=s['growth_2026q1'])
        ws3.cell(row=r, column=6, value=s['p_2025ann'])
        ws3.cell(row=r, column=7, value=s['p_2024ann'])
        ws3.cell(row=r, column=8, value=s['p_2026q1'])
        ws3.cell(row=r, column=9, value=s['p_2025q1'])
        if s['growth_2025ann'] and s['growth_2025ann'] > 50:
            ws3.cell(row=r, column=4).fill = RED_FILL
        if s['growth_2026q1'] and s['growth_2026q1'] > 50:
            ws3.cell(row=r, column=5).fill = RED_FILL
    sty(ws3, 4, r, 9); wid(ws3, 9)
    
    # ---- Sheet4: 全量数据 ----
    ws4 = wb.create_sheet("全部股票明细")
    ws4.merge_cells('A1:I1')
    ws4.cell(row=1, column=1, value=f"全部{total}只股票增长率明细（排除科创板）").font = TITLE_FONT
    ws4.cell(row=1, column=1).alignment = CENTER
    r = 3
    hdr(ws4, r, ["股票名称","代码","2025年报增长(%)","2026Q1增长(%)","2025年报净利(亿)","2024年报净利(亿)","2026Q1净利(亿)","2025Q1净利(亿)","年报期次"])
    for s in results:
        r += 1
        ws4.cell(row=r, column=1, value=s['name'])
        ws4.cell(row=r, column=2, value=s['code'])
        ws4.cell(row=r, column=3, value=s['growth_2025ann'])
        ws4.cell(row=r, column=4, value=s['growth_2026q1'])
        ws4.cell(row=r, column=5, value=s['p_2025ann'])
        ws4.cell(row=r, column=6, value=s['p_2024ann'])
        ws4.cell(row=r, column=7, value=s['p_2026q1'])
        ws4.cell(row=r, column=8, value=s['p_2025q1'])
        ws4.cell(row=r, column=9, value=s['ann_head'])
        # 标色
        g25 = s['growth_2025ann']
        g26 = s['growth_2026q1']
        if g25 is not None and g25 > 50:
            ws4.cell(row=r, column=3).fill = RED_FILL
        elif g25 is not None and g25 < 0:
            ws4.cell(row=r, column=3).fill = GREEN_FILL
        if g26 is not None and g26 > 50:
            ws4.cell(row=r, column=4).fill = RED_FILL
        elif g26 is not None and g26 < 0:
            ws4.cell(row=r, column=4).fill = GREEN_FILL
        if g25 is None and g26 is None:
            for c in range(1, 10):
                ws4.cell(row=r, column=c).font = Font(name='微软雅黑', size=10, color='999999')
    sty(ws4, 4, r, 9); wid(ws4, 9)
    
    # 保存
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    outpath = os.path.join(desktop, "股票增长率分析_2025年报_2026一季报_排除科创板.xlsx")
    wb.save(outpath)
    print(f"\nExcel saved: {outpath}")

if __name__ == '__main__':
    main()
