#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证 Excel 数据: 从Excel读取 + NeoData API 对比
"""
import openpyxl
import requests
import os
import time
import re

# === 1. 读取Excel ===
wb = openpyxl.load_workbook(r'D:\Project\AnnualScorer\股票业绩评价_20260426_204545.xlsx')
ws = wb.active
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

# 找2只特定股票的行
targets = ["603350.SH", "000852.SZ"]  # 安乃达, 石化机械
excel_data = {}
for r in range(2, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if code in targets:
        row = {}
        for i, h in enumerate(headers, 1):
            row[h] = ws.cell(r, i).value
        excel_data[code] = row

# === 2. NeoData API 查询 ===
token_file = os.path.expanduser("~/.workbuddy/.neodata_token")
with open(token_file, "r", encoding="utf-8") as f:
    token = f.read().strip()

URL = "https://copilot.tencent.com/agenttool/v1/neodata"
hdrs = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def neodata_query(q):
    for attempt in range(3):
        try:
            resp = requests.post(URL, json={"query": q}, headers=hdrs, timeout=50)
            data = resp.json()
            if data.get("code") == "200":
                return data.get("data", {}).get("content", "")
            print(f"  API错误: {data.get('msg')}")
            time.sleep(2)
        except Exception as e:
            print(f"  异常: {e}")
            time.sleep(2)
    return ""

def extract_metric(text, keywords):
    """从API返回文本中提取指标值"""
    if not text:
        return None
    for line in text.split("\n"):
        line = line.strip()
        if any(kw in line for kw in keywords):
            # 提取数字
            nums = re.findall(r'[-+]?\d+\.?\d*', line)
            if nums:
                return float(nums[0])
    return None

# === 3. 对比验证 ===
for ts_code in targets:
    row = excel_data.get(ts_code, {})
    name = row.get("股票名称", "")
    
    print(f"\n{'='*60}")
    print(f"股票: {ts_code} {name}")
    print(f"{'='*60}")
    
    # Excel数据
    print(f"\n  [Excel数据]")
    print(f"    行业: {row.get('申万一级行业', 'N/A')}")
    print(f"    ROE: {row.get('ROE(%)', 'N/A')}%")
    print(f"    毛利率: {row.get('毛利率(%)', 'N/A')}%")
    print(f"    净利率: {row.get('净利率(%)', 'N/A')}%")
    print(f"    营收同比: {row.get('营收同比(%)', 'N/A')}%")
    print(f"    净利润同比: {row.get('净利润同比(%)', 'N/A')}%")
    print(f"    资产负债率: {row.get('资产负债率(%)', 'N/A')}%")
    print(f"    归母净利润: {row.get('净利润(元)', 'N/A')}")
    print(f"    经营现金流: {row.get('经营现金流(元)', 'N/A')}")
    print(f"    OCF/净利润: {row.get('OCF/净利润(%)', 'N/A')}%")
    print(f"    总分: {row.get('总分', 'N/A')}  评级: {row.get('评级', 'N/A')}")
    print(f"    置信度: {row.get('置信度', 'N/A')}")
    
    # API数据
    print(f"\n  [NeoData API 查询中...]")
    content = neodata_query(f"{ts_code} {name} 2024年报")
    
    if content:
        print(f"  [API返回关键数据]")
        for line in content.split("\n"):
            line = line.strip()
            if any(kw in line for kw in ["统计截止日期", "营业收入", "归母净利润", "毛利率", "净利率", "ROE", "净资产收益率", "资产负债率", "经营活动现金流"]):
                print(f"    {line}")
        
        # 提取对比
        api_roe = extract_metric(content, ["净资产收益率", "ROE"])
        api_gross = extract_metric(content, ["毛利率"])
        api_net = extract_metric(content, ["净利率"])
        api_debt = extract_metric(content, ["资产负债率"])
        
        print(f"\n  [对比验证]")
        excel_roe = row.get('ROE(%)')
        if api_roe and excel_roe:
            diff = abs(float(api_roe) - float(excel_roe))
            status = "✅" if diff < 1 else f"⚠️ 差{diff:.2f}"
            print(f"    ROE: Excel={excel_roe}% API={api_roe}% {status}")
        
        excel_gross = row.get('毛利率(%)')
        if api_gross and excel_gross:
            diff = abs(float(api_gross) - float(excel_gross))
            status = "✅" if diff < 1 else f"⚠️ 差{diff:.2f}"
            print(f"    毛利率: Excel={excel_gross}% API={api_gross}% {status}")
        
        excel_net = row.get('净利率(%)')
        if api_net and excel_net:
            diff = abs(float(api_net) - float(excel_net))
            status = "✅" if diff < 1 else f"⚠️ 差{diff:.2f}"
            print(f"    净利率: Excel={excel_net}% API={api_net}% {status}")
        
        excel_debt = row.get('资产负债率(%)')
        if api_debt and excel_debt:
            diff = abs(float(api_debt) - float(excel_debt))
            status = "✅" if diff < 1 else f"⚠️ 差{diff:.2f}"
            print(f"    资产负债率: Excel={excel_debt}% API={api_debt}% {status}")
    else:
        print("  API查询失败")
    
    time.sleep(3)

print(f"\n{'='*60}")
print("验证完成")
