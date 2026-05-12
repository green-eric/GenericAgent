#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import random
import json
import os
import sys

# 加载token
token_file = os.path.expanduser("~/.workbuddy/.neodata_token")
with open(token_file, "r", encoding="utf-8") as f:
    token = f.read().strip()

# 加载Excel
wb = openpyxl.load_workbook(r'D:\Project\AnnualScorer\股票业绩评价_20260426_204545.xlsx')
ws = wb.active
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

print('=== Excel 结构 ===')
print('Sheet:', wb.sheetnames)
print('总行数:', ws.max_row)
print('列头:', headers)
print()

# 找有效数据行
valid_rows = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 2).value
    score = ws.cell(r, 3).value
    if name and score is not None:
        valid_rows.append(r)

print(f'有效数据行: {len(valid_rows)}')
print()

# 随机选2只
chosen = random.sample(valid_rows, min(2, len(valid_rows)))

for r in chosen:
    print(f'=== 行{r} ===')
    row_data = {}
    for i, h in enumerate(headers, 1):
        val = ws.cell(r, i).value
        row_data[h] = val
        print(f'  {h}: {val}')
    
    # 打印关键验证字段
    ts_code = row_data.get('股票代码', '')
    name = row_data.get('股票名称', '')
    score = row_data.get('评分', '')
    grade = row_data.get('评级', '')
    industry = row_data.get('行业', '')
    roe = row_data.get('ROE(%)', '')
    net_profit = row_data.get('归母净利润(元)', '')
    revenue = row_data.get('营业收入(元)', '')
    
    print()
    print(f'  [验证摘要] {ts_code} {name}')
    print(f'    评分: {score}  评级: {grade}  行业: {industry}')
    print(f'    ROE: {roe}%  归母净利润: {net_profit}  营收: {revenue}')
    print()
