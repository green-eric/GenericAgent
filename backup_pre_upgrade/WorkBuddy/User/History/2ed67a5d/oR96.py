#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# 业绩分析系统 V5.0.0 

import os, re, json, time, sys, logging, argparse, sqlite3, threading, glob
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED, ALL_COMPLETED, Future
from typing import Dict, List, Optional, Any
import subprocess
import requests as req_lib
from logging.handlers import RotatingFileHandler

# ====================== 解决控制台中文乱码（使用 UTF-8）======================
if hasattr(sys.stdout, 'buffer'):
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
elif hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# ====================== 配置中心 ======================
class Config:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    QUERY_SCRIPT = os.environ.get(
        "NEODATA_QUERY_PATH",
        os.path.expanduser("~/.workbuddy/plugins/marketplaces/cb_teams_marketplace/"
                           "plugins/finance-data/skills/neodata-financial-search/scripts/query.py")
    )
    DEFAULT_STOCK_FILE = os.path.join(BASE_DIR, "xuan.txt")
    OUTPUT_DIR = BASE_DIR
    INDUSTRY_MAP_FILE = os.path.join(BASE_DIR, "industry_map.json")
    DB_FILE = os.path.join(BASE_DIR, "stock_cache.db")
    INDUSTRY_MAP_URL = None

    FINANCE_WORKERS = 16
    API_RETRY_TIMES = 2
    API_RETRY_BACKOFF_BASE = 3.0
    API_TIMEOUT = 50
    GLOBAL_DELAY = 0.5
    PAUSE_CONSECUTIVE_EMPTY = 10
    PAUSE_DURATION = 20
    GLOBAL_TIMEOUT = 7200

    MIN_INDUSTRY_SAMPLES = 5
    SCORE_WEIGHTS = {"profit":0.35, "growth":0.30, "ocf_quality":0.15, "debt_risk":0.20}
    PROFIT_SUB = {"roe":0.4, "gross_margin":0.3, "net_margin":0.3}
    GROWTH_SUB = {"revenue_yoy":0.4, "profit_yoy":0.6}
    NEGATIVE_PROFIT_PENALTY = 15.0
    MARKET_FALLBACK_DISCOUNT = 0.95
    LOW_COMPLETENESS_PENALTY = 0.9
    COMPLETENESS_HIGH = 0.7
    COMPLETENESS_LOW = 0.4
    ULTRA_LOW_COMPLETENESS_THRESHOLD = 1      # ≤1 指标
    ULTRA_LOW_COMPLETENESS_PENALTY = 0.75     # 额外折扣

    RETRY_INDUSTRY_FOR_UNCLASSIFIED = True
    INDUSTRY_API_WORKERS = 15
    FORCE_REFRESH = False

    CACHE_MAX_AGE_ANNUAL = 400                 # 仅保留年报缓存
    INDUSTRY_CACHE_DAYS = 365
    ANNUAL_DISCLOSURE_DEADLINE_MONTH = 4
    ANNUAL_DISCLOSURE_DEADLINE_DAY = 30

    LOG_FILE = os.path.join(BASE_DIR, "stock_analyzer.log")
    LOG_MAX_BYTES = 10 * 1024 * 1024
    LOG_BACKUP_COUNT = 5
    LOG_CLEANUP_DAYS = 30

# ====================== 全局异常日志（加锁） ======================
error_log: List[Dict] = []
_error_lock = threading.Lock()
logger = None

def safe_error_append(entry: dict):
    with _error_lock:
        error_log.append(entry)

# ====================== 内置最小兜底行业表（仅 100 条，完整表请使用 industry_map.json） ======================
_FALLBACK_INDUSTRY_MAP = {
    "000001":"银行","000002":"房地产","000004":"医药生物","000005":"公用事业",
    "000006":"房地产","000007":"房地产","000008":"机械设备","000009":"有色金属",
    "000010":"建筑装饰","000011":"房地产","000012":"建筑材料","000014":"房地产",
    "000016":"家用电器","000017":"汽车","000019":"食品饮料","000020":"电子",
    "000021":"电子","000025":"汽车","000026":"纺织服饰","000027":"公用事业",
    "000028":"医药生物","000029":"房地产","000030":"汽车","000031":"房地产",
    "000032":"计算机","000034":"计算机","000035":"环保","000036":"房地产",
    "000037":"公用事业","000038":"计算机","000039":"机械设备","000040":"公用事业",
    "000042":"房地产","000043":"国防军工","000045":"电子","000046":"房地产",
    "000048":"农林牧渔","000049":"电子","000050":"电子","000055":"建筑装饰",
    "000056":"房地产","000058":"电子","000059":"基础化工","000060":"有色金属",
    "000061":"农林牧渔","000062":"电子","000063":"通信","000065":"建筑装饰",
    "000066":"计算机","000068":"电子","000069":"房地产","000070":"通信",
    "000078":"医药生物","000088":"交通运输","000089":"交通运输","000090":"房地产",
    "000096":"石油石化","000099":"交通运输","000100":"电子","000150":"医药生物",
    "000151":"机械设备","000153":"医药生物","000155":"公用事业","000156":"传媒",
    "000157":"机械设备","000158":"纺织服饰","000159":"基础化工","000166":"非银金融",
    "000168":"医药生物","000170":"机械设备","000171":"有色金属","000172":"纺织服饰",
    "000173":"医药生物","000175":"汽车","000176":"电力设备","000177":"纺织服饰",
    "000178":"电子","000179":"交通运输","000180":"医药生物","000181":"环保",
    "000182":"建筑材料","000183":"医药生物","000184":"国防军工","000185":"电力设备",
    "000186":"食品饮料","000187":"通信","000188":"交通运输","000189":"建筑装饰",
    "000190":"交通运输","000191":"农林牧渔","000192":"有色金属","000193":"建筑装饰",
    "000195":"汽车","000196":"医药生物","000197":"食品饮料","000198":"通信",
    "000199":"医药生物","000200":"汽车",
}

# ====================== 其他静态表 ======================
SECONDARY_TO_PRIMARY = {
    "化学制药":"医药生物","生物制品":"医药生物","医疗器械":"医药生物","医药商业":"医药生物",
    "白酒":"食品饮料","啤酒":"食品饮料","乳品":"食品饮料","调味品":"食品饮料",
    "股份制银行":"银行","城商行":"银行","农商行":"银行","国有大型银行":"银行",
    "集成电路":"电子","分立器件":"电子","印制电路板":"电子","LED":"电子",
    "软件开发":"计算机","IT服务":"计算机","安防设备":"计算机",
    "动力电池":"电力设备","光伏设备":"电力设备","风电设备":"电力设备","电网设备":"电力设备",
    "证券":"非银金融","保险":"非银金融","多元金融":"非银金融",
    "住宅开发":"房地产","商业地产":"房地产","园区开发":"房地产",
    "农药":"基础化工","化肥":"基础化工","化学纤维":"基础化工","塑料":"基础化工",
    "工程机械":"机械设备","专用设备":"机械设备","通用设备":"机械设备",
    "乘用车":"汽车","商用车":"汽车","汽车零部件":"汽车",
    "航空装备":"国防军工","航天装备":"国防军工","地面兵装":"国防军工",
    "工业金属":"有色金属","贵金属":"有色金属","能源金属":"有色金属",
    "普钢":"钢铁","特钢":"钢铁",
    "水泥":"建筑材料","玻璃":"建筑材料",
    "种植业":"农林牧渔","养殖业":"农林牧渔","饲料":"农林牧渔",
    "纺织制造":"纺织服饰","服装家纺":"纺织服饰",
    "铁路运输":"交通运输","港口":"交通运输","机场":"交通运输",
    "火电":"公用事业","水电":"公用事业","风电":"公用事业","光伏发电":"公用事业",
}

NAME_KEYWORD_INDUSTRY = {
    "银行":"银行","医药":"医药生物","医疗":"医药生物","生物":"医药生物",
    "证券":"非银金融","保险":"非银金融","金融":"非银金融",
    "地产":"房地产","置业":"房地产","园区":"房地产",
    "汽车":"汽车","客车":"汽车","车":"汽车",
    "钢铁":"钢铁","钢":"钢铁",
    "化工":"基础化工","化学":"基础化工","化纤":"基础化工",
    "机械":"机械设备","重工":"机械设备",
    "航空":"国防军工","航天":"国防军工","军工":"国防军工",
    "电力":"电力设备","电气":"电力设备","新能源":"电力设备","光伏":"电力设备","锂电":"电力设备",
    "食品":"食品饮料","饮料":"食品饮料","酒":"食品饮料","奶":"食品饮料",
    "水泥":"建筑材料","玻璃":"建筑材料","建材":"建筑材料",
    "煤":"煤炭","炭":"煤炭","焦":"煤炭",
    "石油":"石油石化","石化":"石油石化",
    "软件":"计算机","信息":"计算机","数据":"计算机",
    "芯片":"电子","电子":"电子","光电":"电子","半导体":"电子",
    "通信":"通信","电信":"通信",
    "农牧":"农林牧渔","畜牧":"农林牧渔","种业":"农林牧渔","粮食":"农林牧渔",
    "珠宝":"纺织服饰","服装":"纺织服饰","纺织":"纺织服饰",
    "旅游":"社会服务","酒店":"社会服务","教育":"社会服务",
    "环保":"环保",
}
CODE_PREFIX_INDUSTRY = {
    "600":"银行","601":"银行",
    "602":"基础化工","603":"机械设备",
    "605":"轻工制造","000":"房地产","001":"基础化工","002":"电子",
    "003":"纺织服饰","300":"医药生物","301":"电力设备",
}

def standardize_industry(raw: str) -> Optional[str]:
    if not raw: return None
    raw = raw.strip().replace("申万", "").replace("SW", "").replace("sw", "")
    mapping = {
        "医药":"医药生物","银行":"银行","食品":"食品饮料","饮料":"食品饮料",
        "电力":"电力设备","新能源":"电力设备","计算机":"计算机","电子":"电子",
        "非银":"非银金融","房地产":"房地产","汽车":"汽车","机械":"机械设备",
        "化工":"基础化工","国防军工":"国防军工","军工":"国防军工",
        "有色金属":"有色金属","钢铁":"钢铁","建筑材料":"建筑材料",
        "农林牧渔":"农林牧渔","纺织服饰":"纺织服饰","公用事业":"公用事业",
        "交通运输":"交通运输","建筑装饰":"建筑装饰","家用电器":"家用电器",
        "商贸零售":"商贸零售","社会服务":"社会服务","传媒":"传媒",
        "环保":"环保","通信":"通信","轻工制造":"轻工制造",
        "煤炭":"煤炭","石油石化":"石油石化",
    }
    return mapping.get(raw, raw)

# ====================== 日志 ======================
def cleanup_old_logs():
    try:
        log_pattern = os.path.join(Config.BASE_DIR, "stock_analyzer*.log*")
        for f in glob.glob(log_pattern):
            if os.path.basename(f) == os.path.basename(Config.LOG_FILE):
                continue
            age = time.time() - os.path.getmtime(f)
            if age > Config.LOG_CLEANUP_DAYS * 24 * 3600:
                os.remove(f)
    except Exception:
        pass

def setup_logging():
    log_file = Config.LOG_FILE
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)
    rotating = RotatingFileHandler(
        log_file, maxBytes=Config.LOG_MAX_BYTES,
        backupCount=Config.LOG_BACKUP_COUNT, encoding='utf-8'
    )
    stream = logging.StreamHandler(sys.stdout)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[rotating, stream]
    )
    cleanup_old_logs()
    return logging.getLogger(__name__)

# ====================== 行业映射文件加载 ======================
def get_industry_map() -> dict:
    if os.path.exists(Config.INDUSTRY_MAP_FILE):
        try:
            with open(Config.INDUSTRY_MAP_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict) and len(data) > 100:
                    logger.info(f"已加载外部行业映射表，{len(data)} 条")
                    return data
        except Exception as e:
            logger.warning(f"加载外部映射表失败: {e}")
    logger.warning("使用内置最小兜底行业表（仅 100 条），建议放置完整 industry_map.json")
    return _FALLBACK_INDUSTRY_MAP

# ====================== API 调用 ======================
_ND_ENDPOINT = "https://copilot.tencent.com/agenttool/v1/neodata"
_ND_TOKEN_FILE = os.path.join(os.path.expanduser("~"), ".workbuddy", ".neodata_token")

def _nd_read_token() -> str:
    try:
        with open(_ND_TOKEN_FILE, 'r') as f:
            t = f.read().strip()
            if t: return t
    except: pass
    logger.error("NeoData token 未找到")
    return ""

def run_neodata(query: str, session: req_lib.Session = None) -> dict:
    """调用 NeoData API，支持传入 requests.Session 复用连接"""
    token = _nd_read_token()
    if not token:
        return {"code":"error","msg":"no_token"}
    headers = {"Content-Type":"application/json", "Authorization":f"Bearer {token}"}
    payload = {"query":query, "channel":"neodata", "sub_channel":"workbuddy", "data_type":"api"}
    for attempt in range(1, Config.API_RETRY_TIMES+2):
        try:
            if session is None:
                resp = req_lib.post(_ND_ENDPOINT, headers=headers, json=payload, timeout=Config.API_TIMEOUT)
            else:
                resp = session.post(_ND_ENDPOINT, headers=headers, json=payload, timeout=Config.API_TIMEOUT)
            resp.raise_for_status()
            data = resp.json()
            if data.get("code") == "200": return data
            time.sleep(Config.API_RETRY_BACKOFF_BASE)
        except req_lib.Timeout:
            time.sleep(Config.API_RETRY_BACKOFF_BASE * attempt)
        except req_lib.RequestException as e:
            logger.warning(f"API调用异常: {e}")
            time.sleep(Config.API_RETRY_BACKOFF_BASE)
    return {"code":"error","msg":"exhausted"}

# ====================== 数值与文本提取 ======================


def extract_industry_from_content(content: str) -> Optional[str]:
    if not content: return None
    patterns = [
        r'所属一级行业[：:]\s*([^\s，,、；;]+)',
        r'申万.*?行业[：:]\s*([^\s，,、；;]+)',
        r'一级行业[为是]?\s*([^\s，,、；;]+)',
        r'行业分类[：:]\s*([^\s，,、；;]+)',
        r'(?:申万|SW)\s*([\u4e00-\u9fa5]+)行业',
        r'属于\s*([\u4e00-\u9fa5]+)行业',
        r'所处行业[为是]?([\u4e00-\u9fa5]+)',
        r'主营业务为([\u4e00-\u9fa5]+)',
    ]
    for pat in patterns:
        m = re.search(pat, content)
        if m:
            raw = m.group(1).strip()
            if 2<=len(raw)<=15: return standardize_industry(raw)
    sec_m = re.search(r'所属二级行业[：:]\s*([^\s，,、；;]+)', content)
    if sec_m:
        sec = sec_m.group(1).strip()
        return SECONDARY_TO_PRIMARY.get(sec)
    return None

# ====================== 年报解析器（V5：直接段落匹配，无正则切分） ======================

def _extract_annual_block(text: str) -> Optional[str]:
    """
    从返回的全文中，精确提取年报段落。
    财务复合指标段落的格式为：
      "统计截止日期为YYYY1231的年报，主要财务指标如下..."
    前面可能有 Q1/Q2/Q3 单季报段落，必须跳过。
    返回年报段落的文本（从"资产负债结构"到下一个段落标题之前），
    找不到则返回 None。
    """
    # 精确匹配 "统计截止日期为YYYY1231的年报"
    # 注意：同一个 Recall 里可能有多个报告（Q1、年报、Q4、Q3...）
    # 我们只取第一个出现的年报段落
    annual_header_pat = re.compile(
        r'统计截止日期为(\d{4})1231的年报'
    )
    
    m = annual_header_pat.search(text)
    if not m:
        return None
    
    year = m.group(1)
    start = m.start()
    
    # 找到年报段落的结束位置：下一个"统计截止日期为"或文本结束
    next_section = re.search(r'统计截止日期为', text[start + 1:])
    if next_section:
        end = start + 1 + next_section.start()
    else:
        end = len(text)
    
    return text[start:end]


def _extract_metric_line(block: str, keywords: list) -> Optional[str]:
    """
    在年报段落中，按关键词列表逐行搜索，返回第一个匹配行的完整文本。
    例如关键词 ["资产负债率"] 会匹配到 "资产负债率30.18%" 这一行。
    """
    if not block:
        return None
    for line in block.split('\n'):
        line = line.strip()
        if not line:
            continue
        for kw in keywords:
            if kw in line:
                return line
    return None


def _parse_pct_from_line(line: str, keyword: str = None) -> Optional[float]:
    """
    从一行文本中提取百分比数值。
    如果指定 keyword，则从关键词之后提取第一个百分比（解决一行多百分比问题）。
    例如：keyword="净利率" 时，从 "销售毛利率 45.60%，销售净利率 16.00%" 中提取 16.00
    """
    if not line:
        return None
    if keyword:
        # 找到关键词位置，只从关键词之后搜索
        idx = line.find(keyword)
        if idx >= 0:
            line = line[idx:]
    m = re.search(r'([-+]?\d+\.?\d*)%', line)
    return float(m.group(1)) if m else None


def _parse_num_from_line(line: str) -> Optional[float]:
    """从一行文本中提取带单位的数值，如 '营业总收入382.40亿元' -> 382.40e8"""
    if not line:
        return None
    m = re.search(r'([-+]?\d+\.?\d*)\s*(万[亿]?元|亿元|万元|万亿元|千元|元)', line)
    if not m:
        return None
    num, unit = float(m.group(1)), m.group(2)
    if '万亿' in unit: return num * 1e12
    if '亿' in unit: return num * 1e8
    if '万' in unit: return num * 1e4
    if '千' in unit: return num * 1e3
    return num


def _parse_yoy_from_line(line: str) -> Optional[float]:
    """
    从一行文本中提取同比增长率。
    格式如：'营业收入同比增长60.25%' 或 '归母净利润同比增长108.78%'
    """
    if not line:
        return None
    m = re.search(r'同比增长\s*([-+]?\d+\.?\d*)%', line)
    return float(m.group(1)) if m else None


def parse_financial_all(content: str) -> dict:
    """
    解析 NeoData 返回的全文，提取年报财务指标。
    策略：直接定位 "统计截止日期为YYYY1231的年报" 段落，然后逐行匹配指标。
    完全避免正则切分，从根本上杜绝段落混淆问题。
    """
    res = {k: None for k in [
        "annual_roe", "annual_gross_margin", "annual_net_margin",
        "annual_revenue_yoy", "annual_profit_yoy", "annual_debt_ratio",
        "annual_net_profit", "annual_deducted_profit", "annual_revenue",
        "annual_ocf_to_profit", "annual_ocf_abs",
        "total_asset_turnover", "ar_turnover",
        "annual_report_date"
    ]}

    # Step 1: 提取年报段落
    block = _extract_annual_block(content)
    
    if not block:
        # 兜底：尝试从业绩趋势段落提取（Recall 0 格式）
        res["annual_report_date"] = _guess_date_from_trend(content)
        return res

    # Step 2: 从年报段落提取年份
    year_m = re.search(r'统计截止日期为(\d{4})1231的年报', block)
    if year_m:
        res["annual_report_date"] = year_m.group(1) + "1231"

    # Step 3: 逐行匹配各指标
    # ROE
    line = _extract_metric_line(block, ["加权净资产收益率ROE", "净资产收益率ROE", "加权净资产收益率"])
    if line:
        res["annual_roe"] = _parse_pct_from_line(line)

    # 毛利率 — 用 keyword 参数确保从"毛利率"后提取，避免一行多百分比时取错
    line = _extract_metric_line(block, ["销售毛利率"])
    if line:
        res["annual_gross_margin"] = _parse_pct_from_line(line, keyword="毛利率")

    # 净利率 — 同上，从"净利率"后提取
    line = _extract_metric_line(block, ["销售净利率"])
    if line:
        res["annual_net_margin"] = _parse_pct_from_line(line, keyword="净利率")

    # 营收同比
    line = _extract_metric_line(block, ["营业收入同比增长", "营收同比增长", "营业总收入同比增长"])
    if line:
        res["annual_revenue_yoy"] = _parse_yoy_from_line(line)

    # 净利润同比（归母）
    line = _extract_metric_line(block, ["归母净利润同比增长"])
    if line:
        res["annual_profit_yoy"] = _parse_yoy_from_line(line)

    # 资产负债率
    line = _extract_metric_line(block, ["资产负债率"])
    if line:
        res["annual_debt_ratio"] = _parse_pct_from_line(line)

    # 营业收入（亿元）
    line = _extract_metric_line(block, ["营业总收入", "营业收入"])
    if line:
        res["annual_revenue"] = _parse_num_from_line(line)

    # 净利润（取"净利润"行，不取归母/扣非）
    # 策略：遍历年报段落每一行，找到以"净利润"开头但不含"归母"/"扣非"的行
    line = None
    for l in block.split('\n'):
        l = l.strip()
        if l.startswith('净利润') and '归母' not in l and '扣非' not in l:
            line = l
            break
    if line:
        val = _parse_num_from_line(line)
        if val is not None:
            res["annual_net_profit"] = val

    # 扣非净利润
    line = _extract_metric_line(block, ["扣非净利润"])
    if line:
        res["annual_deducted_profit"] = _parse_num_from_line(line)

    # 经营活动产生的现金流量净额（排除"每股"）
    line = _extract_metric_line(block, ["经营活动产生的现金流量净额"])
    if line:
        res["annual_ocf_abs"] = _parse_num_from_line(line)

    # OCF/净利润
    if res["annual_net_profit"] and res["annual_ocf_abs"] and res["annual_net_profit"] != 0:
        res["annual_ocf_to_profit"] = res["annual_ocf_abs"] / res["annual_net_profit"]

    # 总资产周转率
    line = _extract_metric_line(block, ["总资产周转率"])
    if line:
        m = re.search(r'([\d.]+)\s*次', line)
        if m:
            res["total_asset_turnover"] = float(m.group(1))

    # 应收账款周转率
    line = _extract_metric_line(block, ["应收账款周转率"])
    if line:
        m = re.search(r'([\d.]+)\s*次', line)
        if m:
            res["ar_turnover"] = float(m.group(1))

    return res


def _guess_date_from_trend(content: str) -> Optional[str]:
    """从业绩趋势段落推断年报日期"""
    m = re.search(r'(20\d{2})年，营业总收入', content)
    if m:
        return m.group(1) + '1231'
    return None

# ====================== 单股票获取 ======================
def fetch_stock_finance(ts_code, name, session=None):
    # 简化查询，避免因为关键词太多而搜不到结果
    query = f"{ts_code} {name} 年报"
    result = run_neodata(query, session=session)
    base = {"ts_code": ts_code, "name": name, "industry_l1_parsed": None, "fetch_success": False}
    if result.get("code") != "200":
        safe_error_append({"股票代码": ts_code, "股票名称": name, "阶段": "财务数据获取", "错误详情": f"API错误码 {result.get('code')}"})
        return base
    recalls = result.get("data", {}).get("apiData", {}).get("apiRecall", [])
    all_content = "\n".join(r.get("content", "") for r in recalls)
    if not all_content:
        safe_error_append({"股票代码": ts_code, "股票名称": name, "阶段": "财务数据获取", "错误详情": "API返回空文本"})
        return base
    base["fetch_success"] = True
    try:
        base.update(parse_financial_all(all_content))
    except Exception as e:
        safe_error_append({"股票代码": ts_code, "股票名称": name, "阶段": "财务指标解析", "错误详情": str(e)})
    ind = extract_industry_from_content(all_content)
    if ind:
        base["industry_l1_parsed"] = ind
    return base

def fetch_stock_batch(stocks, workers):
    """高并发批量获取，使用 requests.Session 连接池复用"""
    results, consecutive_errors, api_errors = [], 0, 0
    total = len(stocks)
    logger.info(f"[批次] {total}只, {workers}线程 (requests直连)")
    start_all = time.time()

    # 创建共享 Session，复用 TCP 连接
    session = req_lib.Session()
    adapter = req_lib.adapters.HTTPAdapter(
        pool_connections=workers, pool_maxsize=workers, max_retries=2
    )
    session.mount("https://", adapter)

    executor = ThreadPoolExecutor(max_workers=workers)
    future_to_stock = {
        executor.submit(fetch_stock_finance, s["ts_code"], s["name"], session): s
        for s in stocks
    }
    processed_futures = set()
    try:
        for f in as_completed(future_to_stock):
            elapsed = time.time() - start_all
            if Config.GLOBAL_TIMEOUT > 0 and elapsed > Config.GLOBAL_TIMEOUT:
                logger.warning(f"全局超时 {Config.GLOBAL_TIMEOUT}s, 取消剩余任务")
                for unf in future_to_stock:
                    if not unf.done():
                        unf.cancel()
                break
            s = future_to_stock[f]
            processed_futures.add(f)
            try:
                r = f.result()
                results.append(r)
                # 区分"API 错误"和"正常但无数据"
                if r.get("error"):
                    # API 调用本身失败（网络/超时等）
                    consecutive_errors += 1
                    api_errors += 1
                elif all(r.get(k) is None for k in ["annual_roe", "annual_gross_margin", "annual_debt_ratio"]):
                    # API 正常但无年报数据（新股/退市等），不算错误
                    consecutive_errors = 0
                else:
                    consecutive_errors = 0
            except Exception as e:
                logger.error(f"异常 {s['ts_code']}: {e}")
                safe_error_append({"股票代码": s["ts_code"], "股票名称": s["name"], "阶段": "并发获取", "错误详情": str(e)})
                results.append({"ts_code": s["ts_code"], "name": s["name"], "error": str(e)})
                consecutive_errors += 1
                api_errors += 1

            done = len(results)
            if done % 100 == 0:
                elapsed = time.time() - start_all
                rate = done / elapsed if elapsed > 0 else 0
                remaining = (total - done) / rate if rate > 0 else 0
                logger.info(f"进度: {done}/{total} ({rate:.1f}/s, 已用{elapsed:.0f}s, 剩余{remaining:.0f}s, API错误:{api_errors})")
            # 只有连续 API 错误才暂停（说明服务端可能限流）
            if consecutive_errors >= Config.PAUSE_CONSECUTIVE_EMPTY:
                logger.warning(f"暂停{Config.PAUSE_DURATION}s (连续{consecutive_errors}次API错误)")
                time.sleep(Config.PAUSE_DURATION)
                consecutive_errors = 0
    finally:
        for f, s in future_to_stock.items():
            if f in processed_futures:
                continue
            if not f.done():
                cancelled = f.cancel()
                if not cancelled:
                    try:
                        f.result(timeout=0.1)
                        results.append(f.result())
                        continue
                    except:
                        pass
                safe_error_append({"股票代码": s["ts_code"], "股票名称": s["name"], "阶段": "并发获取", "错误详情": "任务超时被取消或未完成"})
                results.append({"ts_code": s["ts_code"], "name": s["name"], "error": "超时未完成"})
            else:
                try:
                    results.append(f.result())
                except Exception as e:
                    safe_error_append({"股票代码": s["ts_code"], "股票名称": s["name"], "阶段": "并发获取", "错误详情": str(e)})
                    results.append({"ts_code": s["ts_code"], "name": s["name"], "error": str(e)})
        if sys.version_info >= (3, 9):
            executor.shutdown(wait=False, cancel_futures=True)
        else:
            executor.shutdown(wait=False)
    session.close()
    elapsed = time.time() - start_all
    rate = total / elapsed if elapsed > 0 else 0
    logger.info(f"批次完成: {len(results)}/{total}, API错误:{api_errors}, 耗时{elapsed:.0f}s ({rate:.1f}/s)")
    return results

# ====================== 行业映射与补调 ======================
def infer_industry_from_name(name):
    for kw, ind in NAME_KEYWORD_INDUSTRY.items():
        if kw in name: return ind
    return None

def infer_industry_from_code_prefix(code):
    for prefix, ind in CODE_PREFIX_INDUSTRY.items():
        if code.startswith(prefix): return ind
    return None

def fetch_industry_by_api(ts_code, name, session=None):
    """通过 API 查询单只股票的行业分类，支持传入 session 复用连接池"""
    for q in [f"{ts_code} 所属申万行业", f"{ts_code} 所属行业"]:
        try:
            res = run_neodata(q, session=session)
            if res.get("code") == "200":
                content = "\n".join(r.get("content", "") for r in res.get("data", {}).get("apiData", {}).get("apiRecall", []))
                ind = extract_industry_from_content(content)
                if ind: return ind
        except Exception as e:
            logger.debug(f"行业API查询异常 {ts_code}: {e}")
    safe_error_append({"股票代码": ts_code, "股票名称": name, "阶段": "行业API补调", "错误详情": "所有查询均失败"})
    return None

def batch_industry_patch(stocks):
    """行业补调：使用 requests.Session 连接池 + 高并发，分批执行避免 API 限流"""
    if not Config.RETRY_INDUSTRY_FOR_UNCLASSIFIED or not stocks: return {}
    mapping = {}
    total = len(stocks)
    logger.info(f"行业补调 {total} 只，使用 {Config.INDUSTRY_API_WORKERS} 线程 (Session连接池)")
    api_failed = []
    done_count = 0
    t_start = time.time()

    # 创建共享 Session，复用 TCP 连接
    session = req_lib.Session()
    adapter = req_lib.adapters.HTTPAdapter(
        pool_connections=Config.INDUSTRY_API_WORKERS,
        pool_maxsize=Config.INDUSTRY_API_WORKERS,
        max_retries=1
    )
    session.mount("https://", adapter)

    # 分批提交，每批最多 100 只，批间休眠 1s 避免触发限流
    BATCH_SIZE = 100
    total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE
    for batch_start in range(0, total, BATCH_SIZE):
        batch = stocks[batch_start: batch_start + BATCH_SIZE]
        batch_num = batch_start // BATCH_SIZE + 1
        logger.info(f"行业补调批次 {batch_num}/{total_batches} ({len(batch)} 只)")

        with ThreadPoolExecutor(max_workers=Config.INDUSTRY_API_WORKERS) as ex:
            # 传入 session 参数以复用连接池
            fut = {ex.submit(fetch_industry_by_api, s["ts_code"], s["name"], session): s for s in batch}
            for f in as_completed(fut):
                s = fut[f]
                done_count += 1
                try:
                    ind = f.result(timeout=60)
                    if ind:
                        mapping[s["ts_code"]] = ind
                    else:
                        api_failed.append(s)
                except Exception as e:
                    logger.debug(f"行业补调异常 {s['ts_code']}: {e}")
                    api_failed.append(s)

                if done_count % 50 == 0:
                    elapsed = time.time() - t_start
                    rate = done_count / elapsed if elapsed > 0 else 0
                    logger.info(f"行业补调进度: {done_count}/{total} ({rate:.1f}/s, 已补调 {len(mapping)} 只)")

        # 批间休眠，避免 API 限流（最后一批不休眠）
        if batch_start + BATCH_SIZE < total:
            time.sleep(1)

    session.close()

    # API 失败的尝试用名称推断
    if api_failed:
        name_matched = 0
        for s in api_failed:
            ind = infer_industry_from_name(s["name"])
            if ind:
                mapping[s["ts_code"]] = ind
                name_matched += 1
        logger.info(f"行业补调完成: API补调 {len(mapping) - name_matched} 只, 名称推断 {name_matched} 只, 仍失败 {len(api_failed) - name_matched} 只")

    return mapping

# ====================== 评分系统 ======================
def percentile_score(value, values, higher_better=True):
    if value is None: return None
    valid = [v for v in values if isinstance(v, (int, float))]
    if len(valid) < Config.MIN_INDUSTRY_SAMPLES: return None
    below = sum(1 for v in valid if v < value)
    equal = sum(1 for v in valid if v == value)
    pct = (below + 0.5 * equal) / len(valid) * 100
    return round(pct, 1) if higher_better else round(100 - pct, 1)

def compute_completeness(r):
    keys = ["annual_roe", "annual_gross_margin", "annual_net_margin", "annual_debt_ratio",
            "annual_revenue_yoy", "annual_profit_yoy", "annual_ocf_to_profit"]
    present = sum(1 for k in keys if r.get(k) is not None)
    ratio = present / len(keys)
    if ratio >= Config.COMPLETENESS_HIGH: return "高", present
    if ratio >= Config.COMPLETENESS_LOW: return "中", present
    return "低", present

def calc_score(stock, industry_stats, fallback):
    ind = stock.get("industry_l1", "未分类")
    use_ind = ind != "未分类" and ind in industry_stats
    ind_data = industry_stats.get(ind, {}) if use_ind else {}
    completeness, present_count = compute_completeness(stock)
    indicators = [
        ("roe", "annual_roe", "roe_list", True),
        ("gross_margin", "annual_gross_margin", "gross_margin_list", True),
        ("net_margin", "annual_net_margin", "net_margin_list", True),
        ("revenue_yoy", "annual_revenue_yoy", "revenue_yoy_list", True),
        ("profit_yoy", "annual_profit_yoy", "profit_yoy_list", True),
        ("ocf_ratio", "annual_ocf_to_profit", "ocf_ratio_list", True),
        ("debt", "annual_debt_ratio", "debt_ratio_list", False)
    ]
    scores, bases = {}, {}
    for nm, vk, lk, higher in indicators:
        val = stock.get(vk)
        if val is None:
            scores[nm] = None
            bases[nm] = "缺失"
            continue
        if nm == "roe" and isinstance(val, (int, float)) and val < 0:
            scores[nm] = 0.0
            bases[nm] = "ROE为负(0分)"
            continue
        if not use_ind:
            p = percentile_score(val, fallback[lk], higher)
            if p is not None: p *= Config.MARKET_FALLBACK_DISCOUNT
            scores[nm] = p
            bases[nm] = "全市场" if p is not None else "缺失"
        else:
            p = percentile_score(val, ind_data.get(lk, []), higher)
            if p is not None:
                scores[nm] = p
                bases[nm] = f"行业({ind})"
            else:
                p = percentile_score(val, fallback[lk], higher)
                if p is not None: p *= Config.MARKET_FALLBACK_DISCOUNT
                scores[nm] = p
                bases[nm] = "全市场(行业不足)" if p is not None else "缺失"

    profit_w = sum(w for nm, w in Config.PROFIT_SUB.items() if scores[nm] is not None)
    profit_s = sum(scores[nm] * w for nm, w in Config.PROFIT_SUB.items() if scores[nm] is not None)
    profit_score = profit_s / profit_w if profit_w > 0 else 50.0

    growth_w = sum(w for nm, w in Config.GROWTH_SUB.items() if scores[nm] is not None)
    growth_s = sum(scores[nm] * w for nm, w in Config.GROWTH_SUB.items() if scores[nm] is not None)
    growth_score = growth_s / growth_w if growth_w > 0 else 50.0

    ocf_score = scores["ocf_ratio"] if scores["ocf_ratio"] is not None else 50.0
    debt_score = scores["debt"] if scores["debt"] is not None else 50.0

    ms = Config.SCORE_WEIGHTS
    act = 0.0
    if profit_w > 0: act += ms["profit"]
    if growth_w > 0: act += ms["growth"]
    if scores["ocf_ratio"] is not None: act += ms["ocf_quality"]
    if scores["debt"] is not None: act += ms["debt_risk"]
    if act == 0:
        total = 50.0
    else:
        total = (profit_score * (ms["profit"] / act if profit_w > 0 else 0) +
                 growth_score * (ms["growth"] / act if growth_w > 0 else 0) +
                 ocf_score * (ms["ocf_quality"] / act if scores["ocf_ratio"] is not None else 0) +
                 debt_score * (ms["debt_risk"] / act if scores["debt"] is not None else 0))

    if stock.get("annual_net_profit") and stock["annual_net_profit"] < 0 and stock.get("annual_ocf_abs") and stock["annual_ocf_abs"] < 0:
        total = min(total, Config.NEGATIVE_PROFIT_PENALTY)

    if completeness == "低":
        total *= Config.LOW_COMPLETENESS_PENALTY
        if present_count <= Config.ULTRA_LOW_COMPLETENESS_THRESHOLD:
            total *= Config.ULTRA_LOW_COMPLETENESS_PENALTY

    rating = "A" if total >= 75 else ("B" if total >= 55 else ("C" if total >= 40 else ("D" if total >= 25 else "E")))
    confidence = "高" if completeness == "高" else ("中" if completeness == "中" else "低")

    return {"total_score": round(total, 2), "rating": rating,
            "detail": {"score_profit": round(profit_score, 1), "score_growth": round(growth_score, 1),
                       "score_ocf": ocf_score, "score_debt": debt_score},
            "score_base": bases, "completeness": completeness, "confidence": confidence}

def build_industry_stats(results):
    stats, fb = {}, {k: [] for k in ["roe_list", "gross_margin_list", "net_margin_list",
                                     "revenue_yoy_list", "profit_yoy_list", "ocf_ratio_list", "debt_ratio_list"]}
    for r in results:
        if r.get("error"): continue
        ind = r.get("industry_l1", "未分类")
        if ind != "未分类" and ind not in stats:
            stats[ind] = {k: [] for k in fb}
        for lk, dk in [("roe_list", "annual_roe"), ("gross_margin_list", "annual_gross_margin"),
                       ("net_margin_list", "annual_net_margin"), ("revenue_yoy_list", "annual_revenue_yoy"),
                       ("profit_yoy_list", "annual_profit_yoy"), ("ocf_ratio_list", "annual_ocf_to_profit"),
                       ("debt_ratio_list", "annual_debt_ratio")]:
            val = r.get(dk)
            if val is not None:
                fb[lk].append(val)
                if ind != "未分类":
                    stats[ind][lk].append(val)
    return stats, fb

# ====================== 数据库操作 ======================
def init_db():
    try:
        conn = sqlite3.connect(Config.DB_FILE)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute('''CREATE TABLE IF NOT EXISTS stocks (
            ts_code TEXT PRIMARY KEY,
            name TEXT,
            industry_l1 TEXT,
            industry_l2 TEXT,
            last_industry_update TIMESTAMP,
            last_full_update TIMESTAMP
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS financial_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts_code TEXT NOT NULL,
            report_date TEXT NOT NULL,
            report_type TEXT NOT NULL,
            roe REAL, gross_margin REAL, net_margin REAL,
            revenue_yoy REAL, profit_yoy REAL, debt_ratio REAL,
            net_profit REAL, deducted_profit REAL, revenue REAL,
            ocf_to_profit REAL, ocf_abs REAL,
            asset_turnover REAL, ar_turnover REAL,
            fetch_success INTEGER DEFAULT 1,
            last_update TIMESTAMP DEFAULT (datetime('now','localtime')),
            UNIQUE(ts_code, report_date, report_type)
        )''')
        conn.commit()
        return conn
    except sqlite3.OperationalError as e:
        logger.error(f"数据库初始化失败: {e}")
        sys.exit(1)

def save_reports_batch(conn, reports: list):
    fields = ['ts_code', 'report_date', 'report_type',
              'roe', 'gross_margin', 'net_margin', 'revenue_yoy', 'profit_yoy', 'debt_ratio',
              'net_profit', 'deducted_profit', 'revenue', 'ocf_to_profit', 'ocf_abs',
              'asset_turnover', 'ar_turnover', 'fetch_success', 'last_update']
    sql = f"INSERT OR REPLACE INTO financial_reports ({','.join(fields)}) VALUES ({','.join(['?']*len(fields))})"
    data = [[rep.get(f, None) for f in fields] for rep in reports]
    conn.executemany(sql, data)

def update_stocks_batch(conn, stocks_info: list):
    sql = '''INSERT OR REPLACE INTO stocks (ts_code, name, industry_l1, industry_l2, last_industry_update, last_full_update)
             VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)'''
    data = [(s['ts_code'], s.get('name', ''), s.get('industry_l1', ''), s.get('industry_l2', '')) for s in stocks_info]
    conn.executemany(sql, data)

def get_latest_report(conn, ts_code, report_types: list) -> Optional[dict]:
    placeholders = ','.join(['?'] * len(report_types))
    cur = conn.execute(f'''SELECT * FROM financial_reports 
                           WHERE ts_code=? AND report_type IN ({placeholders}) AND fetch_success=1
                           ORDER BY report_date DESC LIMIT 1''', [ts_code] + report_types)
    row = cur.fetchone()
    if row:
        cols = [desc[0] for desc in cur.description]
        return dict(zip(cols, row))
    return None

def is_report_outdated(report_date: str, report_type: str) -> bool:
    if not report_date: return True
    try:
        dt = datetime.strptime(report_date, "%Y%m%d")
    except ValueError:
        return True
    age = (datetime.now() - dt).days
    if report_type == 'annual':
        year = dt.year
        deadline = datetime(year + 1, Config.ANNUAL_DISCLOSURE_DEADLINE_MONTH, Config.ANNUAL_DISCLOSURE_DEADLINE_DAY)
        if datetime.now() > deadline:
            return True
        else:
            return age > Config.CACHE_MAX_AGE_ANNUAL
    else:
        return True

def should_refresh(conn, ts_code) -> bool:
    annual = get_latest_report(conn, ts_code, ['annual'])
    if not annual:
        return True
    return is_report_outdated(annual['report_date'], 'annual')

def merge_latest_reports(conn, ts_code) -> dict:
    annual = get_latest_report(conn, ts_code, ['annual']) or {}
    cur = conn.execute("SELECT * FROM stocks WHERE ts_code=?", (ts_code,))
    stock_info = cur.fetchone()
    if stock_info:
        cols = [desc[0] for desc in cur.description]
        stock_info = dict(zip(cols, stock_info))
    else:
        stock_info = {}
    merged = {
        "ts_code": ts_code,
        "name": stock_info.get("name", ""),
        "industry_l1": stock_info.get("industry_l1", "未分类"),
        "industry_l2": stock_info.get("industry_l2", ""),
        "annual_roe": annual.get("roe"),
        "annual_gross_margin": annual.get("gross_margin"),
        "annual_net_margin": annual.get("net_margin"),
        "annual_revenue_yoy": annual.get("revenue_yoy"),
        "annual_profit_yoy": annual.get("profit_yoy"),
        "annual_debt_ratio": annual.get("debt_ratio"),
        "annual_net_profit": annual.get("net_profit"),
        "annual_deducted_profit": annual.get("deducted_profit"),
        "annual_revenue": annual.get("revenue"),
        "annual_ocf_to_profit": annual.get("ocf_to_profit"),
        "annual_ocf_abs": annual.get("ocf_abs"),
        "total_asset_turnover": annual.get("asset_turnover"),
        "ar_turnover": annual.get("ar_turnover"),
        "fetch_success": True
    }
    return merged

# ====================== 报告生成 ======================
def generate_report(results, output_dir, error_list=None, failed_stocks=None):
    try:
        import pandas as pd
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("需要安装 pandas 和 openpyxl: pip install pandas openpyxl")
        return ""

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(output_dir, f"股票业绩评价_{ts}.xlsx")
    rows = []
    for r in results:
        d = r.get("detail", {})
        base = "; ".join(f"{k}:{v}" for k, v in r.get("score_base", {}).items())
        rows.append({
            "股票代码": r.get("ts_code"),
            "股票名称": r.get("name"),
            "一级行业": r.get("industry_l1"),
            "加权ROE(%)": r.get("annual_roe"),
            "毛利率(%)": r.get("annual_gross_margin"),
            "净利率(%)": r.get("annual_net_margin"),
            "营收同比(%)": r.get("annual_revenue_yoy"),
            "净利润同比(%)": r.get("annual_profit_yoy"),
            "经营现金流/净利润": r.get("annual_ocf_to_profit"),
            "资产负债率(%)": r.get("annual_debt_ratio"),
            "盈利评分": d.get("score_profit"),
            "成长评分": d.get("score_growth"),
            "现金流评分": d.get("score_ocf"),
            "偿债评分": d.get("score_debt"),
            "总评分": r.get("total_score"),
            "评级": r.get("rating"),
            "置信度": r.get("confidence", "高"),
            "评分基准": base,
            "数据完整度": r.get("completeness")
        })
    df = pd.DataFrame(rows)
    if df.empty:
        logger.warning("无分析结果，生成空报告。")
        df.to_excel(path, index=False)
        return path

    df["评级"] = df["评级"].astype(pd.CategoricalDtype(categories=["A", "B", "C", "D", "E"], ordered=True))
    df.sort_values(["评级", "总评分"], ascending=[True, False], inplace=True)

    try:
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='综合评价结果', index=False)
            ws = writer.sheets['综合评价结果']
            col_idx = list(df.columns).index("经营现金流/净利润") + 1
            col_letter = get_column_letter(col_idx)
            for row in range(2, len(df) + 2):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is not None:
                    cell.number_format = '0.0000'

            for rate in ["A", "B", "C", "D", "E"]:
                sub = df[df["评级"] == rate]
                if not sub.empty:
                    sub.to_excel(writer, sheet_name=f'{rate}级股票', index=False)
            low_conf = df[df["置信度"] == "低"]
            if not low_conf.empty:
                low_conf.to_excel(writer, sheet_name='低置信度股票', index=False)
            if error_list:
                pd.DataFrame(error_list).to_excel(writer, sheet_name='异常日志', index=False)
            if failed_stocks:
                pd.DataFrame(failed_stocks).to_excel(writer, sheet_name='获取失败股票', index=False)
            stats = [
                {"项目": "股票总数", "数值": len(df)},
                {"项目": "A级", "数值": int((df["评级"] == "A").sum())},
                {"项目": "B级", "数值": int((df["评级"] == "B").sum())},
                {"项目": "C级", "数值": int((df["评级"] == "C").sum())},
                {"项目": "D级", "数值": int((df["评级"] == "D").sum())},
                {"项目": "E级", "数值": int((df["评级"] == "E").sum())},
                {"项目": "平均评分", "数值": round(df["总评分"].mean(), 2)},
                {"项目": "最高评分", "数值": round(df["总评分"].max(), 2)},
                {"项目": "完整-高", "数值": int((df["数据完整度"] == "高").sum())},
                {"项目": "完整-中", "数值": int((df["数据完整度"] == "中").sum())},
                {"项目": "完整-低", "数值": int((df["数据完整度"] == "低").sum())}
            ]
            pd.DataFrame(stats).to_excel(writer, sheet_name='统计概览', index=False)
    except Exception as e:
        logger.error(f"Excel生成失败: {e}")
        csv_path = os.path.join(output_dir, f"股票业绩评价_{ts}.csv")
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        logger.info(f"已保存CSV备份: {csv_path}")
        return csv_path

    return path

# ====================== 自测模式 ======================
def run_self_test():
    """
    使用模拟返回格式的自测数据，验证 parse_financial_all() 解析正确性。
    数据格式与真实 NeoData 返回一致：包含 "统计截止日期为YYYY1231的年报" 段落标记。
    """
    # 模拟返回的财务复合指标段落（含 Q1 + 年报两个段落）
    test_content = """
统计截止日期为20240331的Q1单季报，主要财务指标如下：
营业总收入100.00亿元，同比增长 8.00%；
归母净利润同比增长 262.28%；
销售毛利率 35.20%，销售净利率 5.10%；
加权净资产收益率ROE 3.50%；
资产负债率 50.10%。

统计截止日期为20241231的年报，主要财务指标如下：
资产负债结构：资产负债率40.50%；
营业总收入500.00亿元，营业收入同比增长 12.50%；
归母净利润同比增长 18.20%；
净利润 80.00亿元；
扣非净利润 75.50亿元；
加权净资产收益率ROE 18.00%；
销售毛利率 45.60%，销售净利率 16.00%；
经营活动产生的现金流量净额 95.00亿元；
总资产周转率 0.85次；
应收账款周转率 6.20次。

统计截止日期为20241001的Q4单季报，主要财务指标如下：
资产负债率 42.00%；
营业总收入380.00亿元，同比增长 10.00%。
"""
    print("=" * 50)
    print("自测模式 — 验证年报段落提取与指标解析")
    print("=" * 50)

    # 先验证年报段落提取
    block = _extract_annual_block(test_content)
    assert block is not None, "自测失败：未找到年报段落"
    assert "20241231" in block, "年报段落应包含 20241231"
    assert "262.28%" not in block, "年报段落不应包含 Q1 的 262.28%"
    assert "销售毛利率 45.60%" in block, "年报段落应包含正确的毛利率"
    print("[OK] 年报段落提取正确（Q1数据已排除）")

    # 再验证完整解析
    data = parse_financial_all(test_content)
    print("\n自测解析结果：")
    for k, v in data.items():
        print(f"  {k}: {v}")

    # 核心断言：确保取到年报数据而非 Q1
    assert data.get("annual_report_date") == "20241231", f"报告日期应为 20241231, 实际 {data.get('annual_report_date')}"
    assert data.get("annual_roe") == 18.0, f"ROE 应为 18.0, 实际 {data.get('annual_roe')}"
    assert data.get("annual_gross_margin") == 45.6, f"毛利率应为 45.6, 实际 {data.get('annual_gross_margin')}"
    assert data.get("annual_net_margin") == 16.0, f"净利率应为 16.0, 实际 {data.get('annual_net_margin')}"
    assert data.get("annual_revenue_yoy") == 12.5, f"营收同比应为 12.5, 实际 {data.get('annual_revenue_yoy')}"
    assert data.get("annual_profit_yoy") == 18.2, f"净利润同比应为 18.2, 实际 {data.get('annual_profit_yoy')}"
    assert data.get("annual_debt_ratio") == 40.5, f"资产负债率应为 40.5, 实际 {data.get('annual_debt_ratio')}"
    assert data.get("annual_net_profit") == 80e8, f"净利润应为 80e8, 实际 {data.get('annual_net_profit')}"
    assert data.get("annual_deducted_profit") == 75.5e8, f"扣非净利润应为 75.5e8, 实际 {data.get('annual_deducted_profit')}"
    assert data.get("annual_ocf_abs") == 95e8, f"经营现金流应为 95e8, 实际 {data.get('annual_ocf_abs')}"
    assert abs(data.get("annual_ocf_to_profit") - 95/80) < 0.01, f"OCF/净利润应为 {95/80}, 实际 {data.get('annual_ocf_to_profit')}"
    assert data.get("total_asset_turnover") == 0.85, f"总资产周转率应为 0.85, 实际 {data.get('total_asset_turnover')}"
    assert data.get("ar_turnover") == 6.2, f"应收账款周转率应为 6.2, 实际 {data.get('ar_turnover')}"

    print("\n✅ 自测通过！所有指标解析正确，Q1/年报段落无混淆。")
    print("=" * 50)

def load_stock_list(path):
    stocks = []
    if not os.path.exists(path):
        logger.warning(f"股票列表文件不存在: {path}")
        return stocks
    try:
        with open(path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                parts = line.split(None, 1)
                symbol = parts[0].zfill(6)
                name = parts[1] if len(parts) > 1 else ""
                if symbol.startswith(('688', '430', '83', '87')):
                    continue
                suffix = '.SZ' if symbol.startswith(('0', '3')) else '.SH'
                stocks.append({"ts_code": symbol + suffix, "symbol": symbol, "name": name})
    except UnicodeDecodeError as e:
        logger.error(f"股票文件编码错误（需要UTF-8）: {e}")
    except Exception as e:
        logger.error(f"读取股票文件失败: {e}")
    return stocks

# ====================== 主流程 ======================
def main():
    global error_log, logger
    error_log.clear()

    parser = argparse.ArgumentParser(description="业绩分析系统 V5.0.0",
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--base-dir", default=Config.BASE_DIR, help="工作目录（默认脚本所在目录）")
    parser.add_argument("--stock-file", default="xuan.txt", help="股票列表文件名")
    parser.add_argument("--workers", type=int, default=Config.FINANCE_WORKERS, help="并发线程数")
    parser.add_argument("--force-refresh", action="store_true", help="忽略缓存全量更新")
    parser.add_argument("--no-industry-patch", action="store_true", help="禁用行业API补调")
    parser.add_argument("--timeout", type=int, default=Config.GLOBAL_TIMEOUT, help="全局超时秒数（0=不限）")
    parser.add_argument("--test", action="store_true", help="运行内置自测并退出")
    args = parser.parse_args()

    if args.test:
        print("启动自测模式...")
        run_self_test()
        return

    Config.BASE_DIR = os.path.abspath(args.base_dir)
    os.makedirs(Config.BASE_DIR, exist_ok=True)
    if not os.path.isabs(args.stock_file):
        Config.DEFAULT_STOCK_FILE = os.path.join(Config.BASE_DIR, args.stock_file)
    else:
        Config.DEFAULT_STOCK_FILE = args.stock_file
    Config.OUTPUT_DIR = Config.BASE_DIR
    Config.INDUSTRY_MAP_FILE = os.path.join(Config.BASE_DIR, "industry_map.json")
    Config.DB_FILE = os.path.join(Config.BASE_DIR, "stock_cache.db")
    Config.LOG_FILE = os.path.join(Config.BASE_DIR, "stock_analyzer.log")
    Config.FINANCE_WORKERS = args.workers
    Config.GLOBAL_TIMEOUT = args.timeout
    Config.FORCE_REFRESH = args.force_refresh
    if args.no_industry_patch:
        Config.RETRY_INDUSTRY_FOR_UNCLASSIFIED = False

    logger = setup_logging()
    logger.info("=" * 60)
    logger.info("业绩分析系统 V5.0.0 启动")
    logger.info(f"工作目录: {Config.BASE_DIR}")
    logger.info("=" * 60)

    try:
        import pandas, openpyxl
    except ImportError:
        logger.error("请先安装依赖: pip install pandas openpyxl")
        sys.exit(1)

    stocks = load_stock_list(Config.DEFAULT_STOCK_FILE)
    if not stocks:
        logger.error("股票列表为空")
        return

    conn = init_db()
    industry_map_full = get_industry_map()

    if Config.FORCE_REFRESH:
        need_fetch = stocks
        cached_data = []
    else:
        need_fetch = []
        cached_data = []
        for s in stocks:
            if should_refresh(conn, s['ts_code']):
                need_fetch.append(s)
            else:
                cached_data.append(merge_latest_reports(conn, s['ts_code']))
        logger.info(f"缓存命中 {len(cached_data)} 只，需获取 {len(need_fetch)} 只")

    fresh_data = fetch_stock_batch(need_fetch, Config.FINANCE_WORKERS) if need_fetch else []

    reports_to_save = []
    stocks_to_update = []
    for r in fresh_data:
        ts = r['ts_code']
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if r.get('annual_report_date'):
            rep = {
                'ts_code': ts,
                'report_date': r['annual_report_date'],
                'report_type': 'annual',
                'roe': r.get('annual_roe'),
                'gross_margin': r.get('annual_gross_margin'),
                'net_margin': r.get('annual_net_margin'),
                'revenue_yoy': r.get('annual_revenue_yoy'),
                'profit_yoy': r.get('annual_profit_yoy'),
                'debt_ratio': r.get('annual_debt_ratio'),
                'net_profit': r.get('annual_net_profit'),
                'deducted_profit': r.get('annual_deducted_profit'),
                'revenue': r.get('annual_revenue'),
                'ocf_to_profit': r.get('annual_ocf_to_profit'),
                'ocf_abs': r.get('annual_ocf_abs'),
                'asset_turnover': r.get('total_asset_turnover'),
                'ar_turnover': r.get('ar_turnover'),
                'fetch_success': 1 if r.get('fetch_success') else 0,
                'last_update': now_str
            }
            reports_to_save.append(rep)

        ind = r.pop("industry_l1_parsed", None)
        if ind:
            r["industry_l1"] = ind
        else:
            code = ts.split('.')[0]
            ind = industry_map_full.get(code)
            if ind:
                r["industry_l1"] = ind
            else:
                r["industry_l1"] = "未分类"
        stocks_to_update.append({
            'ts_code': ts,
            'name': r.get('name', ''),
            'industry_l1': r["industry_l1"],
            'industry_l2': r.get('industry_l2', '')
        })

    if reports_to_save:
        save_reports_batch(conn, reports_to_save)
    if stocks_to_update:
        update_stocks_batch(conn, stocks_to_update)
    conn.commit()

    all_results = cached_data + fresh_data

    unclassified = []
    for r in all_results:
        if r.get("industry_l1") in [None, "未分类"]:
            code = r["ts_code"].split('.')[0]
            ind = industry_map_full.get(code)
            if ind:
                r["industry_l1"] = ind
            else:
                sec = r.get("industry_l2", "")
                if sec in SECONDARY_TO_PRIMARY:
                    r["industry_l1"] = SECONDARY_TO_PRIMARY[sec]
                else:
                    ind = infer_industry_from_name(r["name"])
                    if ind:
                        r["industry_l1"] = ind
                    else:
                        unclassified.append(r)

    patch = batch_industry_patch(unclassified) if Config.RETRY_INDUSTRY_FOR_UNCLASSIFIED else {}
    for r in unclassified:
        if r["ts_code"] in patch:
            r["industry_l1"] = patch[r["ts_code"]]

    for r in all_results:
        if r.get("industry_l1") == "未分类":
            ind = infer_industry_from_code_prefix(r["ts_code"].split('.')[0])
            if ind:
                r["industry_l1"] = ind

    final_unclass = sum(1 for r in all_results if r.get("industry_l1") == "未分类")
    logger.info(f"最终未分类: {final_unclass} 只")

    ind_stats, fallback = build_industry_stats(all_results)
    scored = 0
    for r in all_results:
        if r.get("error"):
            continue
        sr = calc_score(r, ind_stats, fallback)
        r.update(sr)
        scored += 1
    logger.info(f"评分完成: {scored} 只")

    failed_list = []
    for r in all_results:
        if r.get("error") or (r.get("fetch_success") is False):
            failed_list.append({
                "股票代码": r.get("ts_code"),
                "股票名称": r.get("name"),
                "失败原因": r.get("error", "数据缺失")
            })

    report_path = generate_report(all_results, Config.OUTPUT_DIR, error_log, failed_list)
    logger.info(f"报告: {report_path}")

    jp = os.path.join(Config.OUTPUT_DIR, f"股票分析数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    sorted_r = sorted([r for r in all_results if "total_score" in r], key=lambda x: x["total_score"], reverse=True)
    with open(jp, 'w', encoding='utf-8') as f:
        json.dump(sorted_r, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f"JSON: {jp}")

    if error_log:
        logger.info(f"异常记录 {len(error_log)} 条已写入Excel")
    conn.close()
    logger.info("任务完成。")

if __name__ == "__main__":
    main()