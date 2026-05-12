#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A股智能选股系统 V5.0.0
基于 NeoData 真实 API 数据，对 A 股年报进行结构化段落匹配与业绩评分。
"""

import os, sys, json, time, sqlite3, logging, argparse, re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional
import requests as req_lib
from requests.adapters import HTTPAdapter

# ============================================================
# 配置
# ============================================================
class Config:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DEFAULT_STOCK_FILE = os.path.join(BASE_DIR, "xuan.txt")
    OUTPUT_DIR = BASE_DIR
    INDUSTRY_MAP_FILE = os.path.join(BASE_DIR, "industry_map.json")
    DB_FILE = os.path.join(BASE_DIR, "stock_cache.db")
    FINANCE_WORKERS = 16
    API_RETRY_TIMES = 2
    API_RETRY_BACKOFF_BASE = 3.0
    API_TIMEOUT = 50
    PAUSE_CONSECUTIVE_EMPTY = 10
    PAUSE_DURATION = 20
    GLOBAL_TIMEOUT = 7200
    MIN_INDUSTRY_SAMPLES = 5
    CACHE_MAX_AGE_ANNUAL = 400
    NEGATIVE_PROFIT_PENALTY = 15.0
    MARKET_FALLBACK_DISCOUNT = 0.95
    LOW_COMPLETENESS_PENALTY = 0.9
    ULTRA_LOW_COMPLETENESS_PENALTY = 0.75
    INDUSTRY_API_WORKERS = 15
    INDUSTRY_CACHE_DAYS = 365
    ANNUAL_DISCLOSURE_DEADLINE_MONTH = 4
    ANNUAL_DISCLOSURE_DEADLINE_DAY = 30
    NEODATA_URL = "https://copilot.tencent.com/agenttool/v1/neodata"
    NEODATA_TOKEN_FILE = os.path.expanduser("~/.workbuddy/.neodata_token")

# ============================================================
# 日志
# ============================================================
def setup_logging():
    log_file = os.path.join(Config.BASE_DIR, "stock_analyzer.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8", mode="a"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ============================================================
# 工具函数
# ============================================================
def load_token() -> str:
    if not os.path.exists(Config.NEODATA_TOKEN_FILE):
        raise FileNotFoundError(f"NeoData Token 文件不存在: {Config.NEODATA_TOKEN_FILE}")
    with open(Config.NEODATA_TOKEN_FILE, "r", encoding="utf-8") as f:
        token = f.read().strip()
    if not token:
        raise ValueError("NeoData Token 文件为空")
    return token

def parse_num(text: str) -> Optional[float]:
    if not text:
        return None
    text = text.strip()
    m = re.search(r"([-+]?\d+\.?\d*)%", text)
    if m:
        return float(m.group(1))
    m = re.search(r"([-+]?\d+\.?\d*)", text)
    if m:
        return float(m.group(1))
    return None

def _parse_num_from_line(line: str) -> Optional[float]:
    """从财报行提取带单位数值，单位匹配顺序：万亿元>亿元>万元>千元>元"""
    m = re.search(r'([-+]?\d+\.?\d*)\s*(万亿元|亿元|万元|千元|元)', line)
    if not m:
        return None
    val = float(m.group(1))
    unit = m.group(2)
    if unit == "万亿元": return val * 1e12
    elif unit == "亿元": return val * 1e8
    elif unit == "万元": return val * 1e4
    elif unit == "千元": return val * 1e3
    else: return val

def year_of_date(date_str: str) -> int:
    return int(str(date_str)[:4])

# ============================================================
# 自测数据
# ============================================================
MOCK_NEODATA_RESPONSE = """
统计截止日期为20230331的季报
营业收入4182651234.56元
归母净利润同比增长12.34%

统计截止日期为20241231的年报
加权净资产收益率ROE15.67%
销售毛利率42.35%
销售净利率18.22%
营业收入同比增长28.45%
归母净利润同比增长35.67%
资产负债率38.92%
营业总收入18654321098.76元
净利润1642130865.33元
扣非净利润1523456789.01元
经营活动产生的现金流量净额2156789012.34元
净利润现金含量160.44%
净利润增长率25.80%
总资产周转率0.85次
应收账款周转率6.78次
经营现金流/净利润145.67%

统计截止日期为20240930的季报
营业收入1234567890.12元
归母净利润同比增长15.67%
"""

# ============================================================
# 年报解析（V5 核心：直接段落匹配）
# ============================================================
def _extract_annual_block(text: str, year: Optional[int] = None) -> str:
    """精确提取年报段落，锚点：统计截止日期为YYYY1231的年报"""
    if year:
        target = f"统计截止日期为{year}1231的年报"
    else:
        matches = list(re.finditer(r"统计截止日期为(\d{4})1231的年报", text))
        if not matches:
            return ""
        last = matches[-1]
        target = last.group(0)
    start = text.find(target)
    if start == -1:
        return ""
    start += len(target)
    next_anchor = text.find("统计截止日期为", start + 1)
    if next_anchor == -1:
        block = text[start:]
    else:
        block = text[start:next_anchor]
    return block.strip()

def _guess_date_from_trend(text: str) -> Dict:
    return {}

def _extract_metric_line(block: str, keywords: List[str]) -> Optional[str]:
    """在年报段落内按关键词逐行搜索"""
    for line in block.split("\n"):
        line = line.strip()
        if not line:
            continue
        for kw in keywords:
            if kw in line:
                return line
    return None

def parse_financial_all(block: str) -> Dict:
    """从年报段落提取 13 个财务指标"""
    result = {}
    # 盈利能力
    line = _extract_metric_line(block, ["加权净资产收益率ROE", "净资产收益率ROE", "加权净资产收益率"])
    result["roe"] = parse_num(line) if line else None
    line = _extract_metric_line(block, ["销售毛利率"])
    result["gross_margin"] = parse_num(line[line.find("毛利率"):]) if line and "毛利率" in line else None
    line = _extract_metric_line(block, ["销售净利率"])
    result["net_margin"] = parse_num(line[line.find("净利率"):]) if line and "净利率" in line else None
    # 成长性
    line = _extract_metric_line(block, ["营业收入同比增长", "营收同比增长", "营业总收入同比增长"])
    result["revenue_yoy"] = parse_num(line) if line else None
    line = _extract_metric_line(block, ["归母净利润同比增长"])
    result["profit_yoy"] = parse_num(line) if line else None
    # 偿债风险
    line = _extract_metric_line(block, ["资产负债率"])
    result["debt_ratio"] = parse_num(line) if line else None
    # 规模与现金流
    line = _extract_metric_line(block, ["营业总收入", "营业收入"])
    result["revenue"] = _parse_num_from_line(line) if line else None
    # 净利润（特殊逻辑：不含归母/扣非/现金含量/增长率/同比）
    result["net_profit"] = None
    for l in block.split("\n"):
        l = l.strip()
        if l.startswith("净利润") and all(x not in l for x in ["归母","扣非","现金含量","增长率","同比"]):
            result["net_profit"] = _parse_num_from_line(l)
            break
    line = _extract_metric_line(block, ["扣非净利润"])
    result["deducted_profit"] = _parse_num_from_line(line) if line else None
    line = _extract_metric_line(block, ["经营活动产生的现金流量净额"])
    result["ocf_abs"] = _parse_num_from_line(line) if line else None
    # 经营现金流/净利润（计算值）
    if result.get("net_profit") and result.get("ocf_abs") and result["net_profit"] != 0:
        result["ocf_to_profit"] = round(result["ocf_abs"] / result["net_profit"] * 100, 2)
    else:
        result["ocf_to_profit"] = None
    # 运营效率
    line = _extract_metric_line(block, ["总资产周转率"])
    result["asset_turnover"] = parse_num(line) if line else None
    line = _extract_metric_line(block, ["应收账款周转率"])
    result["ar_turnover"] = parse_num(line) if line else None
    return result

# ============================================================
# 行业归属确定
# ============================================================
_FALLBACK_INDUSTRY_MAP: Dict[str, str] = {}
SECONDARY_TO_PRIMARY: Dict[str, str] = {}
NAME_KEYWORD_INDUSTRY: Dict[str, str] = {
    "银行":"银行","证券":"非银金融","保险":"非银金融","地产":"房地产","房地产":"房地产",
    "钢铁":"钢铁","煤炭":"煤炭","有色":"有色金属","化工":"基础化工","医药":"医药生物",
    "生物":"医药生物","电子":"电子","计算机":"计算机","通信":"通信","汽车":"汽车",
    "机械":"机械设备","电力":"公用事业","食品":"食品饮料","饮料":"食品饮料","家电":"家用电器",
    "纺织":"纺织服饰","建筑":"建筑装饰","军工":"国防军工","传媒":"传媒","光伏":"电力设备",
    "电池":"电力设备","半导体":"电子","芯片":"电子",
}
CODE_PREFIX_INDUSTRY: Dict[str, str] = {"60":"银行","00":"房地产","30":"医药生物","68":"电子"}

def load_industry_map() -> Dict:
    if os.path.exists(Config.INDUSTRY_MAP_FILE):
        with open(Config.INDUSTRY_MAP_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def extract_industry_from_content(text: str) -> Optional[str]:
    for p in [r"所属一级行业[：:]\s*(\S+)", r"申万一级行业[：:]\s*(\S+)", r"行业分类[：:]\s*(\S+)"]:
        m = re.search(p, text)
        if m:
            return m.group(1).strip()
    return None

def infer_industry_from_name(name: str) -> Optional[str]:
    for kw, industry in NAME_KEYWORD_INDUSTRY.items():
        if kw in name:
            return industry
    return None

def determine_industry(ts_code, name, content, industry_map, use_api=True, session=None) -> Optional[str]:
    """多级策略确定申万一级行业：文本解析→本地映射→二级转一级→名称规则→API补调→代码前缀"""
    ind = extract_industry_from_content(content)
    if ind: return ind
    code_short = ts_code.split(".")[0]
    if ts_code in industry_map: return industry_map[ts_code].get("industry_l1")
    if code_short in industry_map: return industry_map[code_short].get("industry_l1")
    ind = infer_industry_from_name(name)
    if ind: return ind
    if use_api:
        ind = fetch_industry_by_api(ts_code, name, session)
        if ind: return ind
    prefix = code_short[:2]
    return CODE_PREFIX_INDUSTRY.get(prefix)

def fetch_industry_by_api(ts_code, name, session=None) -> Optional[str]:
    token = load_token()
    query = f"{ts_code} {name} 所属行业"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        if session:
            resp = session.post(Config.NEODATA_URL, json={"query": query}, headers=headers, timeout=30)
        else:
            resp = req_lib.post(Config.NEODATA_URL, json={"query": query}, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        text = data.get("data", {}).get("text", "") if isinstance(data.get("data"), dict) else str(data)
        return extract_industry_from_content(text)
    except Exception as e:
        logger.debug(f"行业API查询失败 {ts_code}: {e}")
        return None

def batch_industry_patch(stocks: List[Dict], industry_map: Dict, workers=Config.INDUSTRY_API_WORKERS):
    """批量 API 补调行业，每批 100 只，批间休眠 1s，15线程+Session连接池"""
    BATCH_SIZE = 100
    need_patch = [s for s in stocks if not s.get("industry_l1")]
    total = len(need_patch)
    if total == 0: return
    logger.info(f"行业补调: 共 {total} 只股票需要补调")
    adapter = HTTPAdapter(pool_connections=workers, pool_maxsize=workers, max_retries=1)
    session = req_lib.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    api_errors = 0
    results = []
    for batch_start in range(0, total, BATCH_SIZE):
        batch = need_patch[batch_start:batch_start+BATCH_SIZE]
        batch_num = batch_start // BATCH_SIZE + 1
        total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(determine_industry, s["ts_code"], s["name"], s.get("content",""), industry_map, True, session): s for s in batch}
            for future in as_completed(futures):
                s = futures[future]
                try:
                    ind = future.result(timeout=60)
                    if ind:
                        s["industry_l1"] = ind
                        results.append((s["ts_code"], ind))
                except Exception as e:
                    api_errors += 1
                    logger.debug(f"行业补调失败 {s['ts_code']}: {e}")
        done = min(batch_start + BATCH_SIZE, total)
        logger.info(f"行业补调批次 {batch_num}/{total_batches}: {done}/{total}, 成功{len(results)}, API错误{api_errors}")
        if batch_start + BATCH_SIZE < total:
            time.sleep(1)
    session.close()
    logger.info(f"行业补调完成: 成功 {len(results)}/{total}")

# ============================================================
# 数据库（SQLite 缓存）
# ============================================================
def init_db(db_path=Config.DB_FILE):
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=10000")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS stocks (
            ts_code TEXT PRIMARY KEY, name TEXT, industry_l1 TEXT, industry_l2 TEXT,
            last_industry_update TEXT, last_full_update TEXT
        );
        CREATE TABLE IF NOT EXISTS financial_reports (
            ts_code TEXT, report_date TEXT, report_type TEXT DEFAULT 'annual',
            roe REAL, gross_margin REAL, net_margin REAL, revenue_yoy REAL, profit_yoy REAL,
            debt_ratio REAL, net_profit REAL, deducted_profit REAL, revenue REAL,
            ocf_to_profit REAL, ocf_abs REAL, asset_turnover REAL, ar_turnover REAL,
            fetch_success INTEGER DEFAULT 0, last_update TEXT,
            PRIMARY KEY (ts_code, report_date, report_type)
        );
        CREATE INDEX IF NOT EXISTS idx_reports_ts_type ON financial_reports(ts_code, report_type);
    """)
    conn.commit()
    return conn

def should_refresh(conn, ts_code, year) -> bool:
    cur = conn.execute("SELECT report_date, last_update FROM financial_reports WHERE ts_code=? AND report_type='annual' AND fetch_success=1 ORDER BY report_date DESC LIMIT 1", (ts_code,))
    row = cur.fetchone()
    if not row: return True
    report_year = year_of_date(str(row[0]))
    deadline = datetime(report_year + 1, Config.ANNUAL_DISCLOSURE_DEADLINE_MONTH, Config.ANNUAL_DISCLOSURE_DEADLINE_DAY)
    if datetime.now() > deadline and report_year < year: return True
    if row[1]:
        age_days = (datetime.now() - datetime.fromisoformat(row[1])).days
        if age_days > Config.CACHE_MAX_AGE_ANNUAL: return True
    return report_year < year

def save_report(conn, ts_code, report_date, metrics, success):
    now_str = datetime.now().isoformat()
    conn.execute(
        "INSERT OR REPLACE INTO financial_reports (ts_code,report_date,report_type,roe,gross_margin,net_margin,revenue_yoy,profit_yoy,debt_ratio,net_profit,deducted_profit,revenue,ocf_to_profit,ocf_abs,asset_turnover,ar_turnover,fetch_success,last_update) VALUES (?,?,'annual',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (ts_code, report_date, metrics.get("roe"), metrics.get("gross_margin"), metrics.get("net_margin"),
         metrics.get("revenue_yoy"), metrics.get("profit_yoy"), metrics.get("debt_ratio"),
         metrics.get("net_profit"), metrics.get("deducted_profit"), metrics.get("revenue"),
         metrics.get("ocf_to_profit"), metrics.get("ocf_abs"), metrics.get("asset_turnover"),
         metrics.get("ar_turnover"), 1 if success else 0, now_str))
    conn.commit()

def save_stock_industry(conn, ts_code, name, industry_l1, industry_l2=""):
    now_str = datetime.now().isoformat()
    conn.execute("INSERT OR REPLACE INTO stocks (ts_code,name,industry_l1,industry_l2,last_industry_update,last_full_update) VALUES (?,?,?,?,?,?)",
                 (ts_code, name, industry_l1, industry_l2, now_str, now_str))
    conn.commit()

def merge_latest_reports(conn, stocks):
    """从数据库取出最新 annual 记录，与 stocks 表合并"""
    result = []
    for s in stocks:
        ts_code = s["ts_code"]
        cur = conn.execute("SELECT * FROM financial_reports WHERE ts_code=? AND report_type='annual' AND fetch_success=1 ORDER BY report_date DESC LIMIT 1", (ts_code,))
        row = cur.fetchone()
        cur2 = conn.execute("SELECT industry_l1, industry_l2 FROM stocks WHERE ts_code=?", (ts_code,))
        stock_row = cur2.fetchone()
        item = {
            "ts_code": ts_code, "name": s.get("name", ""),
            "industry_l1": stock_row[0] if stock_row else s.get("industry_l1", ""),
            "industry_l2": stock_row[1] if stock_row else "",
            "fetch_success": False,
        }
        if row:
            cols = [d[0] for d in cur.description]
            report = dict(zip(cols, row))
            item.update({
                "roe": report.get("roe"), "gross_margin": report.get("gross_margin"),
                "net_margin": report.get("net_margin"), "revenue_yoy": report.get("revenue_yoy"),
                "profit_yoy": report.get("profit_yoy"), "debt_ratio": report.get("debt_ratio"),
                "net_profit": report.get("net_profit"), "deducted_profit": report.get("deducted_profit"),
                "revenue": report.get("revenue"), "ocf_to_profit": report.get("ocf_to_profit"),
                "ocf_abs": report.get("ocf_abs"), "asset_turnover": report.get("asset_turnover"),
                "ar_turnover": report.get("ar_turnover"),
                "fetch_success": bool(report.get("fetch_success")),
                "annual_report_date": report.get("report_date"),
                "report_date": report.get("report_date"),
            })
        else:
            for f in ["roe","gross_margin","net_margin","revenue_yoy","profit_yoy","debt_ratio",
                       "net_profit","deducted_profit","revenue","ocf_to_profit","ocf_abs",
                       "asset_turnover","ar_turnover"]:
                item[f] = None
            item["annual_report_date"] = None
            item["report_date"] = None
        result.append(item)
    return result

# ============================================================
# NeoData API 调用
# ============================================================
def run_neodata(query, token, timeout=Config.API_TIMEOUT) -> str:
    """调用 NeoData API，带重试和指数退避。API_RETRY_TIMES=2 时总尝试 3 次。
    
    API 返回结构:
    {
      "code": "200",
      "data": {
        "apiData": {
          "apiRecall": [
            {"type": "财务主要复合指标", "content": "...统计截止日期为20251231的年报..."},
            {"type": "主营构成与业绩趋势", "content": "..."}
          ]
        }
      }
    }
    优先返回包含"统计截止日期为YYYY1231的年报"的 content，否则拼接所有 content。
    """
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    payload = {"query": query}
    for attempt in range(1, Config.API_RETRY_TIMES + 2):
        try:
            resp = req_lib.post(Config.NEODATA_URL, json=payload, headers=headers, timeout=timeout)
            resp.raise_for_status()
            data = resp.json()
            inner = data.get("data", {})
            if isinstance(inner, dict):
                # 优先从 apiRecall 中提取年报内容
                api_data = inner.get("apiData", {})
                recall_list = api_data.get("apiRecall", [])
                if isinstance(recall_list, list) and recall_list:
                    # 优先找包含"统计截止日期为...年报"的段落
                    for item in recall_list:
                        content = item.get("content", "")
                        if content and "统计截止日期为" in content and "年报" in content:
                            return content
                    # 其次找 type 包含"财务"的段落
                    for item in recall_list:
                        content = item.get("content", "")
                        if content and "财务" in item.get("type", ""):
                            return content
                    # 最后拼接所有 content
                    parts = [item.get("content", "") for item in recall_list if item.get("content")]
                    if parts:
                        return "\n\n".join(parts)
                # 兼容旧格式: data.text
                if isinstance(inner.get("text"), str) and inner["text"]:
                    return inner["text"]
            elif isinstance(inner, str):
                return inner
            return json.dumps(data, ensure_ascii=False)
        except req_lib.exceptions.Timeout:
            logger.warning(f"API 超时 (attempt {attempt}/{Config.API_RETRY_TIMES+1}): {query[:50]}")
        except req_lib.exceptions.HTTPError as e:
            code = e.response.status_code if e.response else "?"
            logger.warning(f"API HTTP {code} (attempt {attempt}): {query[:50]}")
        except Exception as e:
            logger.warning(f"API 错误 (attempt {attempt}): {e}")
        if attempt < Config.API_RETRY_TIMES + 1:
            wait_time = Config.API_RETRY_BACKOFF_BASE ** attempt
            logger.info(f"等待 {wait_time:.0f}s 后重试...")
            time.sleep(wait_time)
    return ""

def fetch_stock_finance(ts_code, name, token, session, year=None) -> Dict:
    """获取单只股票财务数据"""
    query = f"{ts_code} {name} 年报"
    try:
        text = run_neodata(query, token)
        if not text:
            return {"metrics": {}, "content": "", "report_date": "", "fetch_success": False}
        block = _extract_annual_block(text, year) if year else _extract_annual_block(text)
        if not block:
            return {"metrics": {}, "content": text, "report_date": "", "fetch_success": False}
        metrics = parse_financial_all(block)
        m = re.search(r"统计截止日期为(\d{4})1231的年报", text)
        report_date = m.group(1) + "1231" if m else ""
        return {"metrics": metrics, "content": text, "report_date": report_date, "fetch_success": True}
    except Exception as e:
        logger.error(f"获取财务数据异常 {ts_code} {name}: {e}")
        return {"metrics": {}, "content": "", "report_date": "", "fetch_success": False}

def fetch_stock_batch(stocks, token, workers=Config.FINANCE_WORKERS, force_refresh=False, conn=None) -> List[Dict]:
    """批量获取财务数据，Session 连接池 + 多线程"""
    total = len(stocks)
    results = []
    api_errors = 0
    consecutive_errors = 0
    start_time = time.time()
    adapter = HTTPAdapter(pool_connections=workers, pool_maxsize=workers, max_retries=2)
    session = req_lib.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    logger.info(f"开始获取财务数据: 共 {total} 只, {workers} 线程")
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {}
        for s in stocks:
            if not force_refresh and conn and not should_refresh(conn, s["ts_code"], datetime.now().year - 1):
                continue
            future = executor.submit(fetch_stock_finance, s["ts_code"], s["name"], token, session)
            futures[future] = s
        done_count = 0
        for future in as_completed(futures):
            s = futures[future]
            done_count += 1
            try:
                result = future.result(timeout=Config.API_TIMEOUT + 10)
                results.append({**s, **result})
                if not result.get("fetch_success"):
                    if not result.get("content"):
                        api_errors += 1
                        consecutive_errors += 1
                    else:
                        consecutive_errors = 0
                else:
                    consecutive_errors = 0
                if consecutive_errors >= Config.PAUSE_CONSECUTIVE_EMPTY:
                    logger.warning(f"连续 {consecutive_errors} 次 API 错误，暂停 {Config.PAUSE_DURATION}s")
                    time.sleep(Config.PAUSE_DURATION)
                    consecutive_errors = 0
            except Exception as e:
                api_errors += 1
                consecutive_errors += 1
                results.append({**s, "metrics": {}, "content": "", "report_date": "", "fetch_success": False})
                logger.error(f"任务异常 {s['ts_code']}: {e}")
            if done_count % 100 == 0 and done_count > 0:
                elapsed = time.time() - start_time
                rate = done_count / elapsed if elapsed > 0 else 0
                eta = (total - done_count) / rate if rate > 0 else 0
                logger.info(f"进度: {done_count}/{total} ({done_count/total*100:.1f}%), 速率:{rate:.1f}/s, 已用:{elapsed:.0f}s, 剩余:{eta:.0f}s, API错误:{api_errors}")
    session.close()
    elapsed = time.time() - start_time
    logger.info(f"获取完成: {len(results)}/{total}, API错误:{api_errors}, 耗时:{elapsed:.0f}s")
    return results

# ============================================================
# 评分计算
# ============================================================
CORE_METRICS = ["roe", "gross_margin", "net_margin", "revenue_yoy", "profit_yoy", "ocf_to_profit", "debt_ratio"]

def calc_completeness(metrics):
    """计算数据完整度，返回 (比例, 等级)"""
    non_null = sum(1 for m in CORE_METRICS if metrics.get(m) is not None)
    ratio = non_null / len(CORE_METRICS)
    if ratio >= 0.7: return ratio, "high"
    elif ratio >= 0.4: return ratio, "medium"
    elif non_null <= 1: return ratio, "ultra_low"
    else: return ratio, "low"

def percentile_rank(value, values, reverse=False) -> float:
    """计算百分位排名 0~100"""
    if not values: return 50.0
    sorted_vals = sorted(values, reverse=reverse)
    n = len(sorted_vals)
    for i, v in enumerate(sorted_vals):
        if value >= v:
            return (i / (n - 1)) * 100 if n > 1 else 50.0
    return 0.0

def calc_score(stock, industry_stocks, all_stocks) -> Dict:
    """计算单只股票四维评分和综合评级"""
    ts_code = stock["ts_code"]
    industry = stock.get("industry_l1", "")
    metrics = stock
    pool = industry_stocks.get(industry, [])
    use_market_fallback = len(pool) < Config.MIN_INDUSTRY_SAMPLES
    if use_market_fallback:
        pool = all_stocks
    discount = Config.MARKET_FALLBACK_DISCOUNT if use_market_fallback else 1.0

    def pool_values(key):
        return [s[key] for s in pool if s.get(key) is not None]

    # 盈利能力 (35%): ROE(40%) + 毛利率(30%) + 净利率(30%)
    roe_score = 0.0
    if metrics.get("roe") is not None:
        roe_score = 0.0 if metrics["roe"] < 0 else percentile_rank(metrics["roe"], pool_values("roe"))
    gross_score = percentile_rank(metrics["gross_margin"], pool_values("gross_margin")) if metrics.get("gross_margin") is not None else 0.0
    net_score = percentile_rank(metrics["net_margin"], pool_values("net_margin")) if metrics.get("net_margin") is not None else 0.0
    profit_score = (roe_score * 0.4 + gross_score * 0.3 + net_score * 0.3) * discount

    # 成长性 (30%): 营收同比(40%) + 净利润同比(60%)
    rev_score = percentile_rank(metrics["revenue_yoy"], pool_values("revenue_yoy")) if metrics.get("revenue_yoy") is not None else 0.0
    prof_score = percentile_rank(metrics["profit_yoy"], pool_values("profit_yoy")) if metrics.get("profit_yoy") is not None else 0.0
    growth_score = (rev_score * 0.4 + prof_score * 0.6) * discount

    # 现金流质量 (15%)
    ocf_score = 0.0
    if metrics.get("ocf_to_profit") is not None:
        ocf_score = percentile_rank(metrics["ocf_to_profit"], pool_values("ocf_to_profit"))
    ocf_score *= discount

    # 偿债风险 (20%)：资产负债率越低越好
    debt_score = 0.0
    if metrics.get("debt_ratio") is not None:
        debt_score = percentile_rank(metrics["debt_ratio"], pool_values("debt_ratio"), reverse=True)
    debt_score *= discount

    # 总分
    total = profit_score * 0.35 + growth_score * 0.30 + ocf_score * 0.15 + debt_score * 0.20

    # 完整度惩罚
    completeness, level = calc_completeness(metrics)
    if level == "low":
        total *= Config.LOW_COMPLETENESS_PENALTY
    elif level == "ultra_low":
        total *= Config.LOW_COMPLETENESS_PENALTY * Config.ULTRA_LOW_COMPLETENESS_PENALTY

    # 净利润+现金流双负惩罚
    if (metrics.get("net_profit") is not None and metrics["net_profit"] < 0 and
        metrics.get("ocf_abs") is not None and metrics["ocf_abs"] < 0):
        total = min(total, Config.NEGATIVE_PROFIT_PENALTY)

    # 评级
    if total >= 75: grade = "A"
    elif total >= 55: grade = "B"
    elif total >= 40: grade = "C"
    elif total >= 25: grade = "D"
    else: grade = "E"

    confidence = {"high": "高", "medium": "中", "low": "低", "ultra_low": "低"}[level]

    return {
        "ts_code": ts_code, "name": stock.get("name", ""), "industry_l1": industry,
        "total_score": round(total, 2),
        "profit_score": round(profit_score, 2), "growth_score": round(growth_score, 2),
        "ocf_score": round(ocf_score, 2), "debt_score": round(debt_score, 2),
        "grade": grade, "confidence": confidence,
        "completeness": completeness, "completeness_level": level,
        "roe": metrics.get("roe"), "gross_margin": metrics.get("gross_margin"),
        "net_margin": metrics.get("net_margin"), "revenue_yoy": metrics.get("revenue_yoy"),
        "profit_yoy": metrics.get("profit_yoy"), "debt_ratio": metrics.get("debt_ratio"),
        "net_profit": metrics.get("net_profit"), "deducted_profit": metrics.get("deducted_profit"),
        "revenue": metrics.get("revenue"), "ocf_to_profit": metrics.get("ocf_to_profit"),
        "ocf_abs": metrics.get("ocf_abs"), "asset_turnover": metrics.get("asset_turnover"),
        "ar_turnover": metrics.get("ar_turnover"),
        "fetch_success": metrics.get("fetch_success", False),
        "annual_report_date": metrics.get("annual_report_date"),
        "market_fallback": use_market_fallback,
    }

# ============================================================
# 股票列表加载
# ============================================================
def load_stock_list(file_path=Config.DEFAULT_STOCK_FILE) -> List[Dict]:
    """从 xuan.txt 加载股票列表，滤除北交所/科创板等"""
    stocks = []
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line: continue
            parts = line.split()
            if len(parts) < 2: parts = line.split(",")
            if len(parts) < 2: parts = line.split("\t")
            if len(parts) < 2: continue
            code = parts[0].strip()
            name = parts[1].strip()
            if code.startswith(("688", "430", "83", "87")): continue
            if "." not in code:
                code = code + ".SH" if code.startswith("6") else code + ".SZ"
            stocks.append({"ts_code": code, "name": name})
    logger.info(f"加载股票列表: {len(stocks)} 只")
    return stocks

# ============================================================
# 自测模式
# ============================================================
def run_self_test():
    """运行内置自测，验证段落提取与指标解析"""
    logger.info("=" * 60)
    logger.info("开始自测模式 (V5.0.0)")
    logger.info("=" * 60)
    passed = 0
    failed = 0

    # 1. 验证年报段落提取
    block = _extract_annual_block(MOCK_NEODATA_RESPONSE)
    if block and "加权净资产收益率ROE15.67%" in block:
        logger.info("[PASS] 年报段落提取正确")
        passed += 1
    else:
        logger.error("[FAIL] 年报段落提取失败")
        failed += 1

    # 验证 Q1/Q3 不在年报段落中
    if "20230331" not in block and "20240930" not in block:
        logger.info("[PASS] Q1/Q4 段落已正确排除")
        passed += 1
    else:
        logger.error("[FAIL] Q1/Q4 段落未排除")
        failed += 1

    # 2. 验证指标解析
    metrics = parse_financial_all(block)
    tests = [
        ("roe", 15.67), ("gross_margin", 42.35), ("net_margin", 18.22),
        ("revenue_yoy", 28.45), ("profit_yoy", 35.67), ("debt_ratio", 38.92),
        ("asset_turnover", 0.85), ("ar_turnover", 6.78),
    ]
    for key, expected in tests:
        actual = metrics.get(key)
        if actual is not None and abs(actual - expected) < 0.01:
            logger.info(f"[PASS] {key}: {actual}")
            passed += 1
        else:
            logger.error(f"[FAIL] {key}: 期望 {expected}, 实际 {actual}")
            failed += 1

    # 3. 验证净利润（关键：不能是净利润现金含量行）
    net_profit = metrics.get("net_profit")
    if net_profit is not None and abs(net_profit - 1642130865.33) < 1:
        logger.info(f"[PASS] 净利润: {net_profit} (正确提取，非现金含量行)")
        passed += 1
    else:
        logger.error(f"[FAIL] 净利润: {net_profit} (期望 1642130865.33)")
        failed += 1

    # 4. 验证扣非净利润
    deducted = metrics.get("deducted_profit")
    if deducted is not None and abs(deducted - 1523456789.01) < 1:
        logger.info(f"[PASS] 扣非净利润: {deducted}")
        passed += 1
    else:
        logger.error(f"[FAIL] 扣非净利润: {deducted}")
        failed += 1

    # 5. 验证经营现金流
    ocf = metrics.get("ocf_abs")
    if ocf is not None and abs(ocf - 2156789012.34) < 1:
        logger.info(f"[PASS] 经营现金流净额: {ocf}")
        passed += 1
    else:
        logger.error(f"[FAIL] 经营现金流净额: {ocf}")
        failed += 1

    # 6. 验证 OCF/净利润
    ocf_ratio = metrics.get("ocf_to_profit")
    expected_ratio = round(2156789012.34 / 1642130865.33 * 100, 2)
    if ocf_ratio is not None and abs(ocf_ratio - expected_ratio) < 0.1:
        logger.info(f"[PASS] OCF/净利润: {ocf_ratio}%")
        passed += 1
    else:
        logger.error(f"[FAIL] OCF/净利润: {ocf_ratio} (期望 {expected_ratio})")
        failed += 1

    logger.info(f"自测完成: {passed} 通过, {failed} 失败")
    return failed == 0

# ============================================================
# 输出模块
# ============================================================
def output_excel(results: List[Dict], output_dir=Config.OUTPUT_DIR):
    """输出 Excel 报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.error("openpyxl 未安装，跳过 Excel 输出")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"股票业绩评价_{timestamp}.xlsx")
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    grade_fills = {
        "A": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
        "B": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        "C": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "D": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "E": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }

    # Sheet 1: 综合评价结果
    ws = wb.active
    ws.title = "综合评价结果"
    headers = ["排名","股票代码","股票名称","申万一级行业","总分","评级","置信度",
               "盈利能力","成长性","现金流质量","偿债风险",
               "ROE(%)","毛利率(%)","净利率(%)","营收同比(%)","净利润同比(%)",
               "资产负债率(%)","净利润(元)","扣非净利润(元)","营业总收入(元)",
               "经营现金流(元)","OCF/净利润(%)","总资产周转率","应收账款周转率",
               "年报日期","数据完整度","市场基准"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sorted_results = sorted(results, key=lambda x: x.get("total_score", 0), reverse=True)
    for rank, r in enumerate(sorted_results, 1):
        row = [
            rank, r.get("ts_code",""), r.get("name",""), r.get("industry_l1",""),
            r.get("total_score",0), r.get("grade",""), r.get("confidence",""),
            r.get("profit_score",0), r.get("growth_score",0),
            r.get("ocf_score",0), r.get("debt_score",0),
            r.get("roe"), r.get("gross_margin"), r.get("net_margin"),
            r.get("revenue_yoy"), r.get("profit_yoy"), r.get("debt_ratio"),
            r.get("net_profit"), r.get("deducted_profit"), r.get("revenue"),
            r.get("ocf_abs"), r.get("ocf_to_profit"),
            r.get("asset_turnover"), r.get("ar_turnover"),
            r.get("annual_report_date"), f"{r.get('completeness',0)*100:.0f}%",
            "是" if r.get("market_fallback") else "否",
        ]
        ws.append(row)
        # 评级着色
        grade = r.get("grade","")
        if grade in grade_fills:
            ws.cell(row=ws.max_row, column=6).fill = grade_fills[grade]

    # 调整列宽
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    # Sheet 2-6: 按评级分组
    for g in ["A","B","C","D","E"]:
        ws_g = wb.create_sheet(f"{g}级股票")
        for col, h in enumerate(headers, 1):
            cell = ws_g.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r in sorted_results:
            if r.get("grade") == g:
                row_data = [
                    0, r.get("ts_code",""), r.get("name",""), r.get("industry_l1",""),
                    r.get("total_score",0), r.get("grade",""), r.get("confidence",""),
                    r.get("profit_score",0), r.get("growth_score",0),
                    r.get("ocf_score",0), r.get("debt_score",0),
                    r.get("roe"), r.get("gross_margin"), r.get("net_margin"),
                    r.get("revenue_yoy"), r.get("profit_yoy"), r.get("debt_ratio"),
                    r.get("net_profit"), r.get("deducted_profit"), r.get("revenue"),
                    r.get("ocf_abs"), r.get("ocf_to_profit"),
                    r.get("asset_turnover"), r.get("ar_turnover"),
                    r.get("annual_report_date"), f"{r.get('completeness',0)*100:.0f}%",
                    "是" if r.get("market_fallback") else "否",
                ]
                ws_g.append(row_data)

    # Sheet: 统计概览
    ws_stats = wb.create_sheet("统计概览")
    ws_stats.append(["项目","数值"])
    ws_stats.append(["总股票数", len(results)])
    ws_stats.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    for g in ["A","B","C","D","E"]:
        cnt = sum(1 for r in results if r.get("grade") == g)
        ws_stats.append([f"{g}级股票数", cnt])
    success_cnt = sum(1 for r in results if r.get("fetch_success"))
    ws_stats.append(["成功获取数据", success_cnt])
    ws_stats.append(["获取失败", len(results) - success_cnt])

    wb.save(file_path)
    logger.info(f"Excel 报告已保存: {file_path}")
    return file_path

def output_json(results: List[Dict], output_dir=Config.OUTPUT_DIR):
    """输出 JSON 数据，带包装层"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"股票分析数据_{timestamp}.json")
    sorted_results = sorted(results, key=lambda x: x.get("total_score", 0), reverse=True)
    output = {
        "data_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_stocks": len(results),
        "stocks": sorted_results,
    }
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f"JSON 数据已保存: {file_path}")
    return file_path

# ============================================================
# 主函数
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="A股智能选股系统 V5.0.0")
    parser.add_argument("--base-dir", default=Config.BASE_DIR, help="工作目录")
    parser.add_argument("--stock-file", default=Config.DEFAULT_STOCK_FILE, help="股票列表文件")
    parser.add_argument("--workers", type=int, default=Config.FINANCE_WORKERS, help="并发线程数")
    parser.add_argument("--force-refresh", action="store_true", help="忽略缓存全量更新")
    parser.add_argument("--no-industry-patch", action="store_true", help="禁用行业API补调")
    parser.add_argument("--timeout", type=int, default=Config.GLOBAL_TIMEOUT, help="全局超时秒数")
    parser.add_argument("--test", action="store_true", help="运行内置自测并退出")
    args = parser.parse_args()

    if args.test:
        ok = run_self_test()
        sys.exit(0 if ok else 1)

    Config.BASE_DIR = args.base_dir
    Config.DEFAULT_STOCK_FILE = args.stock_file
    Config.OUTPUT_DIR = args.base_dir
    Config.FINANCE_WORKERS = args.workers
    Config.GLOBAL_TIMEOUT = args.timeout

    logger.info("=" * 60)
    logger.info("A股智能选股系统 V5.0.0 启动")
    logger.info("=" * 60)

    # 1. 加载 token
    try:
        token = load_token()
    except Exception as e:
        logger.error(f"加载 Token 失败: {e}")
        sys.exit(1)

    # 2. 加载股票列表
    stocks = load_stock_list(args.stock_file)
    if not stocks:
        logger.error("股票列表为空")
        sys.exit(1)

    # 3. 初始化数据库
    conn = init_db()

    # 4. 获取财务数据
    fetch_results = fetch_stock_batch(stocks, token, workers=args.workers,
                                       force_refresh=args.force_refresh, conn=conn)

    # 5. 保存到数据库
    industry_map = load_industry_map()
    for r in fetch_results:
        ts_code = r["ts_code"]
        name = r.get("name", "")
        metrics = r.get("metrics", {})
        report_date = r.get("report_date", "")
        success = r.get("fetch_success", False)
        content = r.get("content", "")

        if success and report_date:
            save_report(conn, ts_code, report_date, metrics, True)
        elif not success:
            save_report(conn, ts_code, "", {"fetch_success": False}, False)

        # 确定行业
        ind = determine_industry(ts_code, name, content, industry_map, use_api=False)
        if ind:
            save_stock_industry(conn, ts_code, name, ind)

    # 6. 行业补调
    if not args.no_industry_patch:
        # 从数据库重新加载以获取已保存的行业
        all_stocks = merge_latest_reports(conn, stocks)
        batch_industry_patch(all_stocks, industry_map)
        # 保存补调结果
        for s in all_stocks:
            if s.get("industry_l1"):
                save_stock_industry(conn, s["ts_code"], s["name"], s["industry_l1"], s.get("industry_l2",""))
    else:
        all_stocks = merge_latest_reports(conn, stocks)

    # 7. 评分计算
    # 按行业分组
    industry_groups: Dict[str, List[Dict]] = {}
    for s in all_stocks:
        ind = s.get("industry_l1", "未知")
        if ind not in industry_groups:
            industry_groups[ind] = []
        industry_groups[ind].append(s)

    scored = []
    for s in all_stocks:
        score = calc_score(s, industry_groups, all_stocks)
        scored.append(score)

    # 8. 输出
    excel_path = output_excel(scored)
    json_path = output_json(scored)

    conn.close()
    logger.info("=" * 60)
    logger.info(f"分析完成！Excel: {excel_path}, JSON: {json_path}")
    logger.info("=" * 60)

if __name__ == "__main__":
    main()



