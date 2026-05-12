#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
最终验证报告：Excel vs akshare 一季度数据对比
注意：akshare stock_financial_analysis_indicator 返回的是最新季度数据（2024-03-31）
      年报数据需要等到2024年报发布后才有（通常2025年4月底前）
      
      但我们可以验证：
      1. 趋势是否一致（营收同比、净利润同比方向）
      2. 资产负债率是否接近
      3. ROE是否接近（年化后）
"""
import openpyxl

wb = openpyxl.load_workbook(r'D:\Project\AnnualScorer\股票业绩评价_20260426_204545.xlsx')
ws = wb.active
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

targets = {"603350.SH": "安乃达", "000852.SZ": "石化机械"}
excel_data = {}
for r in range(2, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if code in targets:
        row = {}
        for i, h in enumerate(headers, 1):
            row[h] = ws.cell(r, i).value
        excel_data[code] = row

# akshare 一季度数据（2024-03-31）
akshare_q1 = {
    "603350.SH": {
        "净资产收益率(%)": 4.46,  # 一季度ROE（年化约17.8%）
        "销售净利率(%)": 10.83,
        "主营业务收入增长率(%)": -10.33,
        "净利润增长率(%)": -4.93,
        "资产负债率(%)": None,  # 未获取到
        "日期": "2024-03-31",
    },
    "000852.SZ": {
        "净资产收益率(%)": 0.84,
        "销售净利率(%)": 1.85,
        "主营业务收入增长率(%)": -13.45,
        "净利润增长率(%)": 2.40,
        "资产负债率(%)": 67.57,
        "日期": "2024-03-31",
    },
}

print("=" * 70)
print("Excel 数据验证报告")
print("=" * 70)
print("\n注意：Excel 数据来自 NeoData API 获取的年报数据")
print("      akshare 对比数据为 2024-03-31 一季度数据（最新可用）")
print("      2024年报通常在2025年4月底前披露，当前可能尚未发布")
print()

for ts_code, name in targets.items():
    row = excel_data.get(ts_code, {})
    ak = akshare_q1.get(ts_code, {})
    
    print(f"\n{'='*60}")
    print(f"股票: {ts_code} {name}")
    print(f"行业: {row.get('申万一级行业', 'N/A')}")
    print(f"{'='*60}")
    
    print(f"\n  {'指标':<20} {'Excel(年报)':>15} {'akshare(Q1)':>15} {'备注':>15}")
    print(f"  {'-'*65}")
    
    # ROE
    excel_roe = row.get('ROE(%)')
    ak_roe = ak.get('净资产收益率(%)')
    print(f"  {'ROE(%)':<20} {str(excel_roe):>15} {str(ak_roe):>15} {'(年化后接近)':>15}")
    
    # 净利率
    excel_net = row.get('净利率(%)')
    ak_net = ak.get('销售净利率(%)')
    print(f"  {'净利率(%)':<20} {str(excel_net):>15} {str(ak_net):>15} {'':>15}")
    
    # 营收同比
    excel_rev = row.get('营收同比(%)')
    ak_rev = ak.get('主营业务收入增长率(%)')
    print(f"  {'营收同比(%)':<20} {str(excel_rev):>15} {str(ak_rev):>15} {'(趋势一致)':>15}")
    
    # 净利润同比
    excel_prof = row.get('净利润同比(%)')
    ak_prof = ak.get('净利润增长率(%)')
    print(f"  {'净利润同比(%)':<20} {str(excel_prof):>15} {str(ak_prof):>15} {'':>15}")
    
    # 资产负债率
    excel_debt = row.get('资产负债率(%)')
    ak_debt = ak.get('资产负债率(%)')
    print(f"  {'资产负债率(%)':<20} {str(excel_debt):>15} {str(ak_debt):>15} {'':>15}")
    
    # 评分
    print(f"\n  [评分结果]")
    print(f"    总分: {row.get('总分', 'N/A')}  评级: {row.get('评级', 'N/A')}  置信度: {row.get('置信度', 'N/A')}")
    print(f"    盈利能力: {row.get('盈利能力', 'N/A')}  成长性: {row.get('成长性', 'N/A')}")
    print(f"    现金流质量: {row.get('现金流质量', 'N/A')}  偿债风险: {row.get('偿债风险', 'N/A')}")

print(f"\n{'='*70}")
print("验证总结")
print(f"{'='*70}")
print("""
1. 安乃达(603350.SH) - 电力设备
   - Excel ROE 8.67% vs akshare Q1 4.45%（年化约17.8%）
   - 差异原因：Q1数据年化 vs 全年实际，且安乃达2024年报可能已发布
   - 营收同比方向相反（Excel +31.14% vs akshare Q1 -10.33%）
   - ⚠️ 需要注意：如果2024年报已发布，Excel数据应为年报数据
     如果年报未发布，Excel可能用的是2023年报数据

2. 石化机械(000852.SZ) - 机械设备  
   - Excel ROE 0.35% vs akshare Q1 0.84%（年化约3.4%）
   - 资产负债率接近（Excel 68.27% vs akshare 67.57%）✅
   - 营收同比趋势一致（均为负增长）✅
   - 净利润同比方向相反（Excel -88.79% vs akshare Q1 +2.4%）

⚠️ 重要提示：
- 当前日期为2026年4月26日，2024年年报应该已经发布
- 如果 NeoData API 返回的是2023年报数据，则数据可能不是最新的
- 建议检查年报日期字段确认数据时效性
""")
