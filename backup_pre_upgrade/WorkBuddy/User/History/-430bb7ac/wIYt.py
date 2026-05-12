#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用 akshare 验证 Excel 中2只股票的财务数据
"""
import akshare as ak
import openpyxl
import warnings
warnings.filterwarnings("ignore")

# 读取Excel
wb = openpyxl.load_workbook(r'D:\Project\AnnualScorer\股票业绩评价_20260426_204545.xlsx')
ws = wb.active
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

targets = {"603350.SH": "安乃达", "000852.SZ": "石化机械"}
excel_data = {}
for r in range(2, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if code in targets:
        row = {}
        for i, h in enumerate(headers, 1):
            row[h] = ws.cell(r, i).value
        excel_data[code] = row

# akshare 验证
for ts_code, name in targets.items():
    code = ts_code.split(".")[0]
    row = excel_data.get(ts_code, {})
    
    print(f"\n{'='*60}")
    print(f"股票: {ts_code} {name}")
    print(f"{'='*60}")
    
    print(f"\n  [Excel数据]")
    print(f"    ROE: {row.get('ROE(%)', 'N/A')}%")
    print(f"    毛利率: {row.get('毛利率(%)', 'N/A')}%")
    print(f"    净利率: {row.get('净利率(%)', 'N/A')}%")
    print(f"    营收同比: {row.get('营收同比(%)', 'N/A')}%")
    print(f"    净利润同比: {row.get('净利润同比(%)', 'N/A')}%")
    print(f"    资产负债率: {row.get('资产负债率(%)', 'N/A')}%")
    print(f"    归母净利润: {row.get('净利润(元)', 'N/A')}")
    print(f"    经营现金流: {row.get('经营现金流(元)', 'N/A')}")
    print(f"    OCF/净利润: {row.get('OCF/净利润(%)', 'N/A')}%")
    print(f"    总分: {row.get('总分', 'N/A')}  评级: {row.get('评级', 'N/A')}")
    
    # akshare 财务指标
    print(f"\n  [akshare 财务指标]")
    try:
        df = ak.stock_financial_analysis_indicator(symbol=code, start_year="2024")
        if df is not None and len(df) > 0:
            col_map = {
                "净资产收益率": "ROE(%)",
                "销售毛利率": "毛利率(%)",
                "销售净利率": "净利率(%)",
                "资产负债率": "资产负债率(%)",
            }
            for ak_col, excel_col in col_map.items():
                if ak_col in df.columns:
                    ak_val = df[ak_col].iloc[0]
                    excel_val = row.get(excel_col, "N/A")
                    try:
                        diff = abs(float(ak_val) - float(excel_val))
                        status = "✅" if diff < 0.5 else f"⚠️ 差{diff:.2f}"
                    except:
                        status = "N/A"
                    print(f"    {excel_col}: akshare={ak_val}  Excel={excel_val}  {status}")
        else:
            print("  无数据")
    except Exception as e:
        print(f"  错误: {e}")
    
    # akshare 利润表 - 营收和净利润
    print(f"\n  [akshare 利润表]")
    try:
        df2 = ak.stock_financial_report_sina(stock=ts_code, symbol="利润表")
        if df2 is not None and len(df2) > 0:
            # 找2024年报
            date_col = df2.columns[0]
            for idx, r in df2.iterrows():
                date_val = str(r[date_col])
                if "2024" in date_val and "1231" in date_val:
                    print(f"    报表日期: {date_val}")
                    for c in df2.columns[1:]:
                        v = r[c]
                        if v and str(v).strip():
                            print(f"      {c}: {v}")
                    break
    except Exception as e:
        print(f"  错误: {e}")

print(f"\n{'='*60}")
print("验证完成")
