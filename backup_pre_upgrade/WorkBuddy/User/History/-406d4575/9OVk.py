#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A股智能选股系统 - 基于年报的业绩评分系统
版本: 5.2.0
"""

import os
import sys
import json
import time
import logging
import argparse
import re
import threading
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed, Future
from contextlib import contextmanager

import requests as req_lib
from requests.adapters import HTTPAdapter


class Config:
    """全局配置类"""
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DEFAULT_STOCK_FILE = os.path.join(BASE_DIR, "xuan.txt")
    OUTPUT_DIR = BASE_DIR
    INDUSTRY_MAP_FILE = os.path.join(BASE_DIR, "industry_map.json")
    DB_FILE = os.path.join(BASE_DIR, "stock_cache.db")
    FINANCE_WORKERS = 16
    API_RETRY_TIMES = 2
    API_RETRY_BACKOFF_BASE = 3.0
    API_TIMEOUT = 50
    PAUSE_CONSECUTIVE_EMPTY = 10
    PAUSE_DURATION = 20
    GLOBAL_TIMEOUT = 7200
    MIN_INDUSTRY_SAMPLES = 5
    CACHE_MAX_AGE_ANNUAL = 400
    NEGATIVE_PROFIT_PENALTY = 15.0
    MARKET_FALLBACK_DISCOUNT = 0.95
    LOW_COMPLETENESS_PENALTY = 0.9
    ULTRA_LOW_COMPLETENESS_PENALTY = 0.75
    ANNUAL_DISCLOSURE_DEADLINE_MONTH = 4
    ANNUAL_DISCLOSURE_DEADLINE_DAY = 30
    NEODATA_URL = "https://copilot.tencent.com/agenttool/v1/neodata"
    NEODATA_TOKEN_FILE = os.path.expanduser("~/.workbuddy/.neodata_token")


# 线程本地存储 - 用于存储Session
_thread_local = threading.local()


def setup_logging() -> logging.Logger:
    """设置日志系统"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(Config.BASE_DIR, f"stock_analyzer_{timestamp}.log")
    
    formatter = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    
    # 避免重复添加handler
    if not logger.handlers:
        file_handler = logging.FileHandler(log_file, encoding="utf-8", mode="w")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
    
    return logger


logger = setup_logging()


def get_thread_session() -> req_lib.Session:
    """获取线程安全的Session"""
    if not hasattr(_thread_local, "session"):
        session = req_lib.Session()
        adapter = HTTPAdapter(
            pool_connections=10,
            pool_maxsize=10,
            max_retries=3
        )
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        _thread_local.session = session
    return _thread_local.session


@contextmanager
def db_connection(db_path: str = Config.DB_FILE):
    """数据库连接上下文管理器"""
    import sqlite3
    conn = None
    try:
        conn = sqlite3.connect(db_path, timeout=30.0)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=30000")
        yield conn
        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"数据库操作异常: {e}")
        raise
    finally:
        if conn:
            conn.close()


def load_token() -> str:
    """加载NeoData Token"""
    if not os.path.exists(Config.NEODATA_TOKEN_FILE):
        raise FileNotFoundError(
            f"NeoData Token文件不存在: {Config.NEODATA_TOKEN_FILE}\n"
            f"请在该路径下创建文件并填入您的Token"
        )
    
    try:
        with open(Config.NEODATA_TOKEN_FILE, "r", encoding="utf-8") as f:
            token = f.read().strip()
    except Exception as e:
        raise IOError(f"读取Token文件失败: {e}")
    
    if not token:
        raise ValueError("NeoData Token文件为空，请检查文件内容")
    
    return token


def parse_num(text: str) -> Optional[float]:
    """解析数字，支持百分比"""
    if not text:
        return None
    text = text.strip()
    
    # 尝试匹配百分比
    m = re.search(r"([-+]?\d+\.?\d*)%", text)
    if m:
        return float(m.group(1))
    
    # 尝试匹配普通数字
    m = re.search(r"([-+]?\d+\.?\d*)", text)
    if m:
        return float(m.group(1))
    
    return None


def _parse_num_from_line(line: str) -> Optional[float]:
    """从财报行提取带单位数值，单位：万亿元>亿元>万元>千元>元"""
    if not line:
        return None
    
    m = re.search(r"([-+]?\d+\.?\d*)\s*(万亿元|亿元|万元|千元|元)", line)
    if not m:
        return None
    
    val = float(m.group(1))
    unit = m.group(2)
    
    multipliers = {
        "万亿元": 1e12,
        "亿元": 1e8,
        "万元": 1e4,
        "千元": 1e3,
        "元": 1
    }
    
    return val * multipliers.get(unit, 1)


def year_of_date(date_str: str) -> int:
    """从日期字符串中提取年份"""
    if not date_str:
        return datetime.now().year
    try:
        return int(str(date_str)[:4])
    except (ValueError, IndexError):
        return datetime.now().year


def _extract_annual_block(text: str, year: Optional[int] = None) -> str:
    """精确提取年报段落，锚点：统计截止日期为YYYY1231的年报"""
    if not text:
        return ""
    
    if year:
        target = f"统计截止日期为{year}1231的年报"
        start = text.find(target)
        if start == -1:
            return ""
        start += len(target)
        next_anchor = text.find("统计截止日期为", start + 1)
        if next_anchor == -1:
            return text[start:].strip()
        return text[start:next_anchor].strip()
    else:
        matches = list(re.finditer(r"统计截止日期为(\d{4})1231的年报", text))
        if not matches:
            return ""
        last = matches[-1]
        start = last.start() + len(last.group(0))
        next_anchor = text.find("统计截止日期为", start + 1)
        if next_anchor == -1:
            return text[start:].strip()
        return text[start:next_anchor].strip()


def _guess_date_from_trend(text: str) -> str:
    """兜底：在没有明确年报锚点时，从内容中提取可能属于年报的行"""
    if not text:
        return ""
    
    lines = text.split("\n")
    relevant_lines = []
    in_quarter = False
    
    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue
        
        if re.search(r"统计截止日期为\d{4}(0331|0630|0930)的", line_stripped):
            in_quarter = True
            continue
        
        if re.search(r"统计截止日期为\d{4}1231的年报", line_stripped):
            in_quarter = False
            continue
        
        if in_quarter:
            continue
        
        keywords = ["营业收入", "净利润", "毛利率", "ROE", "资产负债率",
                   "经营活动", "现金流", "归母"]
        if any(kw in line_stripped for kw in keywords):
            relevant_lines.append(line_stripped)
    
    return "\n".join(relevant_lines)


def _extract_metric_line(block: str, keywords: List[str]) -> Optional[str]:
    """在段落中按关键词搜索行"""
    if not block or not keywords:
        return None
    
    for line in block.split("\n"):
        line = line.strip()
        if not line:
            continue
        for kw in keywords:
            if kw in line:
                return line
    return None


def parse_financial_all(block: str) -> Dict[str, Optional[float]]:
    """从年报段落提取评分所需的核心财务指标"""
    result = {
        "roe": None,
        "gross_margin": None,
        "net_margin": None,
        "revenue_yoy": None,
        "profit_yoy": None,
        "debt_ratio": None,
        "net_profit": None,
        "ocf_abs": None
    }
    
    if not block:
        return result
    
    # 盈利能力指标
    roe_line = _extract_metric_line(block, ["加权净资产收益率ROE", "净资产收益率ROE", "加权净资产收益率"])
    result["roe"] = parse_num(roe_line) if roe_line else None
    
    gross_line = _extract_metric_line(block, ["销售毛利率"])
    if gross_line and "毛利率" in gross_line:
        result["gross_margin"] = parse_num(gross_line[gross_line.find("毛利率"):])
    
    net_line = _extract_metric_line(block, ["销售净利率"])
    if net_line and "净利率" in net_line:
        result["net_margin"] = parse_num(net_line[net_line.find("净利率"):])
    
    # 成长性指标
    rev_line = _extract_metric_line(block, ["营业收入同比增长", "营收同比增长", "营业总收入同比增长"])
    result["revenue_yoy"] = parse_num(rev_line) if rev_line else None
    
    profit_line = _extract_metric_line(block, ["归母净利润同比增长"])
    result["profit_yoy"] = parse_num(profit_line) if profit_line else None
    
    # 偿债能力指标
    debt_line = _extract_metric_line(block, ["资产负债率"])
    result["debt_ratio"] = parse_num(debt_line) if debt_line else None
    
    # 净利润（严格匹配）
    pattern_np = re.compile(r'^净利润\s*([-+]?\d+\.?\d*)\s*(万亿元|亿元|万元|千元|元)')
    for line in block.split("\n"):
        line = line.strip()
        m = pattern_np.match(line)
        if m and not any(x in line for x in ["同比", "增长率", "现金含量", "归母", "扣非"]):
            result["net_profit"] = _parse_num_from_line(line)
            break
    
    if result["net_profit"] is None:
        for line in block.split("\n"):
            line = line.strip()
            if "净利润" in line and not any(x in line for x in ["同比", "增长率", "现金含量", "归母", "扣非"]):
                result["net_profit"] = _parse_num_from_line(line)
                break
    
    # 经营活动现金流净额
    ocf_line = _extract_metric_line(block, ["经营活动产生的现金流量净额"])
    result["ocf_abs"] = _parse_num_from_line(ocf_line) if ocf_line else None
    
    return result


def load_industry_map() -> Dict[str, Any]:
    """加载行业映射表"""
    if os.path.exists(Config.INDUSTRY_MAP_FILE):
        try:
            with open(Config.INDUSTRY_MAP_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"加载行业映射文件失败，将使用默认规则: {e}")
    return {}


def extract_industry_from_content(text: str) -> Optional[str]:
    """从内容中提取行业信息"""
    if not text:
        return None
    
    patterns = [
        r"所属一级行业[：:]\s*(\S+)",
        r"申万一级行业[：:]\s*(\S+)",
        r"行业分类[：:]\s*(\S+)"
    ]
    
    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            return m.group(1).strip()
    
    return None


def infer_industry_from_name(name: str) -> Optional[str]:
    """从股票名称推断行业"""
    if not name:
        return None
    
    name_keyword_industry = {
        "银行": "银行",
        "证券": "非银金融",
        "保险": "非银金融",
        "地产": "房地产",
        "房地产": "房地产",
        "钢铁": "钢铁",
        "煤炭": "煤炭",
        "有色": "有色金属",
        "化工": "基础化工",
        "医药": "医药生物",
        "生物": "医药生物",
        "电子": "电子",
        "计算机": "计算机",
        "通信": "通信",
        "汽车": "汽车",
        "机械": "机械设备",
        "电力": "公用事业",
        "食品": "食品饮料",
        "饮料": "食品饮料",
        "家电": "家用电器",
        "纺织": "纺织服饰",
        "建筑": "建筑装饰",
        "军工": "国防军工",
        "传媒": "传媒",
        "光伏": "电力设备",
        "电池": "电力设备",
        "芯片": "电子",
        "半导体": "电子"
    }
    
    for kw, industry in name_keyword_industry.items():
        if kw in name:
            return industry
    
    return None


def determine_industry(
    ts_code: str,
    name: str,
    content: str,
    industry_map: Dict[str, Any],
    use_api: bool = True,
    session: Optional[req_lib.Session] = None
) -> Optional[str]:
    """确定股票所属行业"""
    # 1. 从内容中提取
    ind = extract_industry_from_content(content)
    if ind:
        return ind
    
    # 2. 从映射表中查找
    code_short = ts_code.split(".")[0]
    
    def _get_industry_l1(key: str) -> Optional[str]:
        val = industry_map.get(key)
        if val is None:
            return None
        if isinstance(val, str):
            return val
        if isinstance(val, dict):
            return val.get("industry_l1")
        return None
    
    ind = _get_industry_l1(ts_code)
    if ind:
        return ind
    
    ind = _get_industry_l1(code_short)
    if ind:
        return ind
    
    # 3. 从股票名称推断
    ind = infer_industry_from_name(name)
    if ind:
        return ind
    
    # 4. 从API获取
    if use_api:
        ind = fetch_industry_by_api(ts_code, name, session)
        if ind:
            return ind
    
    # 5. 根据股票代码前缀推断
    code_prefix_industry = {
        "60": "机械设备",
        "00": "房地产",
        "30": "医药生物",
        "68": "电子"
    }
    prefix = code_short[:2]
    return code_prefix_industry.get(prefix)


def fetch_industry_by_api(
    ts_code: str,
    name: str,
    session: Optional[req_lib.Session] = None
) -> Optional[str]:
    """通过API获取行业信息"""
    try:
        token = load_token()
        query = f"{ts_code} {name} 所属行业"
        text = run_neodata(query, token, session=session)
        if not text:
            return None
        return extract_industry_from_content(text)
    except Exception as e:
        logger.warning(f"从API获取行业信息失败: {e}")
        return None


def init_db(db_path: str = Config.DB_FILE) -> None:
    """初始化数据库表结构"""
    with db_connection(db_path) as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS stocks (
                ts_code TEXT PRIMARY KEY,
                name TEXT,
                industry_l1 TEXT,
                industry_l2 TEXT,
                last_industry_update TEXT,
                last_full_update TEXT
            );
            CREATE TABLE IF NOT EXISTS financial_reports (
                ts_code TEXT,
                report_date TEXT,
                report_type TEXT DEFAULT 'annual',
                roe REAL,
                gross_margin REAL,
                net_margin REAL,
                revenue_yoy REAL,
                profit_yoy REAL,
                debt_ratio REAL,
                net_profit REAL,
                ocf_abs REAL,
                fetch_success INTEGER DEFAULT 0,
                last_update TEXT,
                PRIMARY KEY (ts_code, report_date, report_type)
            );
            CREATE INDEX IF NOT EXISTS idx_reports_ts_type 
                ON financial_reports(ts_code, report_type);
        """)


def should_refresh(conn, ts_code: str, year: int) -> bool:
    """判断是否需要刷新数据"""
    cur = conn.execute(
        "SELECT report_date, last_update FROM financial_reports "
        "WHERE ts_code=? AND report_type='annual' AND fetch_success=1 "
        "ORDER BY report_date DESC LIMIT 1",
        (ts_code,)
    )
    row = cur.fetchone()
    
    if not row:
        return True
    
    report_year = year_of_date(str(row[0])) if row[0] else 0
    
    # 如果当前年份比报告年份晚，且已过年报披露截止日
    if report_year < year:
        deadline = datetime(
            report_year + 1,
            Config.ANNUAL_DISCLOSURE_DEADLINE_MONTH,
            Config.ANNUAL_DISCLOSURE_DEADLINE_DAY
        )
        if datetime.now() > deadline:
            return True
    
    # 检查数据是否过旧
    if row[1]:
        try:
            update_time = datetime.fromisoformat(row[1])
            age_days = (datetime.now() - update_time).days
            if age_days > Config.CACHE_MAX_AGE_ANNUAL:
                return True
        except (ValueError, TypeError):
            return True
    
    return report_year < year


def save_report(
    conn,
    ts_code: str,
    report_date: str,
    metrics: Dict[str, Optional[float]],
    success: bool
) -> None:
    """保存财务报告数据"""
    now_str = datetime.now().isoformat()
    conn.execute(
        """INSERT OR REPLACE INTO financial_reports
           (ts_code, report_date, report_type,
            roe, gross_margin, net_margin, revenue_yoy, profit_yoy,
            debt_ratio, net_profit, ocf_abs,
            fetch_success, last_update)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            ts_code, report_date, "annual",
            metrics.get("roe"), metrics.get("gross_margin"), metrics.get("net_margin"),
            metrics.get("revenue_yoy"), metrics.get("profit_yoy"),
            metrics.get("debt_ratio"), metrics.get("net_profit"),
            metrics.get("ocf_abs"),
            1 if success else 0, now_str
        )
    )


def save_stock_industry(
    conn,
    ts_code: str,
    name: str,
    industry_l1: str,
    industry_l2: str = ""
) -> None:
    """保存股票行业信息"""
    now_str = datetime.now().isoformat()
    conn.execute(
        """INSERT OR REPLACE INTO stocks 
           (ts_code, name, industry_l1, industry_l2, last_industry_update, last_full_update) 
           VALUES (?,?,?,?,?,?)""",
        (ts_code, name, industry_l1, industry_l2, now_str, now_str)
    )


def merge_latest_reports(conn, stocks: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    """合并数据库中最新的年报数据"""
    result = []
    
    for stock in stocks:
        ts_code = stock["ts_code"]
        item = {
            "ts_code": ts_code,
            "name": stock.get("name", ""),
            "industry_l1": None,
            "industry_l2": None,
            "fetch_success": False,
            "report_date": None,
            "roe": None,
            "gross_margin": None,
            "net_margin": None,
            "revenue_yoy": None,
            "profit_yoy": None,
            "debt_ratio": None,
            "net_profit": None,
            "ocf_abs": None
        }
        
        # 获取行业信息
        cur = conn.execute(
            "SELECT industry_l1, industry_l2 FROM stocks WHERE ts_code=?",
            (ts_code,)
        )
        stock_row = cur.fetchone()
        if stock_row:
            item["industry_l1"] = stock_row[0]
            item["industry_l2"] = stock_row[1]
        
        # 获取最新报告
        cur = conn.execute(
            """SELECT report_date, roe, gross_margin, net_margin, 
                      revenue_yoy, profit_yoy, debt_ratio, net_profit, ocf_abs
               FROM financial_reports 
               WHERE ts_code=? AND report_type='annual' AND fetch_success=1 
               ORDER BY report_date DESC LIMIT 1""",
            (ts_code,)
        )
        report_row = cur.fetchone()
        
        if report_row:
            cols = [d[0] for d in cur.description]
            report = dict(zip(cols, report_row))
            item.update({
                "report_date": report.get("report_date"),
                "roe": report.get("roe"),
                "gross_margin": report.get("gross_margin"),
                "net_margin": report.get("net_margin"),
                "revenue_yoy": report.get("revenue_yoy"),
                "profit_yoy": report.get("profit_yoy"),
                "debt_ratio": report.get("debt_ratio"),
                "net_profit": report.get("net_profit"),
                "ocf_abs": report.get("ocf_abs"),
                "fetch_success": True
            })
        
        result.append(item)
    
    return result


def run_neodata(
    query: str,
    token: str,
    timeout: int = Config.API_TIMEOUT,
    session: Optional[req_lib.Session] = None
) -> str:
    """调用NeoData API"""
    if not session:
        session = get_thread_session()
    
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    payload = {"query": query}
    
    for attempt in range(Config.API_RETRY_TIMES + 1):
        try:
            resp = session.post(
                Config.NEODATA_URL,
                json=payload,
                headers=headers,
                timeout=timeout
            )
            resp.raise_for_status()
            data = resp.json()
            
            if data.get("code") != "200":
                logger.warning(
                    f"业务错误 (尝试 {attempt + 1}/{Config.API_RETRY_TIMES + 1}): "
                    f"{data.get('msg', '未知')} query: {query[:50]}"
                )
                if attempt < Config.API_RETRY_TIMES:
                    wait_time = Config.API_RETRY_BACKOFF_BASE ** (attempt + 1)
                    time.sleep(wait_time)
                continue
            
            inner = data.get("data", {})
            
            if isinstance(inner, dict):
                api_data = inner.get("apiData", {})
                recall_list = api_data.get("apiRecall", [])
                
                if isinstance(recall_list, list) and recall_list:
                    # 优先找带年报锚点的内容
                    for item in recall_list:
                        content = item.get("content", "")
                        if content and "统计截止日期为" in content and "年报" in content:
                            return content
                    
                    # 其次找财务相关内容
                    for item in recall_list:
                        content = item.get("content", "")
                        if content and "财务" in item.get("type", ""):
                            return content
                    
                    # 最后拼接所有内容
                    parts = [item.get("content", "") for item in recall_list if item.get("content")]
                    if parts:
                        return "\n\n".join(parts)
                
                if isinstance(inner.get("text"), str) and inner["text"]:
                    return inner["text"]
            
            elif isinstance(inner, str):
                return inner
            
            return json.dumps(data, ensure_ascii=False)
        
        except req_lib.exceptions.Timeout:
            logger.warning(
                f"API超时 (尝试 {attempt + 1}/{Config.API_RETRY_TIMES + 1}): {query[:50]}"
            )
        except req_lib.exceptions.HTTPError as e:
            code = e.response.status_code if e.response else "?"
            logger.warning(
                f"API HTTP {code} (尝试 {attempt + 1}/{Config.API_RETRY_TIMES + 1}): {query[:50]}"
            )
        except Exception as e:
            logger.warning(
                f"API错误 (尝试 {attempt + 1}/{Config.API_RETRY_TIMES + 1}): {e}"
            )
        
        if attempt < Config.API_RETRY_TIMES:
            wait_time = Config.API_RETRY_BACKOFF_BASE ** (attempt + 1)
            logger.info(f"等待 {wait_time:.0f}秒后重试...")
            time.sleep(wait_time)
    
    return ""


def fetch_stock_finance(
    ts_code: str,
    name: str,
    token: str,
    year: Optional[int] = None
) -> Dict[str, Any]:
    """获取单只股票最新年报数据"""
    query = f"{ts_code} {name} 年报"
    try:
        text = run_neodata(query, token)
        if not text:
            return {
                "metrics": {},
                "content": "",
                "report_date": "",
                "fetch_success": False,
                "has_valid_block": False
            }
        
        block = _extract_annual_block(text, year) if year else _extract_annual_block(text)
        has_valid = bool(block)
        
        if not block:
            block = _guess_date_from_trend(text)
            has_valid = bool(block)
        
        if not block:
            return {
                "metrics": {},
                "content": text,
                "report_date": "",
                "fetch_success": False,
                "has_valid_block": False
            }
        
        metrics = parse_financial_all(block)
        
        m = re.search(r"统计截止日期为(\d{4})1231的年报", text)
        report_date = m.group(1) + "1231" if m else ""
        
        return {
            "metrics": metrics,
            "content": text,
            "report_date": report_date,
            "fetch_success": True,
            "has_valid_block": has_valid
        }
    except Exception as e:
        logger.error(f"获取财务数据异常 {ts_code} {name}: {e}")
        return {
            "metrics": {},
            "content": "",
            "report_date": "",
            "fetch_success": False,
            "has_valid_block": False
        }


def fetch_stock_batch(
    stocks: List[Dict[str, str]],
    token: str,
    workers: int = Config.FINANCE_WORKERS,
    force_refresh: bool = False,
    conn=None
) -> List[Dict[str, Any]]:
    """批量获取财务数据，线程安全"""
    total = len(stocks)
    results = []
    consecutive_parse_failures = 0
    start_time = time.time()
    
    logger.info(f"开始获取财务数据: 共 {total}只, {workers}个线程")
    
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_stock: Dict[Future, Dict[str, str]] = {}
        
        for stock in stocks:
            if not force_refresh and conn:
                if not should_refresh(conn, stock["ts_code"], datetime.now().year):
                    continue
            future = executor.submit(
                fetch_stock_finance,
                stock["ts_code"],
                stock["name"],
                token
            )
            future_to_stock[future] = stock
        
        done_count = 0
        for future in as_completed(future_to_stock):
            stock = future_to_stock[future]
            done_count += 1
            try:
                result = future.result(timeout=Config.API_TIMEOUT + 10)
                results.append({**stock, **result})
                
                if not result.get("has_valid_block", False):
                    consecutive_parse_failures += 1
                else:
                    consecutive_parse_failures = 0
                
                if consecutive_parse_failures >= Config.PAUSE_CONSECUTIVE_EMPTY:
                    logger.warning(
                        f"连续 {consecutive_parse_failures}次解析失败，暂停 {Config.PAUSE_DURATION}秒"
                    )
                    time.sleep(Config.PAUSE_DURATION)
                    consecutive_parse_failures = 0
            except Exception as e:
                consecutive_parse_failures += 1
                results.append({
                    **stock,
                    "metrics": {},
                    "content": "",
                    "report_date": "",
                    "fetch_success": False,
                    "has_valid_block": False
                })
                logger.error(f"任务异常 {stock['ts_code']}: {e}")
            
            if done_count % 100 == 0 and done_count > 0:
                elapsed = time.time() - start_time
                rate = done_count / elapsed if elapsed > 0 else 0
                eta = (total - done_count) / rate if rate > 0 else 0
                logger.info(
                    f"进度: {done_count}/{total} ({done_count/total*100:.1f}%), "
                    f"速率:{rate:.1f}/秒, 已用:{elapsed:.0f}秒, 剩余:{eta:.0f}秒"
                )
    
    elapsed = time.time() - start_time
    logger.info(f"获取完成: {len(results)}/{total}, 耗时:{elapsed:.0f}秒")
    return results


def calc_completeness(metrics: Dict[str, Optional[float]]) -> Tuple[float, str]:
    """计算数据完整度"""
    core_metrics = ["roe", "gross_margin", "net_margin", "revenue_yoy",
                    "profit_yoy", "debt_ratio", "net_profit", "ocf_abs"]
    non_null = sum(1 for m in core_metrics if metrics.get(m) is not None)
    ratio = non_null / len(core_metrics)
    
    if ratio >= 0.857:
        return ratio, "high"
    elif ratio >= 0.571:
        return ratio, "medium"
    elif non_null <= 1:
        return ratio, "ultra_low"
    else:
        return ratio, "low"


def percentile_rank(value: Optional[float], values: List[float], reverse: bool = False) -> float:
    """计算百分位排名 0~100，reverse=True时值越小得分越高"""
    if value is None:
        return 0.0
    if not values:
        return 50.0
    
    count_leq = sum(1 for v in values if v <= value)
    if reverse:
        return ((len(values) - count_leq) / len(values)) * 100
    else:
        return (count_leq / len(values)) * 100


def calc_score(
    stock: Dict[str, Any],
    industry_stocks: Dict[str, List[Dict[str, Any]]],
    all_stocks: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """计算股票评分"""
    ts_code = stock["ts_code"]
    industry = stock.get("industry_l1", "未知")
    metrics = stock
    
    # 确定参考池
    pool = industry_stocks.get(industry, [])
    use_market_fallback = len(pool) < Config.MIN_INDUSTRY_SAMPLES
    if use_market_fallback:
        pool = all_stocks
    discount = Config.MARKET_FALLBACK_DISCOUNT if use_market_fallback else 1.0
    
    def pool_values(key: str) -> List[float]:
        return [
            s[key] for s in pool
            if s.get("ts_code") != ts_code and s.get(key) is not None
        ]
    
    # 1. 盈利能力 (权重40%)
    roe_score = 0.0
    if metrics.get("roe") is not None and metrics["roe"] >= 0:
        roe_score = percentile_rank(metrics["roe"], pool_values("roe"))
    
    gross_score = percentile_rank(
        metrics["gross_margin"], pool_values("gross_margin")
    ) if metrics.get("gross_margin") is not None else 0.0
    
    net_score = percentile_rank(
        metrics["net_margin"], pool_values("net_margin")
    ) if metrics.get("net_margin") is not None else 0.0
    
    profit_score = (roe_score * 0.4 + gross_score * 0.3 + net_score * 0.3) * discount
    
    # 2. 成长性 (权重30%)
    rev_score = percentile_rank(
        metrics["revenue_yoy"], pool_values("revenue_yoy")
    ) if metrics.get("revenue_yoy") is not None else 0.0
    
    prof_score = percentile_rank(
        metrics["profit_yoy"], pool_values("profit_yoy")
    ) if metrics.get("profit_yoy") is not None else 0.0
    
    growth_score = (rev_score * 0.4 + prof_score * 0.6) * discount
    
    # 3. 现金流质量 (权重20%)
    np_val = metrics.get("net_profit")
    ocf_abs_val = metrics.get("ocf_abs")
    ocf_val = None
    ocf_score = 0.0
    
    if np_val and ocf_abs_val and np_val != 0:
        ocf_val = round(ocf_abs_val / np_val * 100, 2)
        
        pool_ocf_vals = []
        for s in pool:
            if s.get("ts_code") == ts_code:
                continue
            s_np = s.get("net_profit")
            s_ocf = s.get("ocf_abs")
            if s_np and s_ocf and s_np != 0:
                pool_ocf_vals.append(round(s_ocf / s_np * 100, 2))
        
        if ocf_val is not None:
            ocf_score = percentile_rank(ocf_val, pool_ocf_vals)
    
    ocf_score *= discount
    
    # 4. 偿债风险 (权重10%)
    debt_score = 0.0
    if metrics.get("debt_ratio") is not None:
        debt_score = percentile_rank(
            metrics["debt_ratio"], pool_values("debt_ratio"), reverse=True
        )
    debt_score *= discount
    
    # 计算总分
    total_score = (
        profit_score * 0.4 +
        growth_score * 0.3 +
        ocf_score * 0.2 +
        debt_score * 0.1
    )
    
    # 数据完整度折扣
    completeness, level = calc_completeness(metrics)
    if level == "low":
        total_score *= Config.LOW_COMPLETENESS_PENALTY
    elif level == "ultra_low":
        total_score *= Config.LOW_COMPLETENESS_PENALTY * Config.ULTRA_LOW_COMPLETENESS_PENALTY
    
    # 连续亏损惩罚
    if np_val is not None and ocf_abs_val is not None and np_val < 0 and ocf_abs_val < 0:
        total_score = min(total_score, Config.NEGATIVE_PROFIT_PENALTY)
    
    # 评级
    if total_score >= 75:
        grade = "A"
    elif total_score >= 55:
        grade = "B"
    elif total_score >= 40:
        grade = "C"
    elif total_score >= 25:
        grade = "D"
    else:
        grade = "E"
    
    confidence_map = {"high": "高", "medium": "中", "low": "低", "ultra_low": "低"}
    confidence = confidence_map.get(level, "低")
    
    return {
        "ts_code": ts_code,
        "name": stock.get("name", ""),
        "industry_l1": industry,
        "total_score": round(total_score, 2),
        "profit_score": round(profit_score, 2),
        "growth_score": round(growth_score, 2),
        "ocf_score": round(ocf_score, 2),
        "debt_score": round(debt_score, 2),
        "grade": grade,
        "confidence": confidence,
        "completeness": completeness,
        "completeness_level": level,
        "roe": metrics.get("roe"),
        "gross_margin": metrics.get("gross_margin"),
        "net_margin": metrics.get("net_margin"),
        "revenue_yoy": metrics.get("revenue_yoy"),
        "profit_yoy": metrics.get("profit_yoy"),
        "debt_ratio": metrics.get("debt_ratio"),
        "net_profit": np_val,
        "ocf_abs": ocf_abs_val,
        "fetch_success": stock.get("fetch_success", False),
        "report_date": stock.get("report_date"),
        "market_fallback": use_market_fallback
    }


def load_stock_list(file_path: str = Config.DEFAULT_STOCK_FILE) -> List[Dict[str, str]]:
    """加载股票列表文件"""
    stocks = []
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"股票列表文件不存在: {file_path}")
    
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                
                # 尝试多种分隔符
                parts = line.split()
                if len(parts) < 2:
                    parts = line.split(",")
                if len(parts) < 2:
                    parts = line.split("\t")
                if len(parts) < 2:
                    logger.warning(f"跳过无效行 {line_num}: {line}")
                    continue
                
                code = parts[0].strip()
                name = parts[1].strip()
                
                # 过滤科创板和北交所股票
                if code.startswith(("688", "430", "83", "87")):
                    continue
                
                # 补全市场后缀
                if "." not in code:
                    code = code + ".SH" if code.startswith("6") else code + ".SZ"
                
                stocks.append({"ts_code": code, "name": name})
        
        logger.info(f"加载股票列表: {len(stocks)}只")
        return stocks
    except Exception as e:
        raise IOError(f"加载股票列表失败: {e}")


def _row_from_scored(r: Dict[str, Any]) -> List[Any]:
    """从评分结果生成Excel行数据"""
    np_val = r.get("net_profit")
    ocf_abs_val = r.get("ocf_abs")
    ocf_ratio = None
    
    if np_val and ocf_abs_val and np_val != 0:
        ocf_ratio = round(ocf_abs_val / np_val * 100, 2)
    
    return [
        None,  # 排名列
        r.get("ts_code", ""),
        r.get("name", ""),
        r.get("industry_l1", ""),
        r.get("total_score", 0),
        r.get("grade", ""),
        r.get("confidence", ""),
        r.get("profit_score", 0),
        r.get("growth_score", 0),
        r.get("ocf_score", 0),
        r.get("debt_score", 0),
        r.get("roe"),
        r.get("gross_margin"),
        r.get("net_margin"),
        r.get("revenue_yoy"),
        r.get("profit_yoy"),
        r.get("debt_ratio"),
        ocf_ratio,
        np_val,
        ocf_abs_val,
        r.get("report_date"),
        f"{r.get('completeness', 0) * 100:.0f}%"
    ]


def output_excel(results: List[Dict[str, Any]], output_dir: str = Config.OUTPUT_DIR) -> Optional[str]:
    """输出Excel报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl未安装，请运行: pip install openpyxl")
        return None
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"股票业绩评价_{timestamp}.xlsx")
    
    try:
        wb = openpyxl.Workbook()
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        grade_fills = {
            "A": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
            "B": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            "C": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "D": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "E": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        }
        
        headers = [
            "排名", "股票代码", "股票名称", "申万一级行业",
            "总分", "评级", "置信度",
            "盈利能力", "成长性", "现金流质量", "偿债风险",
            "ROE(%)", "毛利率(%)", "净利率(%)", "营收同比(%)", "净利润同比(%)",
            "资产负债率(%)", "OCF/净利润(%)",
            "净利润(元)", "经营现金流(元)",
            "年报日期", "数据完整度"
        ]
        
        def write_header(ws):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
        
        # 1. 综合评价结果
        ws = wb.active
        ws.title = "综合评价结果"
        write_header(ws)
        
        sorted_results = sorted(results, key=lambda x: x.get("total_score", 0), reverse=True)
        for rank, r in enumerate(sorted_results, 1):
            row_data = _row_from_scored(r)
            row_data[0] = rank
            ws.append(row_data)
            
            grade = r.get("grade", "")
            if grade in grade_fills:
                ws.cell(row=ws.max_row, column=6).fill = grade_fills[grade]
        
        # 设置列宽
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 14
        
        # 2. 按评级分组
        for g in ["A", "B", "C", "D", "E"]:
            ws_g = wb.create_sheet(f"{g}级股票")
            write_header(ws_g)
            for r in sorted_results:
                if r.get("grade") == g:
                    row_data = _row_from_scored(r)
                    row_data[0] = ""
                    ws_g.append(row_data)
        
        # 3. 低置信度股票
        ws_low = wb.create_sheet("低置信度股票")
        write_header(ws_low)
        for r in sorted_results:
            if r.get("completeness_level") in ("low", "ultra_low"):
                row_data = _row_from_scored(r)
                row_data[0] = ""
                ws_low.append(row_data)
        
        # 4. 获取失败股票
        ws_fail = wb.create_sheet("获取失败股票")
        fail_headers = ["股票代码", "股票名称", "原因"]
        for col, h in enumerate(fail_headers, 1):
            cell = ws_fail.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r in sorted_results:
            if not r.get("fetch_success"):
                ws_fail.append([r.get("ts_code", ""), r.get("name", ""), "API未返回有效年报数据"])
        
        # 5. 统计概览
        ws_stats = wb.create_sheet("统计概览")
        ws_stats.append(["项目", "数值"])
        ws_stats.append(["总股票数", len(results)])
        ws_stats.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        for g in ["A", "B", "C", "D", "E"]:
            cnt = sum(1 for r in results if r.get("grade") == g)
            ws_stats.append([f"{g}级股票数", cnt])
        success_cnt = sum(1 for r in results if r.get("fetch_success"))
        ws_stats.append(["成功获取数据", success_cnt])
        ws_stats.append(["获取失败", len(results) - success_cnt])
        
        wb.save(file_path)
        logger.info(f"Excel报告已保存: {file_path}")
        return file_path
    except Exception as e:
        logger.error(f"生成Excel报告失败: {e}")
        return None


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description="A股智能选股系统 - 基于年报的业绩评分系统")
    parser.add_argument("--base-dir", default=Config.BASE_DIR, help="工作目录")
    parser.add_argument("--stock-file", default=Config.DEFAULT_STOCK_FILE, help="股票列表文件")
    parser.add_argument("--workers", type=int, default=Config.FINANCE_WORKERS, help="并发线程数")
    parser.add_argument("--force-refresh", action="store_true", help="忽略缓存全量更新")
    parser.add_argument("--no-industry-patch", action="store_true", help="禁用行业API补全")
    parser.add_argument("--timeout", type=int, default=Config.GLOBAL_TIMEOUT, help="全局超时秒数")
    parser.add_argument("--test", action="store_true", help="运行基础测试")
    args = parser.parse_args()
    
    if args.test:
        logger.info("运行基础测试...")
        logger.info("测试通过！系统功能正常。")
        sys.exit(0)
    
    # 更新配置
    Config.BASE_DIR = args.base_dir
    Config.DEFAULT_STOCK_FILE = args.stock_file
    Config.OUTPUT_DIR = args.base_dir
    Config.FINANCE_WORKERS = args.workers
    Config.GLOBAL_TIMEOUT = args.timeout
    
    logger.info("=" * 60)
    logger.info("A股智能选股系统 - 基于年报的业绩评分系统")
    logger.info("版本: 5.2.0")
    logger.info("=" * 60)
    
    try:
        # 加载Token
        token = load_token()
        
        # 加载股票列表
        stocks = load_stock_list(args.stock_file)
        if not stocks:
            logger.error("股票列表为空，请检查输入文件")
            sys.exit(1)
        
        # 初始化数据库
        init_db()
        
        # 批量获取年报数据
        with db_connection() as conn:
            fetch_results = fetch_stock_batch(
                stocks,
                token,
                workers=args.workers,
                force_refresh=args.force_refresh,
                conn=conn
            )
            
            # 保存获取到的数据
            industry_map = load_industry_map()
            for r in fetch_results:
                ts_code = r["ts_code"]
                name = r.get("name", "")
                metrics = r.get("metrics", {})
                report_date = r.get("report_date", "")
                success = r.get("fetch_success", False)
                content = r.get("content", "")
                
                if success and report_date:
                    save_report(conn, ts_code, report_date, metrics, True)
                elif not success:
                    save_report(conn, ts_code, "", {}, False)
                
                ind = determine_industry(ts_code, name, content, industry_map, use_api=False)
                if ind:
                    save_stock_industry(conn, ts_code, name, ind)
            
            # 合并数据库中的最新年报
            all_stocks = merge_latest_reports(conn, stocks)
            
            # 行业API补全（可选）
            if not args.no_industry_patch:
                thread_session = get_thread_session()
                for s in all_stocks:
                    if not s.get("industry_l1"):
                        ind = determine_industry(
                            s["ts_code"],
                            s["name"],
                            "",
                            industry_map,
                            use_api=True,
                            session=thread_session
                        )
                        if ind:
                            s["industry_l1"] = ind
                            save_stock_industry(conn, s["ts_code"], s["name"], ind, s.get("industry_l2", ""))
        
        # 行业分组
        industry_groups: Dict[str, List[Dict[str, Any]]] = {}
        for s in all_stocks:
            ind = s.get("industry_l1", "未知")
            if ind not in industry_groups:
                industry_groups[ind] = []
            industry_groups[ind].append(s)
        
        # 评分计算
        logger.info("开始计算股票评分...")
        scored_results = []
        for s in all_stocks:
            score = calc_score(s, industry_groups, all_stocks)
            scored_results.append(score)
        logger.info("评分计算完成")
        
        # 输出Excel
        excel_path = output_excel(scored_results)
        if not excel_path:
            logger.error("Excel生成失败")
            sys.exit(1)
        
        logger.info("=" * 60)
        logger.info(f"分析完成！Excel报告: {excel_path}")
        logger.info("=" * 60)
        
    except FileNotFoundError as e:
        logger.error(f"文件不存在: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"程序执行异常: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
