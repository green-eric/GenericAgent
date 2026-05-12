#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成2025年报净利润增长率Excel"""
import json, sys, io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 读取数据
with open("growth_2025_result.json", "r", encoding="utf-8") as f:
    data = json.load(f)

high_growth = data["high_growth_over50"]
all_results = data["all"]

wb = Workbook()

# ===== Sheet1: 增长率>50% =====
ws1 = wb.active
ws1.title = "2025年报增长超50%"

# 样式
header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # 红色
red_font = Font(name="微软雅黑", bold=True, size=11, color="FF0000")
normal_font = Font(name="微软雅黑", size=11)
title_font = Font(name="微软雅黑", bold=True, size=14, color="C00000")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')

# 标题
ws1.merge_cells('A1:G1')
ws1['A1'] = "新征程853股票池 - 2025年报归母净利润增长率 > 50%（排除科创板）"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

# 表头
headers = ["排名", "股票名称", "股票代码", "2025归母净利润(亿元)", "2024归母净利润(亿元)", "增长率(%)", "备注"]
col_widths = [8, 14, 14, 22, 22, 14, 20]
for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(col_idx)].width = w

# 数据
for i, r in enumerate(high_growth):
    row = 4 + i
    v25 = round(r["net2025"] / 1e8, 2) if r["net2025"] else 0
    v24 = round(r["net2024"] / 1e8, 2) if r["net2024"] else 0
    yoy = r["yoy"]
    note = ""
    if v24 < 0:
        note = "2024年亏损，扭亏"
    
    data_row = [i+1, r["name"], r["ts_code"], v25, v24, yoy, note]
    for col_idx, val in enumerate(data_row, 1):
        cell = ws1.cell(row=row, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if col_idx in (1, 6):
            cell.alignment = center_align
        elif col_idx in (4, 5):
            cell.alignment = right_align
            cell.number_format = '#,##0.00'
        elif col_idx == 6:
            cell.number_format = '0.0'
        # 增长率>100%标红
        if col_idx == 6 and yoy > 100:
            cell.font = red_font

# ===== Sheet2: 全部股票增长率 =====
ws2 = wb.create_sheet("全部股票增长率一览")

ws2.merge_cells('A1:G1')
ws2['A1'] = "新征程853股票池 - 2025年报归母净利润增长率（排除科创板，共79只）"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

headers2 = ["排名", "股票名称", "股票代码", "2025归母净利润(亿元)", "2024归母净利润(亿元)", "增长率(%)", "是否>50%"]
col_widths2 = [8, 14, 14, 22, 22, 14, 12]
for col_idx, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

# 排序：有增长率的按增长率降序，无增长率的在后面
valid = [r for r in all_results if r["yoy"] is not None]
invalid = [r for r in all_results if r["yoy"] is None]
valid.sort(key=lambda x: x["yoy"], reverse=True)
sorted_all = valid + invalid

light_red = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
light_green = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
light_gray = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

for i, r in enumerate(sorted_all):
    row = 4 + i
    v25 = round(r["net2025"] / 1e8, 2) if r.get("net2025") else None
    v24 = round(r["net2024"] / 1e8, 2) if r.get("net2024") else None
    yoy = r.get("yoy")
    gt50 = "是" if (yoy is not None and yoy > 50) else "否"
    
    data_row = [i+1, r["name"], r["ts_code"], v25, v24, yoy, gt50]
    for col_idx, val in enumerate(data_row, 1):
        cell = ws2.cell(row=row, column=col_idx, value=val if val is not None else "N/A")
        cell.font = normal_font
        cell.border = thin_border
        if col_idx in (1, 6, 7):
            cell.alignment = center_align
        elif col_idx in (4, 5):
            cell.alignment = right_align
            if val is not None:
                cell.number_format = '#,##0.00'
        if col_idx == 6 and yoy is not None:
            cell.number_format = '0.0'
    
    # 行背景色：>50%浅红，<0浅绿，N/A灰
    if yoy is not None and yoy > 50:
        for col_idx in range(1, 8):
            ws2.cell(row=row, column=col_idx).fill = light_red
    elif yoy is not None and yoy < 0:
        for col_idx in range(1, 8):
            ws2.cell(row=row, column=col_idx).fill = light_green
    elif yoy is None:
        for col_idx in range(1, 8):
            ws2.cell(row=row, column=col_idx).fill = light_gray

# 保存
desktop = "C:\\Users\\green\\Desktop"
filepath = f"{desktop}\\新征程853_2025年报净利润增长率_排除科创板.xlsx"
wb.save(filepath)
print(f"Excel已保存到: {filepath}")
