#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl, io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r'c:\Users\green\WorkBuddy\20260424203734\workplace\股票业绩评价_20260426_000327.xlsx', read_only=True)

# 异常日志
print("=" * 70)
print("【异常日志 sheet】")
print("=" * 70)
ws = wb['异常日志']
rows = list(ws.iter_rows(values_only=True))
print(f"总行数: {len(rows)}")
print(f"表头: {rows[0]}")
print("\n全部错误记录:")
for r in rows[1:]:
    print(f"  {r}")

# 获取失败股票
print(f"\n{'=' * 70}")
print("【获取失败股票 sheet】")
print("=" * 70)
ws2 = wb['获取失败股票']
rows2 = list(ws2.iter_rows(values_only=True))
print(f"总行数: {len(rows2)}")
print(f"表头: {rows2[0]}")
print("\n前20条:")
for r in rows2[1:21]:
    print(f"  {r}")

# 统计错误类型
print(f"\n{'=' * 70}")
print("【错误类型统计】")
print("=" * 70)
from collections import Counter
error_types = Counter()
for r in rows[1:]:
    if r and len(r) > 3:
        err_detail = str(r[3])  # 错误详情列
        # 分类
        if 'API错误码' in err_detail:
            # 提取具体错误码
            code = err_detail.replace('API错误码 ', '').strip()
            error_types[f'API错误码: {code}'] += 1
        elif '所有查询均失败' in err_detail:
            error_types['行业API: 所有查询均失败'] += 1
        elif '空文本' in err_detail:
            error_types['API返回空文本'] += 1
        elif 'no_token' in err_detail:
            error_types['Token缺失'] += 1
        elif 'exhausted' in err_detail:
            error_types['重试耗尽(exhausted)'] += 1
        else:
            error_types[err_detail[:50]] += 1

for err, cnt in error_types.most_common():
    print(f"  {cnt:>5d} 次 | {err}")

print(f"\n总计: {sum(error_types.values())} 条错误记录")
